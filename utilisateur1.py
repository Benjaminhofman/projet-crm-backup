import os
import sys
import re
import csv
import sqlite3
import pandas as pd
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from ttkthemes import ThemedTk
import threading
import time
import logging
import unicodedata
import webbrowser
import openpyxl
from openpyxl.styles import PatternFill
from decimal import Decimal, getcontext
import subprocess
import datetime  # Ajoutez cette ligne

class FECExplorer:
    def __init__(self):
        self.root = ThemedTk(theme="arc")
        self.root.title("FEC Explorer")
        self.root.geometry("800x600")
        self.root.configure(bg='#f5f6fa')

        self.base_path = self.get_base_path()
        self.db_path = self.get_resource_path('fusionfec.db')
        self.log_path = self.get_resource_path('fec_explorer.log')
        self.reports_dir = self.get_resource_path('rapports')

        # Configuration du logging
        logging.basicConfig(filename=self.log_path, level=logging.DEBUG,
                            format='%(asctime)s - %(levelname)s - %(message)s')

        # Création des styles pour l'interface
        self.setup_styles()
        
        # Mise en place de l'interface avec onglets
        self.setup_notebook_ui()

    def get_base_path(self):
        if getattr(sys, 'frozen', False):
            return os.path.dirname(sys.executable)
        else:
            return os.path.dirname(os.path.abspath(__file__))

    def get_resource_path(self, relative_path):
        return os.path.join(self.base_path, relative_path)

    def setup_styles(self):
        self.style = ttk.Style()
        self.style.configure("TFrame", background='#f5f6fa')
        self.style.configure("TLabel", background='#f5f6fa', foreground="#1e2022")
        self.style.configure("TButton", background="#3282b8", foreground="black", font=("Helvetica", 10, "bold"))
        self.style.configure("Accent.TButton", background="#0f4c75", foreground="black", padding=(20, 10), font=("Helvetica", 12, "bold"))
        
        # Style pour les onglets - taille augmentée et texte rouge quand actif
        self.style.configure("TNotebook", background='#f5f6fa', tabmargin=0)
        self.style.configure("TNotebook.Tab", 
                             background="#e0e0e0", 
                             foreground="black", 
                             padding=(20, 10),  # Augmenter la taille des onglets
                             font=("Helvetica", 12, "bold"))  # Police plus grande
        
        # Style pour l'onglet actif - fond plus contrasté et texte en rouge
        self.style.map("TNotebook.Tab", 
                       background=[("selected", "#3282b8")],
                       foreground=[("selected", "#FF0000")],  # Texte en rouge pour l'onglet actif
                       font=[("selected", ("Helvetica", 12, "bold"))])

    def setup_notebook_ui(self):
        """Configure l'interface avec onglets"""
        # Création du cadre de titre avec dégradé
        self.title_frame = GradientFrame(self.root, height=60, color1="#1E88E5", color2="#1DE9B6")
        self.title_frame.pack(fill=tk.X)
        
        self.title_frame.create_text(20, 30, anchor="w", text="FEC Explorer", 
                                    font=('Arial', 24, 'bold'), fill="white")
        
        # Création du notebook (onglets) avec plus d'espace au-dessus
        self.notebook = ttk.Notebook(self.root, padding=(0, 5, 0, 0))  # Ajouter un padding en haut
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # Onglet principal
        self.main_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.main_tab, text="Accueil")
        
        # Configuration de l'interface pour chaque onglet
        self.setup_main_tab_ui()
        
        # Signature en bas de l'interface
        self.setup_signature()
   
    def setup_crm_tab_ui(self):
        """Configure l'interface de l'onglet CRM"""
        crm_frame = ttk.Frame(self.crm_tab)
        crm_frame.pack(fill=tk.BOTH, expand=True)
        
        self.crm_canvas = tk.Canvas(crm_frame, bg='#f5f6fa', highlightthickness=0)
        self.crm_scrollbar = ttk.Scrollbar(crm_frame, orient="vertical", command=self.crm_canvas.yview)
        self.crm_scrollable_frame = ttk.Frame(self.crm_canvas)

        self.crm_canvas.configure(yscrollcommand=self.crm_scrollbar.set)
        self.crm_canvas.pack(side="left", fill="both", expand=True)
        self.crm_scrollbar.pack(side="right", fill="y")

        self.crm_canvas_frame = self.crm_canvas.create_window(
            (0, 0), window=self.crm_scrollable_frame, anchor="nw", width=780
        )

        # Section pour la gestion des clients CRM
        self.create_section("Gestion des données CRM", [
            ("👥 Saisie des données CRM", self.afficher_liste_clients_crm),
            ("🔍 Rechercher un client", self.rechercher_client_crm),
            ("📝 Modifier les missions", self.modifier_missions_crm),
            ("📊 Tableau de bord CRM", self.afficher_tableau_bord_crm),
            ("📤 Exporter données CRM", self.exporter_donnees_crm)
        ], self.crm_scrollable_frame)
        
        # Section pour les rapports et analyses CRM
        self.create_section("Rapports et analyses", [
            ("📈 Statistiques des missions", self.statistiques_missions_crm),
            ("🎯 Opportunités commerciales", self.opportunites_commerciales_crm),
            ("📆 Planning de suivi", self.planning_suivi_crm)
        ], self.crm_scrollable_frame)

        # Configuration pour la gestion du défilement
        self.crm_scrollable_frame.bind("<Configure>", 
            lambda e: self.crm_canvas.configure(scrollregion=self.crm_canvas.bbox("all"))
        )
        self.crm_canvas.bind("<Configure>", 
            lambda e: self.crm_canvas.itemconfig(self.crm_canvas_frame, width=e.width-4)
        )
        
        # Étiquette de statut pour l'onglet CRM
        self.crm_status_label = tk.Label(
            self.crm_scrollable_frame, 
            text="Prêt à gérer les relations clients", 
            font=("Helvetica", 10), 
            bg="#f5f6fa"
        )
        self.crm_status_label.pack(pady=10)

    def afficher_liste_clients_crm(self):
        """Affiche la liste des clients dans la table CRM avec possibilité de modifier l'état des missions"""
        try:
            # Connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Vérifier si la table CRM existe, sinon la créer
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS crm (
                    siret TEXT PRIMARY KEY,
                    nom TEXT,
                    code_naf TEXT,
                    ville TEXT,
                    mission_patrimoniale TEXT,
                    mission_optimisation TEXT,
                    mission_retraite TEXT,
                    mission_juridique TEXT,
                    date_maj TEXT,
                    utilisateur TEXT,
                    commentaire TEXT
                )
            """)
            conn.commit()
            
            # Récupérer les noms des colonnes de la table CRM
            cursor.execute("PRAGMA table_info(crm)")
            column_info = cursor.fetchall()
            db_columns = [col[1] for col in column_info]
            
            # Si la table est vide, proposer de la remplir à partir des clients existants
            cursor.execute("SELECT COUNT(*) FROM crm")
            count = cursor.fetchone()[0]
            
            if count == 0:
                response = messagebox.askyesno(
                    "Base CRM vide",
                    "La base de données CRM est vide. Voulez-vous l'initialiser avec les clients existants ?"
                )
                
                if response:
                    # Récupérer les clients de la table indicateurs
                    cursor.execute("""
                        SELECT siret, nom, code_naf, ville
                        FROM indicateurs
                        WHERE siret != 'moyenne' AND siret IS NOT NULL
                    """)
                    clients = cursor.fetchall()
                    
                    # Insérer les clients dans la table CRM
                    for client in clients:
                        siret, nom, code_naf, ville = client
                        cursor.execute("""
                            INSERT INTO crm (siret, nom, code_naf, ville, 
                                         mission_patrimoniale, mission_optimisation, 
                                         mission_retraite, mission_juridique, 
                                         date_maj, utilisateur)
                            VALUES (?, ?, ?, ?, 'non_detectee', 'non_detectee', 'non_detectee', 'non_detectee', ?, ?)
                        """, (siret, nom, code_naf, ville, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "système"))
                    conn.commit()
                    messagebox.showinfo("Succès", f"{len(clients)} clients ont été ajoutés à la base CRM.")
            
            # Création de la fenêtre d'affichage CRM
            crm_window = tk.Toplevel(self.root)
            crm_window.title("Gestion des données CRM")
            crm_window.geometry(f"{self.root.winfo_screenwidth()}x{self.root.winfo_screenheight()}")
            
            # Cadre principal avec défilement
            main_frame = ttk.Frame(crm_window)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Barre d'outils
            toolbar_frame = ttk.Frame(main_frame)
            toolbar_frame.pack(fill=tk.X, pady=5)
            
            # Champ de recherche
            search_frame = ttk.Frame(toolbar_frame)
            search_frame.pack(side=tk.LEFT, padx=10)
            
            ttk.Label(search_frame, text="Rechercher:").pack(side=tk.LEFT, padx=5)
            search_entry = ttk.Entry(search_frame, width=30)
            search_entry.pack(side=tk.LEFT, padx=5)
            
            self.modified_cells = set()  # Pour suivre les cellules modifiées
            
            # Fonction de recherche
            def search_clients():
                search_text = search_entry.get().strip().lower()
                for item in tree.get_children():
                    tree.delete(item)
                
                query = f"SELECT {', '.join(db_columns)} FROM crm WHERE 1=1"
                params = []
                
                if search_text:
                    query += " AND (LOWER(nom) LIKE ? OR LOWER(siret) LIKE ? OR LOWER(ville) LIKE ?)"
                    params.extend([f'%{search_text}%', f'%{search_text}%', f'%{search_text}%'])
                
                query += " ORDER BY nom"
                
                cursor.execute(query, params)
                results = cursor.fetchall()
                
                for row in results:
                    tree.insert("", "end", values=row, tags=(str(row[0]),))
                
                status_label.config(text=f"{len(results)} client(s) trouvé(s)")
                
                # Appliquer les couleurs aux cellules
                color_cells()
            
            ttk.Button(search_frame, text="Rechercher", command=search_clients).pack(side=tk.LEFT, padx=5)
            
            # Bouton pour actualiser la liste
            ttk.Button(toolbar_frame, text="Actualiser", command=search_clients).pack(side=tk.LEFT, padx=10)
            
            # Bouton pour enregistrer les modifications
            save_button = ttk.Button(
                toolbar_frame, 
                text="Enregistrer les modifications", 
                command=lambda: save_changes())
            save_button.pack(side=tk.RIGHT, padx=10)
            
            # Cadre pour le tableau avec défilement
            table_frame = ttk.Frame(main_frame)
            table_frame.pack(fill=tk.BOTH, expand=True, pady=10)
            
            # Barre de défilement
            scrollbar_y = ttk.Scrollbar(table_frame, orient=tk.VERTICAL)
            scrollbar_x = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL)
            
            # Conversion des noms de colonnes pour l'affichage (première lettre en majuscule)
            display_columns = [col.capitalize().replace('_', ' ') for col in db_columns]
            
            # Création du tableau (Treeview)
            tree = ttk.Treeview(
                table_frame, 
                columns=db_columns, 
                show='headings', 
                yscrollcommand=scrollbar_y.set, 
                xscrollcommand=scrollbar_x.set,
                selectmode='browse'
            )
            
            # Configurer le style pour le tableau
            style = ttk.Style()
            style.configure("Treeview", 
                            foreground="black", 
                            rowheight=25)
            style.configure("Treeview.Heading", 
                            foreground="black", 
                            font=('Helvetica', 10, 'bold'))
            
            # Configurer les colonnes
            mission_columns = ['mission_patrimoniale', 'mission_optimisation', 'mission_retraite', 'mission_juridique']
            
            for i, col in enumerate(db_columns):
                tree.heading(col, text=display_columns[i], command=lambda c=col: sort_tree(tree, c, False))
                
                # Largeurs spécifiques selon le type de colonne
                if col == 'siret':
                    width = 120
                elif col == 'nom':
                    width = 200
                elif col == 'commentaire':
                    width = 200
                elif col in mission_columns:
                    width = 150
                else:
                    width = 100
                    
                tree.column(col, width=width, anchor='center')
            
            # Configurer les barres de défilement
            scrollbar_y.config(command=tree.yview)
            scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
            
            scrollbar_x.config(command=tree.xview)
            scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
            
            tree.pack(fill=tk.BOTH, expand=True)
            
            # Barre d'état
            status_frame = ttk.Frame(main_frame)
            status_frame.pack(fill=tk.X, pady=5)
            
            status_label = ttk.Label(status_frame, text="Prêt")
            status_label.pack(side=tk.LEFT, padx=10)
            
            # Charger les données initiales
            query = f"SELECT {', '.join(db_columns)} FROM crm ORDER BY nom"
            cursor.execute(query)
            
            results = cursor.fetchall()
            for row in results:
                tree.insert("", "end", values=row, tags=(str(row[0]),))
            
            status_label.config(text=f"{len(results)} client(s) trouvé(s)")
            
            # Définir les états possibles pour les missions
            mission_states = ["non_detectee", "detectee", "proposee", "signee", "refusee"]
            mission_colors = {
                "non_detectee": "#FFFFFF",  # blanc
                "detectee": "#FFF9C4",      # jaune clair
                "proposee": "#BBDEFB",      # bleu clair
                "signee": "#C8E6C9",        # vert clair
                "refusee": "#FFCDD2"        # rouge clair
            }
            
            # Indices des colonnes de missions dans db_columns
            mission_indices = [db_columns.index(col) for col in mission_columns if col in db_columns]
            
            # Configurer à l'avance tous les styles de tag pour les états
            for state, color in mission_colors.items():
                for col_name in mission_columns:
                    if col_name in db_columns:
                        col_idx = db_columns.index(col_name)
                        tag_name = f"state_{col_idx}_{state}"
                        tree.tag_configure(tag_name, background=color)
            
            # Fonction pour appliquer les couleurs aux cellules selon leurs états
            def color_cells():
                for item in tree.get_children():
                    values = tree.item(item, 'values')
                    for i in mission_indices:
                        if i < len(values):
                            state = values[i]
                            if state in mission_colors:
                                tag_name = f"state_{i}_{state}"
                                item_tags = list(tree.item(item, 'tags'))
                                
                                # Enlever les anciens tags de cette colonne
                                item_tags = [tag for tag in item_tags if not tag.startswith(f"state_{i}_")]
                                
                                # Ajouter le nouveau tag
                                item_tags.append(tag_name)
                                tree.item(item, tags=item_tags)
            
            # Appliquer les couleurs initiales
            color_cells()
            
            # Fonction pour trier le tableau
            def sort_tree(tree, col, reverse):
                # Récupérer les données
                data = [(tree.set(item, col), item) for item in tree.get_children('')]
                
                # Définir une fonction de comparaison
                def convert_to_number(val):
                    try:
                        return float(val)
                    except (ValueError, TypeError):
                        return val
                
                # Trier les données
                data.sort(key=lambda x: convert_to_number(x[0]), reverse=reverse)
                
                # Réorganiser les items
                for index, (_, item) in enumerate(data):
                    tree.move(item, '', index)
                
                # Configurer le prochain clic pour trier dans l'ordre inverse
                tree.heading(col, command=lambda: sort_tree(tree, col, not reverse))
            
            # Fonction pour gérer le clic sur une cellule de mission
            def on_cell_click(event):
                # Identifier la région où le clic a eu lieu
                region = tree.identify_region(event.x, event.y)
                if region != "cell":
                    return
                
                # Obtenir l'ID de la ligne et la colonne
                row_id = tree.identify_row(event.y)
                if not row_id:
                    return
                    
                col_id = tree.identify_column(event.x)
                if not col_id:
                    return
                
                # Convertir l'ID de colonne en index (de 1 à n)
                col_index = int(col_id.replace('#', '')) - 1
                
                # Vérifier si l'index est valide
                if col_index < 0 or col_index >= len(db_columns):
                    return
                    
                col_name = db_columns[col_index]
                
                # Vérifier si c'est une colonne de mission
                if col_name in mission_columns:
                    # Obtenir les valeurs actuelles
                    values = list(tree.item(row_id, 'values'))
                    if not values or len(values) <= col_index:
                        return
                        
                    client_siret = values[db_columns.index('siret')]
                    
                    # Obtenir l'état actuel de la mission
                    current_state = values[col_index]
                    
                    # Trouver le prochain état
                    if current_state in mission_states:
                        next_index = (mission_states.index(current_state) + 1) % len(mission_states)
                        next_state = mission_states[next_index]
                    else:
                        next_state = "detectee"  # État par défaut si l'état actuel est inconnu
                    
                    # Mettre à jour la valeur dans le tableau
                    values[col_index] = next_state
                    tree.item(row_id, values=values)
                    
                    # Appliquer la couleur
                    tag_name = f"state_{col_index}_{next_state}"
                    current_tags = list(tree.item(row_id, 'tags'))
                    
                    # Enlever les anciens tags de cette colonne
                    current_tags = [tag for tag in current_tags if not tag.startswith(f"state_{col_index}_")]
                    
                    # Ajouter le nouveau tag
                    current_tags.append(tag_name)
                    tree.item(row_id, tags=current_tags)
                    
                    # Ajouter à la liste des modifications
                    self.modified_cells.add((client_siret, col_name, next_state))
                    
                    # Mettre à jour le libellé de statut
                    status_label.config(text=f"Modification en attente d'enregistrement: {client_siret}, {col_name} -> {next_state}")
            
            # Associer l'événement de clic
            tree.bind("<Button-1>", on_cell_click)
            
            # Fonction pour enregistrer les modifications
            def save_changes():
                if not self.modified_cells:
                    messagebox.showinfo("Information", "Aucune modification à enregistrer.")
                    return
                
                try:
                    # Préparer les mises à jour
                    for client_siret, column_name, new_state in self.modified_cells:
                        # Mise à jour dans la base de données
                        cursor.execute(f"""
                            UPDATE crm
                            SET {column_name} = ?,
                                date_maj = ?,
                                utilisateur = ?
                            WHERE siret = ?
                        """, (new_state, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
                              "utilisateur", client_siret))
                    
                    conn.commit()
                    self.modified_cells.clear()
                    
                    messagebox.showinfo("Succès", "Les modifications ont été enregistrées avec succès.")
                    
                    # Actualiser l'affichage
                    search_clients()
                    
                except Exception as e:
                    messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'enregistrement: {str(e)}")
            
            # Fonction à exécuter à la fermeture de la fenêtre
            def on_close():
                try:
                    # Demander si l'utilisateur veut enregistrer les modifications
                    if self.modified_cells:
                        response = messagebox.askyesnocancel(
                            "Enregistrer les modifications", 
                            "Des modifications n'ont pas été enregistrées. Voulez-vous les enregistrer avant de fermer?"
                        )
                        
                        if response is None:  # Annuler la fermeture
                            return
                        elif response:  # Oui, enregistrer
                            save_changes()
                    
                    # Fermer la connexion à la base de données et la fenêtre
                    conn.close()
                    crm_window.destroy()
                except Exception as e:
                    print(f"Erreur lors de la fermeture: {str(e)}")
                    try:
                        conn.close()
                    except:
                        pass
                    crm_window.destroy()
            
            # Associer l'événement de fermeture
            crm_window.protocol("WM_DELETE_WINDOW", on_close)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue: {str(e)}")
            import traceback
            traceback.print_exc()
            if 'conn' in locals():
                try:
                    conn.close()
                except:
                    pass

    def rechercher_client_crm(self):
        """Recherche un client dans la table CRM"""
        messagebox.showinfo("Information", "Fonctionnalité à implémenter.")

    def modifier_missions_crm(self):
        """Interface pour modifier les missions d'un client"""
        messagebox.showinfo("Information", "Fonctionnalité à implémenter.")

    def afficher_tableau_bord_crm(self):
        """Affiche le tableau de bord CRM"""
        messagebox.showinfo("Information", "Fonctionnalité à implémenter.")

    def exporter_donnees_crm(self):
        """Exporte les données CRM"""
        messagebox.showinfo("Information", "Fonctionnalité à implémenter.")

    def statistiques_missions_crm(self):
        """Affiche les statistiques des missions"""
        messagebox.showinfo("Information", "Fonctionnalité à implémenter.")

    def opportunites_commerciales_crm(self):
        """Affiche les opportunités commerciales"""
        messagebox.showinfo("Information", "Fonctionnalité à implémenter.")

    def planning_suivi_crm(self):
        """Affiche le planning de suivi"""
        messagebox.showinfo("Information", "Fonctionnalité à implémenter.")

    def modifier_client_crm(self, tree):
        """Modifie les informations d'un client sélectionné"""
        messagebox.showinfo("Information", "Fonctionnalité à implémenter.")

    def afficher_resultats_recherche_crm(self, results):
        """Affiche les résultats de la recherche CRM"""
        messagebox.showinfo("Information", "Fonctionnalité à implémenter.")


    def setup_reco_tab_ui(self):
        """Configure l'interface de l'onglet Recommandations"""
        reco_frame = ttk.Frame(self.reco_tab)
        reco_frame.pack(fill=tk.BOTH, expand=True)
        
        self.reco_canvas = tk.Canvas(reco_frame, bg='#f5f6fa', highlightthickness=0)
        self.reco_scrollbar = ttk.Scrollbar(reco_frame, orient="vertical", command=self.reco_canvas.yview)
        self.reco_scrollable_frame = ttk.Frame(self.reco_canvas)

        self.reco_canvas.configure(yscrollcommand=self.reco_scrollbar.set)
        self.reco_canvas.pack(side="left", fill="both", expand=True)
        self.reco_scrollbar.pack(side="right", fill="y")

        self.reco_canvas_frame = self.reco_canvas.create_window(
            (0, 0), window=self.reco_scrollable_frame, anchor="nw", width=780
        )

        # Section pour les recommandations globales
        self.create_section("Recommandations", [
            ("🔍 Détection automatique des opportunités", self.detecter_opportunites),
            ("📊 Afficher les dossiers avec recommandations", self.afficher_dossiers_avec_recommandations),
            ("📝 Générer rapport complet des recommandations", self.generer_rapport_recommandations)
        ], self.reco_scrollable_frame)
        
        # Section pour l'analyse individuelle
        self.create_section("Analyse individuelle", [
            ("👤 Sélectionner un dossier pour analyse détaillée", self.analyser_dossier_individuel)
        ], self.reco_scrollable_frame)

        # Configuration similaire aux autres onglets pour la gestion du défilement
        self.reco_scrollable_frame.bind("<Configure>", 
            lambda e: self.reco_canvas.configure(scrollregion=self.reco_canvas.bbox("all"))
        )
        self.reco_canvas.bind("<Configure>", 
            lambda e: self.reco_canvas.itemconfig(self.reco_canvas_frame, width=e.width-4)
        )
        
        # Étiquette de statut pour l'onglet Recommandations
        self.reco_status_label = tk.Label(
            self.reco_scrollable_frame, 
            text="Prêt à générer des recommandations", 
            font=("Helvetica", 10), 
            bg="#f5f6fa"
        )
        self.reco_status_label.pack(pady=10)

    def detecter_opportunites(self):
        """Détecte les opportunités pour chaque dossier selon des critères spécifiques"""
        try:
            # Connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Vérifier si la table 'recommandations' existe, sinon la créer
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS recommandations (
                    siret TEXT PRIMARY KEY,
                    opportunites TEXT,
                    details TEXT,
                    date_detection TEXT
                )
            """)
            conn.commit()
            
            # Effacer les recommandations existantes
            cursor.execute("DELETE FROM recommandations")
            conn.commit()
            
            # Récupérer tous les dossiers
            cursor.execute("""
                SELECT siret, nom, forme_juri, Imposition, emprunt, code_naf, prestation, 
                       CA, Regime_de_TVA, impot_revenu, resultat, m_salariale, trésorerie,
                       compte_70, compte_60
                FROM indicateurs
                WHERE siret != 'moyenne'
            """)
            dossiers = cursor.fetchall()
            
            # Mise à jour de la barre de progression
            total_dossiers = len(dossiers)
            self.progress_bar['value'] = 0
            self.progress_label.config(text=f"Analyse des dossiers: 0/{total_dossiers}")
            self.root.update_idletasks()
            
            # Nombre de dossiers avec des recommandations
            nb_dossiers_avec_reco = 0
            total_recommandations = 0
            
            # Analyser chaque dossier
            for idx, dossier in enumerate(dossiers):
                siret, nom, forme_juri, imposition, emprunt, code_naf, prestation, ca, regime_tva, impot_revenu, resultat, m_salariale, tresorerie, compte_70, compte_60 = dossier
                
                # Liste pour stocker les opportunités détectées
                opportunites = []
                details = []
                
                # Convertir les valeurs None en valeurs par défaut
                forme_juri = forme_juri or ""
                imposition = imposition or ""
                emprunt = float(emprunt or 0)
                code_naf = code_naf or ""
                prestation = prestation or ""
                ca = float(ca or 0)
                regime_tva = regime_tva or ""
                impot_revenu = float(impot_revenu or 0)
                resultat = float(resultat or 0)
                m_salariale = float(m_salariale or 0)
                tresorerie = float(tresorerie or 0)
                compte_70 = float(compte_70 or 0)
                compte_60 = float(compte_60 or 0)
                
                # 1. SCI à l'IR avec emprunt significatif
                if forme_juri == "SCI" and imposition == "IR" and emprunt < -30000:
                    opportunites.append("Étudier passage IS")
                    details.append(f"Forme juridique: SCI, Imposition actuelle: IR, Montant emprunt: {emprunt:.2f}€. Un passage à l'IS pourrait être avantageux pour déduire les intérêts d'emprunt.")
                
                # 2. Exonération CFE presse
                if code_naf == "4726Z" or code_naf == "4762Z":
                    opportunites.append("Étudier exonération CFE presse")
                    details.append(f"Code NAF: {code_naf}. Ce code est lié au secteur de la presse et peut permettre une exonération de CFE.")
                
                # 3. Franchise de TVA pour activités de vente
                if prestation != "presta" and ca < 85000 and forme_juri != "SCI" and ca != 0 and regime_tva != "NI":
                    opportunites.append("Étudier passage en franchise de TVA")
                    details.append(f"CA: {ca:.2f}€, Activité de vente (non prestation), Régime TVA actuel: {regime_tva}. Un CA < 85 000€ permet la franchise de TVA.")
                
                # 4. Franchise de TVA pour activités de prestation
                if prestation == "presta" and ca < 35000 and forme_juri != "SCI" and ca != 0 and regime_tva != "NI":
                    opportunites.append("Étudier passage en franchise de TVA")
                    details.append(f"CA: {ca:.2f}€, Activité de prestation, Régime TVA actuel: {regime_tva}. Un CA < 35 000€ permet la franchise de TVA.")
                
                # 5. Mission patrimoniale IR élevé
                if impot_revenu > 6000:
                    opportunites.append("Étudier mission patrimoniale IR élevé")
                    details.append(f"Impôt sur le revenu: {impot_revenu:.2f}€. L'IR élevé justifie une analyse patrimoniale approfondie pour optimisation fiscale.")
                
                # 6. Passage à l'IS pour entreprises à l'IR avec résultat modéré
                if forme_juri != "SCI" and resultat < 40000 and imposition == "IR":
                    opportunites.append("Étudier passage à l'IS (coût social potentiellement élevé)")
                    details.append(f"Forme juridique: {forme_juri}, Imposition actuelle: IR, Résultat: {resultat:.2f}€. Le passage à l'IS peut réduire les charges sociales.")
                
                # 7. Prime d'activité
                if forme_juri != "SCI" and impot_revenu < 2000 and impot_revenu != 0:
                    opportunites.append("Étudier bénéfice prime d'activité")
                    details.append(f"Impôt sur le revenu: {impot_revenu:.2f}€. Faible IR peut indiquer éligibilité à la prime d'activité.")
                
                # 8. Augmentation rémunération de gérance
                if imposition == "IS" and forme_juri != "SCI" and resultat > 42500:
                    opportunites.append("Étudier augmentation rémunération de gérance")
                    details.append(f"Imposition: IS, Résultat: {resultat:.2f}€. Augmenter la rémunération pourrait optimiser la fiscalité globale.")
                
                # 9. Passage TNS pour SAS avec masse salariale
                if forme_juri == "SAS" and m_salariale > 1 and imposition == "IS":
                    opportunites.append("Étudier passage TNS")
                    details.append(f"Forme juridique: SAS, Masse salariale: {m_salariale:.2f}€, Imposition: IS. Le statut TNS pourrait réduire les charges sociales.")
                
                # 10. Placement de trésorerie excédentaire
                if tresorerie > 100000:
                    opportunites.append("Étudier placement de l'excédent de trésorerie")
                    details.append(f"Trésorerie: {tresorerie:.2f}€. L'excédent pourrait être placé pour générer des revenus supplémentaires.")
                
                # 11. NOUVELLE RECOMMANDATION : Marge négative pour activités de vente
                if prestation != "presta" and forme_juri != "SCI" and compte_70 - compte_60 < 0:
                    opportunites.append("ANOMALIE : Marge négative")
                    details.append(f"Ventes: {compte_70:.2f}€, Achats: {compte_60:.2f}€, Marge: {(compte_70 - compte_60):.2f}€. Attention, la marge (ventes - achats) est négative pour une activité non prestataire, ce qui est anormal.")
                
                # Si des opportunités ont été détectées, les enregistrer
                if opportunites:
                    nb_dossiers_avec_reco += 1
                    total_recommandations += len(opportunites)
                    
                    # Convertir en chaînes de caractères pour stockage
                    opportunites_str = ";;".join(opportunites)
                    details_str = ";;".join(details)
                    
                    # Enregistrer dans la base de données
                    import datetime
                    date_actuelle = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    cursor.execute("""
                        INSERT INTO recommandations (siret, opportunites, details, date_detection)
                        VALUES (?, ?, ?, ?)
                    """, (siret, opportunites_str, details_str, date_actuelle))
                
                # Mise à jour de la barre de progression
                self.progress_bar['value'] = (idx + 1) / total_dossiers * 100
                self.progress_label.config(text=f"Analyse des dossiers: {idx + 1}/{total_dossiers}")
                
                # Mettre à jour l'interface tous les 10 dossiers
                if idx % 10 == 0:
                    self.root.update_idletasks()
            
            # Valider les modifications dans la base de données
            conn.commit()
            conn.close()
            
            # Mettre à jour le statut
            self.reco_status_label.config(
                text=f"Analyse terminée: {nb_dossiers_avec_reco} dossiers avec recommandations, {total_recommandations} recommandations au total"
            )
            
            # Afficher un message de réussite
            messagebox.showinfo(
                "Analyse terminée", 
                f"L'analyse des dossiers est terminée.\n\n"
                f"{nb_dossiers_avec_reco} dossiers sur {total_dossiers} ont des recommandations.\n"
                f"{total_recommandations} recommandations ont été identifiées au total."
            )
            
            # Si des recommandations ont été trouvées, proposer d'afficher les résultats
            if nb_dossiers_avec_reco > 0:
                if messagebox.askyesno("Afficher les résultats?", "Voulez-vous afficher les dossiers avec des recommandations?"):
                    self.afficher_dossiers_avec_recommandations()
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de la détection des opportunités: {str(e)}")
            import traceback
            traceback.print_exc()

    def afficher_dossiers_avec_recommandations(self):
        """Affiche la liste des dossiers avec des recommandations"""
        try:
            # Connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            
            # Récupérer les dossiers avec des recommandations
            query = """
            SELECT r.siret, COALESCE(i.nom, r.siret) as nom, 
                   COUNT(r.opportunites) as nb_opportunites,
                   i.forme_juri, i.code_naf, i.CA,
                   i.nomUniteLegale, i.prenom1UniteLegale
            FROM recommandations r
            LEFT JOIN indicateurs i ON r.siret = i.siret
            GROUP BY r.siret
            ORDER BY nb_opportunites DESC
            """
            df = pd.read_sql_query(query, conn)
            
            # Si aucun résultat
            if df.empty:
                messagebox.showinfo("Information", "Aucune recommandation n'a été détectée. Veuillez exécuter l'algorithme de détection d'abord.")
                conn.close()
                return
            
            # Créer une fenêtre pour afficher les résultats
            results_window = tk.Toplevel(self.root)
            results_window.title("Dossiers avec Recommandations")
            results_window.geometry("1200x768")  # Fenêtre plus large pour accommoder les colonnes supplémentaires
            results_window.configure(bg="#f0f5f9")
            
            # Créer un cadre principal
            main_frame = ttk.Frame(results_window)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            # Ajouter un titre
            title_label = tk.Label(
                main_frame, 
                text="Dossiers avec Recommandations", 
                font=("Helvetica", 18, "bold"), 
                bg="#f0f5f9"
            )
            title_label.pack(pady=(0, 20))
            
            # Créer un cadre pour le tableau
            result_frame = ttk.Frame(main_frame)
            result_frame.pack(fill=tk.BOTH, expand=True)
            
            # Ajouter des barres de défilement
            scrollbar_y = ttk.Scrollbar(result_frame, orient="vertical")
            scrollbar_x = ttk.Scrollbar(result_frame, orient="horizontal")
            
            # Créer un tableau (Treeview) avec les nouvelles colonnes
            columns = ["SIRET", "Nom", "Nombre d'opportunités", "Forme Juridique", "Code NAF", "CA", "Nom Unité Légale", "Prénom Unité Légale"]
            tree = ttk.Treeview(
                result_frame, 
                columns=columns, 
                show='headings', 
                yscrollcommand=scrollbar_y.set, 
                xscrollcommand=scrollbar_x.set
            )
            
            # Configurer les colonnes
            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=150, anchor="center")
            
            # Configurer les barres de défilement
            scrollbar_y.config(command=tree.yview)
            scrollbar_y.pack(side="right", fill="y")
            
            scrollbar_x.config(command=tree.xview)
            scrollbar_x.pack(side="bottom", fill="x")
            
            tree.pack(fill="both", expand=True)
            
            # Remplir le tableau avec les données
            for _, row in df.iterrows():
                # Formater le CA
                ca_formate = f"{row['CA']:,.2f} €" if pd.notna(row['CA']) else "N/A"
                
                # Formater les noms et prénoms d'unité légale
                nom_unite = row['nomUniteLegale'] if pd.notna(row['nomUniteLegale']) else "N/A"
                prenom_unite = row['prenom1UniteLegale'] if pd.notna(row['prenom1UniteLegale']) else "N/A"
                
                tree.insert("", "end", values=[
                    row['siret'], 
                    row['nom'], 
                    row['nb_opportunites'], 
                    row['forme_juri'], 
                    row['code_naf'], 
                    ca_formate,
                    nom_unite,
                    prenom_unite
                ])
            
            # Ajouter un événement de double-clic pour afficher les détails
            tree.bind("<Double-1>", lambda event: self.afficher_details_recommandations(event, tree))
            
            # Ajouter un texte d'aide
            help_text = tk.Label(
                main_frame, 
                text="Double-cliquez sur un dossier pour voir les recommandations détaillées", 
                font=("Helvetica", 10, "italic"), 
                bg="#f0f5f9"
            )
            help_text.pack(pady=10)
            
            # Ajouter des boutons d'action
            button_frame = tk.Frame(main_frame, bg="#f0f5f9")
            button_frame.pack(pady=10)
            
            # Bouton pour exporter en CSV
            export_btn = tk.Button(
                button_frame, 
                text="Exporter en CSV", 
                command=lambda: self.exporter_recommandations_csv(tree),
                bg="#3282b8", 
                fg="white"
            )
            export_btn.pack(side=tk.LEFT, padx=10)
            
            # Bouton pour fermer
            close_btn = tk.Button(
                button_frame, 
                text="Fermer", 
                command=results_window.destroy,
                bg="#0f4c75", 
                fg="white"
            )
            close_btn.pack(side=tk.LEFT, padx=10)
            
            conn.close()
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'affichage des recommandations: {str(e)}")
            if 'conn' in locals():
                conn.close()


    def afficher_details_recommandations(self, event, tree):
        """Affiche les détails des recommandations pour un dossier spécifique"""
        # Identifier l'élément cliqué
        item = tree.identify('item', event.x, event.y)
        if not item:
            return
        
        # Récupérer les valeurs de la ligne sélectionnée
        values = tree.item(item, 'values')
        siret = values[0]
        nom = values[1]
        
        try:
            # Connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Récupérer les détails du dossier
            cursor.execute("""
                SELECT i.*, r.opportunites, r.details
                FROM indicateurs i
                JOIN recommandations r ON i.siret = r.siret
                WHERE i.siret = ?
            """, (siret,))
            
            dossier = cursor.fetchone()
            if not dossier:
                messagebox.showinfo("Information", f"Aucune information détaillée disponible pour le dossier {nom}")
                conn.close()
                return
            
            # Récupérer les colonnes de la table indicateurs
            cursor.execute("PRAGMA table_info(indicateurs)")
            column_info = cursor.fetchall()
            column_names = [col[1] for col in column_info]
            
            # Créer un dictionnaire pour stocker les données du dossier
            dossier_dict = {column_names[i]: dossier[i] for i in range(len(column_names))}
            
            # Récupérer les opportunités et détails
            opportunites = dossier[-2].split(";;")
            details = dossier[-1].split(";;")
            
            # Créer une fenêtre pour afficher les détails
            details_window = tk.Toplevel(self.root)
            details_window.title(f"Recommandations pour {nom}")
            details_window.geometry("900x700")
            details_window.configure(bg="#f0f5f9")
            
            # Créer un cadre principal avec défilement
            main_frame = ttk.Frame(details_window)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            # Ajouter un titre
            title_label = tk.Label(
                main_frame, 
                text=f"Recommandations pour {nom}", 
                font=("Helvetica", 18, "bold"), 
                bg="#f0f5f9"
            )
            title_label.pack(pady=(0, 20))
            
            # Créer un cadre avec défilement
            canvas = tk.Canvas(main_frame, bg='#f0f5f9', highlightthickness=0)
            scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Section avec les informations générales du dossier
            info_frame = ttk.LabelFrame(scrollable_frame, text="Informations du dossier")
            info_frame.pack(fill="x", padx=10, pady=10)
            
            # Sélectionner les informations à afficher (les plus importantes)
            important_infos = ["siret", "nom", "forme_juri", "CA", "resultat", "trésorerie", 
                              "Imposition", "Regime_de_TVA", "code_naf", "impot_revenu", 
                              "emprunt", "m_salariale"]
            
            # Afficher les informations importantes dans une grille
            for i, info in enumerate(important_infos):
                if info in dossier_dict:
                    row = i // 3
                    col = i % 3
                    
                    # Label pour le nom du champ
                    field_label = ttk.Label(
                        info_frame, 
                        text=f"{info}:", 
                        font=("Helvetica", 10, "bold")
                    )
                    field_label.grid(row=row, column=col*2, padx=10, pady=5, sticky="e")
                    
                    # Formater la valeur si c'est un nombre
                    value = dossier_dict[info]
                    try:
                        value = float(value)
                        value_str = f"{value:,.2f}" if abs(value) > 0.01 else "0"
                        # Ajouter € pour les valeurs monétaires
                        if info in ["CA", "resultat", "trésorerie", "impot_revenu", "emprunt", "m_salariale"]:
                            value_str += " €"
                    except (TypeError, ValueError):
                        value_str = str(value) if value is not None else "N/A"
                    
                    # Label pour la valeur
                    value_label = ttk.Label(
                        info_frame, 
                        text=value_str, 
                        font=("Helvetica", 10)
                    )
                    value_label.grid(row=row, column=col*2+1, padx=10, pady=5, sticky="w")
            
            # Section avec les recommandations
            reco_frame = ttk.LabelFrame(scrollable_frame, text="Recommandations")
            reco_frame.pack(fill="x", padx=10, pady=10)
            
            # Afficher chaque recommandation avec son détail
            for i, (opportunite, detail) in enumerate(zip(opportunites, details)):
                # Cadre pour cette recommandation
                reco_item_frame = ttk.Frame(reco_frame)
                reco_item_frame.pack(fill="x", padx=10, pady=5)
                
                # Titre de la recommandation
                reco_title = ttk.Label(
                    reco_item_frame, 
                    text=f"{i+1}. {opportunite}", 
                    font=("Helvetica", 12, "bold")
                )
                reco_title.pack(anchor="w", pady=(5, 2))
                
                # Détail de la recommandation
                reco_detail = ttk.Label(
                    reco_item_frame, 
                    text=detail, 
                    font=("Helvetica", 10),
                    wraplength=800,
                    justify="left"
                )
                reco_detail.pack(anchor="w", padx=20, pady=(0, 5))
                
                # Ajouter un séparateur
                ttk.Separator(reco_frame, orient="horizontal").pack(fill="x", padx=5, pady=5)
            
            # Bouton pour fermer
            close_btn = tk.Button(
                scrollable_frame, 
                text="Fermer", 
                command=details_window.destroy,
                bg="#0f4c75", 
                fg="white",
                font=("Helvetica", 10, "bold"),
                padx=20, 
                pady=5
            )
            close_btn.pack(pady=20)
            
            conn.close()
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'affichage des détails: {str(e)}")
            if 'conn' in locals():
                conn.close()

    def exporter_recommandations_csv(self, tree):
        """Exporte la liste des dossiers avec recommandations en CSV"""
        try:
            # Proposer à l'utilisateur de choisir l'emplacement du fichier
            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv")],
                title="Enregistrer les recommandations"
            )
            
            if not file_path:
                return  # L'utilisateur a annulé
            
            # Récupérer les données du tableau
            data = []
            columns = []
            
            # Récupérer les en-têtes de colonnes
            for idx, col_name in enumerate(tree['columns']):
                columns.append(tree.heading(col_name)['text'])
            
            # Récupérer les données des lignes
            for item in tree.get_children():
                values = tree.item(item, 'values')
                data.append(values)
            
            # Écrire dans le fichier CSV
            import csv
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=';')
                writer.writerow(columns)
                writer.writerows(data)
            
            messagebox.showinfo("Export réussi", f"Les données ont été exportées avec succès vers {file_path}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur s'est produite lors de l'exportation: {str(e)}")
            
    def generer_rapport_recommandations(self):
        """Génère un rapport complet des recommandations au format Excel"""
        try:
            # Vérifier si la bibliothèque openpyxl est disponible
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            # Proposer à l'utilisateur de choisir l'emplacement du fichier
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Enregistrer le rapport de recommandations"
            )
            
            if not file_path:
                return  # L'utilisateur a annulé
            
            # Connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            
            # Récupérer tous les dossiers avec des recommandations
            query = """
            SELECT 
                r.siret, 
                COALESCE(i.nom, r.siret) as nom,
                i.forme_juri,
                i.code_naf,
                i.CA,
                i.resultat,
                i.Imposition,
                i.Regime_de_TVA,
                i.emprunt,
                i.impot_revenu,
                i.trésorerie,
                i.m_salariale,
                i.nomUniteLegale,
                i.prenom1UniteLegale,
                r.opportunites,
                r.details
            FROM recommandations r
            LEFT JOIN indicateurs i ON r.siret = i.siret
            ORDER BY r.siret
            """
            
            df = pd.read_sql_query(query, conn)
            conn.close()
            
            if df.empty:
                messagebox.showinfo("Information", "Aucune recommandation n'a été détectée. Veuillez exécuter l'algorithme de détection d'abord.")
                return
            
            # Créer un nouveau classeur Excel
            wb = openpyxl.Workbook()
            
            # Supprimer la feuille par défaut
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            
            # Créer une feuille pour le résumé
            ws_resume = wb.create_sheet("Résumé")
            
            # Définir les styles
            title_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
            header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
            normal_font = Font(name='Calibri', size=11)
            
            # Couleurs de fond
            header_fill = PatternFill(start_color="3282B8", end_color="3282B8", fill_type="solid")
            alt_row_fill = PatternFill(start_color="EEF4F9", end_color="EEF4F9", fill_type="solid")
            
            # Bordures
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Alignement
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            
            # Titre du rapport
            ws_resume.merge_cells('A1:H1')  # Élargir pour accommoder les colonnes supplémentaires
            cell = ws_resume.cell(row=1, column=1, value="Rapport de Recommandations d'Optimisation")
            cell.font = title_font
            cell.fill = header_fill
            cell.alignment = center_align
            
            # En-têtes de colonnes pour le résumé (avec les nouvelles colonnes)
            headers = ["SIRET", "Nom", "Nombre de recommandations", "Forme Juridique", "CA", "Nom Unité Légale", "Prénom Unité Légale", "Détails"]
            for col, header in enumerate(headers, 1):
                cell = ws_resume.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align
                
                # Ajuster la largeur des colonnes
                if col == 1:  # SIRET
                    ws_resume.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                elif col == 2:  # Nom
                    ws_resume.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30
                elif col == 3:  # Nombre
                    ws_resume.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                elif col == 4:  # Forme juridique
                    ws_resume.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                elif col == 5:  # CA
                    ws_resume.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                elif col == 6:  # Nom Unité Légale
                    ws_resume.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
                elif col == 7:  # Prénom Unité Légale
                    ws_resume.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
                elif col == 8:  # Détails
                    ws_resume.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 50
            
            # Remplir les données du résumé
            row = 4
            for _, dossier in df.iterrows():
                # Créer une nouvelle feuille pour ce dossier
                siret = dossier['siret']
                nom_court = dossier['nom'][:31] if dossier['nom'] else siret  # Limite pour nom de feuille Excel
                # Remplacer les caractères invalides pour un nom d'onglet Excel
                sheet_name = re.sub(r'[\\/*?:\[\]]', '_', nom_court)
                
                # Éviter les doublons de noms de feuilles
                counter = 1
                base_name = sheet_name
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{base_name}_{counter}"
                    counter += 1
                
                # Créer la feuille pour ce dossier
                ws_dossier = wb.create_sheet(sheet_name)
                
                # Extraire et compter les recommandations
                opportunites = dossier['opportunites'].split(';;') if dossier['opportunites'] else []
                details = dossier['details'].split(';;') if dossier['details'] else []
                nb_recommandations = len(opportunites)
                
                # Ajouter dans le résumé
                ws_resume.cell(row=row, column=1, value=siret).font = normal_font
                ws_resume.cell(row=row, column=2, value=dossier['nom']).font = normal_font
                ws_resume.cell(row=row, column=3, value=nb_recommandations).font = normal_font
                ws_resume.cell(row=row, column=4, value=dossier['forme_juri']).font = normal_font
                
                # Formater le CA
                try:
                    ca_value = float(dossier['CA']) if dossier['CA'] is not None and not pd.isna(dossier['CA']) else 0
                    ws_resume.cell(row=row, column=5, value=ca_value).number_format = '#,##0.00 €'
                except (ValueError, TypeError):
                    ws_resume.cell(row=row, column=5, value="N/A")
                
                # Ajouter les noms et prénoms d'unité légale
                ws_resume.cell(row=row, column=6, value=dossier['nomUniteLegale'] if pd.notna(dossier['nomUniteLegale']) else "N/A").font = normal_font
                ws_resume.cell(row=row, column=7, value=dossier['prenom1UniteLegale'] if pd.notna(dossier['prenom1UniteLegale']) else "N/A").font = normal_font
                
                # Liste des recommandations pour la colonne Détails
                details_resume = "; ".join(opportunites)
                ws_resume.cell(row=row, column=8, value=details_resume).font = normal_font
                
                # Appliquer un style alterné aux lignes
                if row % 2 == 0:
                    for col in range(1, 9):  # Mise à jour pour inclure les nouvelles colonnes
                        ws_resume.cell(row=row, column=col).fill = alt_row_fill
                
                # Ajouter des bordures
                for col in range(1, 9):  # Mise à jour pour inclure les nouvelles colonnes
                    ws_resume.cell(row=row, column=col).border = thin_border
                
                # Remplir la feuille détaillée pour ce dossier
                # Titre
                ws_dossier.merge_cells('A1:F1')
                cell = ws_dossier.cell(row=1, column=1, value=f"Analyse détaillée pour {dossier['nom']}")
                cell.font = title_font
                cell.fill = header_fill
                cell.alignment = center_align
                
                # Informations générales du dossier
                ws_dossier.cell(row=3, column=1, value="Informations générales").font = header_font
                
                info_headers = ["Champ", "Valeur"]
                for col, header in enumerate(info_headers, 1):
                    cell = ws_dossier.cell(row=4, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = center_align
                
                # Ajuster les largeurs
                ws_dossier.column_dimensions['A'].width = 20
                ws_dossier.column_dimensions['B'].width = 30
                
                # Informations importantes à afficher, incluant les nouvelles colonnes
                important_fields = [
                    ("SIRET", "siret"),
                    ("Nom", "nom"),
                    ("Nom Unité Légale", "nomUniteLegale"),
                    ("Prénom Unité Légale", "prenom1UniteLegale"),
                    ("Forme juridique", "forme_juri"),
                    ("CA", "CA"),
                    ("Résultat", "resultat"),
                    ("Imposition", "Imposition"),
                    ("Régime TVA", "Regime_de_TVA"),
                    ("Code NAF", "code_naf"),
                    ("Impôt sur le revenu", "impot_revenu"),
                    ("Emprunt", "emprunt"),
                    ("Trésorerie", "trésorerie"),
                    ("Masse salariale", "m_salariale")
                ]
                
                # Remplir les informations
                info_row = 5
                for display_name, field_name in important_fields:
                    ws_dossier.cell(row=info_row, column=1, value=display_name).font = normal_font
                    
                    # Récupérer et formater la valeur
                    value = dossier[field_name]
                    if field_name in ["CA", "resultat", "impot_revenu", "emprunt", "trésorerie", "m_salariale"]:
                        try:
                            numeric_value = float(value) if value is not None and not pd.isna(value) else 0
                            cell = ws_dossier.cell(row=info_row, column=2, value=numeric_value)
                            cell.number_format = '#,##0.00 €'
                        except (ValueError, TypeError):
                            ws_dossier.cell(row=info_row, column=2, value="N/A")
                    else:
                        if pd.isna(value):
                            value = "N/A"
                        ws_dossier.cell(row=info_row, column=2, value=value).font = normal_font
                    
                    # Styles
                    if info_row % 2 == 0:
                        ws_dossier.cell(row=info_row, column=1).fill = alt_row_fill
                        ws_dossier.cell(row=info_row, column=2).fill = alt_row_fill
                    
                    # Bordures
                    ws_dossier.cell(row=info_row, column=1).border = thin_border
                    ws_dossier.cell(row=info_row, column=2).border = thin_border
                    
                    info_row += 1
                
                # Section des recommandations
                reco_start_row = info_row + 2
                ws_dossier.merge_cells(f'A{reco_start_row}:F{reco_start_row}')
                cell = ws_dossier.cell(row=reco_start_row, column=1, value="Recommandations")
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                
                # En-têtes des recommandations
                reco_headers = ["N°", "Recommandation", "Détails"]
                for col, header in enumerate(reco_headers, 1):
                    cell = ws_dossier.cell(row=reco_start_row+1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = center_align
                
                # Ajuster les colonnes
                ws_dossier.column_dimensions['A'].width = 5
                ws_dossier.column_dimensions['B'].width = 30
                ws_dossier.column_dimensions['C'].width = 60
                
                # Remplir les recommandations
                for i, (opportunite, detail) in enumerate(zip(opportunites, details), 1):
                    # Numéro
                    ws_dossier.cell(row=reco_start_row+1+i, column=1, value=i).font = normal_font
                    
                    # Recommandation
                    ws_dossier.cell(row=reco_start_row+1+i, column=2, value=opportunite).font = normal_font
                    
                    # Détails
                    cell = ws_dossier.cell(row=reco_start_row+1+i, column=3, value=detail)
                    cell.font = normal_font
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # Styles alternés
                    if i % 2 == 0:
                        for col in range(1, 4):
                            ws_dossier.cell(row=reco_start_row+1+i, column=col).fill = alt_row_fill
                    
                    # Bordures
                    for col in range(1, 4):
                        ws_dossier.cell(row=reco_start_row+1+i, column=col).border = thin_border
                
                # Passer à la ligne suivante dans le résumé
                row += 1
            
            # Sauvegarder le fichier
            wb.save(file_path)
            
            messagebox.showinfo("Rapport généré", f"Le rapport a été généré avec succès à l'emplacement:\n{file_path}")
            
            # Demander à l'utilisateur s'il souhaite ouvrir le fichier
            if messagebox.askyesno("Ouvrir le rapport?", "Voulez-vous ouvrir le rapport maintenant?"):
                try:
                    import os
                    import subprocess
                    if os.name == 'nt':  # Windows
                        os.startfile(file_path)
                    elif os.name == 'posix':  # macOS et Linux
                        if sys.platform == 'darwin':  # macOS
                            subprocess.call(['open', file_path])
                        else:  # Linux
                            subprocess.call(['xdg-open', file_path])
                except Exception as e:
                    messagebox.showerror("Erreur", f"Impossible d'ouvrir le fichier: {str(e)}")
            
        except ImportError:
            messagebox.showerror("Erreur", "La bibliothèque openpyxl n'est pas installée. Veuillez l'installer pour utiliser cette fonctionnalité.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur s'est produite lors de la génération du rapport: {str(e)}")
            import traceback
            traceback.print_exc()



    def analyser_dossier_individuel(self):
        """Permet à l'utilisateur de sélectionner un dossier pour analyse détaillée"""
        try:
            # Connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Récupérer la liste des dossiers avec nom et prénom d'unité légale
            cursor.execute("""
                SELECT siret, 
                       COALESCE(nom, siret) as nom, 
                       nomUniteLegale, 
                       prenom1UniteLegale 
                FROM indicateurs
                WHERE siret != 'moyenne'
                ORDER BY nom
            """)
            
            dossiers = cursor.fetchall()
            conn.close()
            
            if not dossiers:
                messagebox.showinfo("Information", "Aucun dossier disponible.")
                return
            
            # Créer une fenêtre de sélection
            selection_window = tk.Toplevel(self.root)
            selection_window.title("Sélection de dossier")
            selection_window.geometry("650x550")
            selection_window.configure(bg="#f0f5f9")
            
            # Empêcher le redimensionnement pour une meilleure apparence
            selection_window.resizable(False, False)
            
            # Ajouter un titre avec fond dégradé
            title_frame = GradientFrame(selection_window, height=60, color1="#1E88E5", color2="#1DE9B6")
            title_frame.pack(fill=tk.X)
            
            title_frame.create_text(20, 30, anchor="w", text="Sélectionnez un dossier à analyser", 
                                    font=('Arial', 16, 'bold'), fill="white")
            
            # Créer un cadre principal
            main_frame = ttk.Frame(selection_window, style="TFrame")
            main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            # Cadre de recherche
            search_frame = ttk.Frame(main_frame)
            search_frame.pack(fill="x", pady=10)
            
            search_label = ttk.Label(search_frame, text="Rechercher:", font=("Helvetica", 11))
            search_label.pack(side="left", padx=5)
            
            # Variable pour la recherche
            search_var = tk.StringVar()
            
            # Fonction de mise à jour en temps réel de la liste
            def update_list(*args):
                search_text = search_var.get().lower()
                listbox.delete(0, tk.END)
                
                for siret, nom, nom_unite, prenom_unite in dossiers:
                    # Construire un texte d'affichage avec plus d'informations
                    nom_unite_str = str(nom_unite) if nom_unite else ""
                    prenom_unite_str = str(prenom_unite) if prenom_unite else ""
                    
                    # Format d'affichage: "Nom dossier - Prénom Nom (SIRET)"
                    if nom_unite_str or prenom_unite_str:
                        nom_display = f"{nom} - {prenom_unite_str} {nom_unite_str} ({siret})"
                    else:
                        nom_display = f"{nom} ({siret})"
                    
                    # Recherche dans tous les champs
                    nom_str = str(nom).lower()
                    siret_str = str(siret).lower()
                    nom_unite_str = nom_unite_str.lower()
                    prenom_unite_str = prenom_unite_str.lower()
                    
                    if (search_text == "" or 
                        search_text in nom_str or 
                        search_text in siret_str or
                        search_text in nom_unite_str or
                        search_text in prenom_unite_str):
                        listbox.insert(tk.END, nom_display)
                
                # Sélectionner le premier élément s'il existe
                if listbox.size() > 0:
                    listbox.selection_set(0)
            
            # Associer la fonction à la variable
            search_var.trace_add("write", update_list)
            
            # Champ de recherche
            search_entry = ttk.Entry(search_frame, textvariable=search_var, width=55, font=("Helvetica", 11))
            search_entry.pack(side="left", padx=5, fill="x", expand=True)
            search_entry.focus_set()  # Donner le focus au champ de recherche
            
            # Cadre pour la liste avec défilement
            list_frame = ttk.Frame(main_frame)
            list_frame.pack(fill="both", expand=True, pady=10)
            
            # Barre de défilement
            scrollbar = ttk.Scrollbar(list_frame)
            scrollbar.pack(side="right", fill="y")
            
            # Liste des dossiers
            listbox = tk.Listbox(
                list_frame, 
                yscrollcommand=scrollbar.set, 
                font=("Helvetica", 12), 
                selectbackground="#3282b8",
                selectforeground="white",
                height=15,
                activestyle="none"
            )
            listbox.pack(side="left", fill="both", expand=True)
            
            # Configurer la barre de défilement
            scrollbar.config(command=listbox.yview)
            
            # Remplir la liste initialement
            for siret, nom, nom_unite, prenom_unite in dossiers:
                # Construire un texte d'affichage avec plus d'informations
                nom_unite_str = str(nom_unite) if nom_unite else ""
                prenom_unite_str = str(prenom_unite) if prenom_unite else ""
                
                # Format d'affichage: "Nom dossier - Prénom Nom (SIRET)"
                if nom_unite_str or prenom_unite_str:
                    nom_display = f"{nom} - {prenom_unite_str} {nom_unite_str} ({siret})"
                else:
                    nom_display = f"{nom} ({siret})"
                    
                listbox.insert(tk.END, nom_display)
            
            # Fonction pour gérer la sélection d'un dossier
            def selectionner_dossier():
                selection = listbox.curselection()
                if selection:
                    index = selection[0]
                    displayed_text = listbox.get(index)
                    
                    # Extraire le SIRET entre parenthèses à la fin du texte
                    siret = displayed_text.split('(')[-1].rstrip(')')
                    
                    # Fermer la fenêtre avant d'appeler l'analyse
                    selection_window.destroy()
                    
                    # Lancer l'analyse
                    self.analyser_dossier_specifique(siret)
                else:
                    messagebox.showinfo("Information", "Veuillez sélectionner un dossier.")
            
            # Permettre la sélection par double-clic
            listbox.bind("<Double-1>", lambda event: selectionner_dossier())
            
            # Permettre la sélection par touche Entrée
            def on_enter(event):
                selectionner_dossier()
            listbox.bind("<Return>", on_enter)
            search_entry.bind("<Return>", on_enter)
            
            # Boutons d'action - dans un cadre séparé pour éviter les chevauchements
            button_frame = ttk.Frame(selection_window)
            button_frame.pack(pady=20, fill="x")
            
            # Centrer les boutons
            button_container = ttk.Frame(button_frame)
            button_container.pack(side="top", anchor="center")
            
            # Bouton Analyser avec style
            analyser_btn = tk.Button(
                button_container, 
                text="Analyser", 
                command=selectionner_dossier,
                bg="#3282b8", 
                fg="white",
                font=("Helvetica", 11, "bold"),
                width=15,
                height=2,
                relief="flat"
            )
            analyser_btn.pack(side="left", padx=20)
            
            # Bouton Annuler avec style
            annuler_btn = tk.Button(
                button_container, 
                text="Annuler", 
                command=selection_window.destroy,
                bg="#0f4c75", 
                fg="white",
                font=("Helvetica", 11, "bold"),
                width=15,
                height=2,
                relief="flat"
            )
            annuler_btn.pack(side="left", padx=20)
            
            # Gestion des effets de survol pour les boutons
            def on_enter_button(e, button, color):
                button.config(background=color)
                
            def on_leave_button(e, button, color):
                button.config(background=color)
            
            analyser_btn.bind("<Enter>", lambda e: on_enter_button(e, analyser_btn, "#2471A3"))
            analyser_btn.bind("<Leave>", lambda e: on_leave_button(e, analyser_btn, "#3282b8"))
            
            annuler_btn.bind("<Enter>", lambda e: on_enter_button(e, annuler_btn, "#0A3A5E"))
            annuler_btn.bind("<Leave>", lambda e: on_leave_button(e, annuler_btn, "#0f4c75"))
            
            # Démarrer avec la liste complète
            update_list()
            
            # Centrer la fenêtre par rapport à l'application principale
            selection_window.update_idletasks()
            width = selection_window.winfo_width()
            height = selection_window.winfo_height()
            x = self.root.winfo_x() + (self.root.winfo_width() - width) // 2
            y = self.root.winfo_y() + (self.root.winfo_height() - height) // 2
            selection_window.geometry(f"{width}x{height}+{x}+{y}")
            
            # Rendre la fenêtre modale
            selection_window.transient(self.root)
            selection_window.grab_set()
            self.root.wait_window(selection_window)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue: {str(e)}")
            import traceback
            traceback.print_exc()

    def filter_dossiers(self, search_var, listbox, dossiers):
        """Filtre les dossiers en fonction du texte de recherche"""
        search_text = search_var.get().lower()
        listbox.delete(0, tk.END)
        
        for siret, nom in dossiers:
            if search_text in nom.lower() or search_text in siret.lower():
                listbox.insert(tk.END, f"{nom} ({siret})")

    def analyser_dossier_specifique(self, siret):
        """Analyse un dossier spécifique et affiche les recommandations personnalisées"""
        try:
            # Connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Récupérer les informations du dossier
            cursor.execute("""
                SELECT * FROM indicateurs
                WHERE siret = ?
            """, (siret,))
            
            dossier = cursor.fetchone()
            
            if not dossier:
                messagebox.showinfo("Information", f"Aucune information disponible pour le dossier {siret}")
                conn.close()
                return
            
            # Récupérer les noms des colonnes
            cursor.execute("PRAGMA table_info(indicateurs)")
            columns = [col[1] for col in cursor.fetchall()]
            
            # Créer un dictionnaire pour faciliter l'accès aux valeurs
            dossier_dict = dict(zip(columns, dossier))
            
            # Fermer la connexion à la base de données
            conn.close()
            
            # Récupérer les valeurs importantes (avec gestion des valeurs null/None)
            nom = dossier_dict.get('nom') or siret
            forme_juri = dossier_dict.get('forme_juri') or ""
            imposition = dossier_dict.get('Imposition') or ""
            emprunt = float(dossier_dict.get('emprunt') or 0)
            code_naf = dossier_dict.get('code_naf') or ""
            prestation = dossier_dict.get('prestation') or ""
            ca = float(dossier_dict.get('CA') or 0)
            regime_tva = dossier_dict.get('Regime_de_TVA') or ""
            impot_revenu = float(dossier_dict.get('impot_revenu') or 0)
            resultat = float(dossier_dict.get('resultat') or 0)
            m_salariale = float(dossier_dict.get('m_salariale') or 0)
            tresorerie = float(dossier_dict.get('trésorerie') or 0)
            
            # Liste pour stocker les recommandations
            recommandations = []
            
            # Appliquer les règles métier
            # 1. SCI à l'IR avec emprunt significatif
            if forme_juri == "SCI" and imposition == "IR" and emprunt < -30000:
                recommandations.append({
                    "titre": "Étudier passage IS",
                    "detail": f"Forme juridique: SCI, Imposition actuelle: IR, Montant emprunt: {emprunt:.2f}€. Un passage à l'IS pourrait être avantageux pour déduire les intérêts d'emprunt.",
                    "formule": "forme_juri = SCI ET Imposition = IR ET emprunt < -30000"
                })
            
            # 2. Exonération CFE presse
            if code_naf == "4726Z" or code_naf == "4762Z":
                recommandations.append({
                    "titre": "Étudier exonération CFE presse",
                    "detail": f"Code NAF: {code_naf}. Ce code est lié au secteur de la presse et peut permettre une exonération de CFE.",
                    "formule": "code_naf = 4726Z OU code_naf = 4762Z"
                })
            
            # 3. Franchise de TVA pour activités de vente
            if prestation != "presta" and ca < 85000 and forme_juri != "SCI" and ca != 0 and regime_tva != "NI":
                recommandations.append({
                    "titre": "Étudier passage en franchise de TVA",
                    "detail": f"CA: {ca:.2f}€, Activité de vente (non prestation), Régime TVA actuel: {regime_tva}. Un CA < 85 000€ permet la franchise de TVA.",
                    "formule": "prestation = none ET CA < 85000 ET forme_juri ≠ SCI ET CA ≠ 0 ET Regime_de_TVA ≠ NI"
                })
            
            # 4. Franchise de TVA pour activités de prestation
            if prestation == "presta" and ca < 35000 and forme_juri != "SCI" and ca != 0 and regime_tva != "NI":
                recommandations.append({
                    "titre": "Étudier passage en franchise de TVA",
                    "detail": f"CA: {ca:.2f}€, Activité de prestation, Régime TVA actuel: {regime_tva}. Un CA < 35 000€ permet la franchise de TVA.",
                    "formule": "prestation = presta ET CA < 35000 ET forme_juri ≠ SCI ET CA ≠ 0 ET Regime_de_TVA ≠ NI"
                })
            
            # 5. Mission patrimoniale IR élevé
            if impot_revenu > 6000:
                recommandations.append({
                    "titre": "Étudier mission patrimoniale IR élevé",
                    "detail": f"Impôt sur le revenu: {impot_revenu:.2f}€. L'IR élevé justifie une analyse patrimoniale approfondie pour optimisation fiscale.",
                    "formule": "impot_revenu > 6000"
                })
            
            # 6. Passage à l'IS pour entreprises à l'IR avec résultat modéré
            if forme_juri != "SCI" and resultat < 40000 and imposition == "IR":
                recommandations.append({
                    "titre": "Étudier passage à l'IS (coût social potentiellement élevé)",
                    "detail": f"Forme juridique: {forme_juri}, Imposition actuelle: IR, Résultat: {resultat:.2f}€. Le passage à l'IS peut réduire les charges sociales.",
                    "formule": "forme_juri ≠ SCI ET resultat < 40000 ET Imposition = IR"
                })
            
            # 7. Prime d'activité
            if forme_juri != "SCI" and impot_revenu < 2000 and impot_revenu != 0:
                recommandations.append({
                    "titre": "Étudier bénéfice prime d'activité",
                    "detail": f"Impôt sur le revenu: {impot_revenu:.2f}€. Faible IR peut indiquer éligibilité à la prime d'activité.",
                    "formule": "forme_juri ≠ SCI ET impot_revenu < 2000 ET impot_revenu ≠ 0"
                })
            
            # 8. Augmentation rémunération de gérance
            if imposition == "IS" and forme_juri != "SCI" and resultat > 42500:
                recommandations.append({
                    "titre": "Étudier augmentation rémunération de gérance",
                    "detail": f"Imposition: IS, Résultat: {resultat:.2f}€. Augmenter la rémunération pourrait optimiser la fiscalité globale.",
                    "formule": "Imposition = IS ET forme_juri ≠ SCI ET resultat < 42500"
                })
            
            # 9. Passage TNS pour SAS avec masse salariale
            if forme_juri == "SAS" and m_salariale > 1 and imposition == "IS":
                recommandations.append({
                    "titre": "Étudier passage TNS",
                    "detail": f"Forme juridique: SAS, Masse salariale: {m_salariale:.2f}€, Imposition: IS. Le statut TNS pourrait réduire les charges sociales.",
                    "formule": "forme_juri = SAS ET m_salariale > 1 ET Imposition = IS"
                })
            
            # 10. Placement de trésorerie excédentaire
            if tresorerie > 100000:
                recommandations.append({
                    "titre": "Étudier placement de l'excédent de trésorerie",
                    "detail": f"Trésorerie: {tresorerie:.2f}€. L'excédent pourrait être placé pour générer des revenus supplémentaires.",
                    "formule": "trésorerie > 100000"
                })
            
            # Afficher les résultats
            self.afficher_analyse_dossier(nom, siret, dossier_dict, recommandations)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'analyse du dossier: {str(e)}")
            import traceback
            traceback.print_exc()

    def afficher_analyse_dossier(self, nom, siret, dossier_dict, recommandations):
        """Affiche l'analyse d'un dossier avec ses recommandations"""
        # Créer une fenêtre pour afficher l'analyse
        analyse_window = tk.Toplevel(self.root)
        analyse_window.title(f"Analyse du dossier {nom}")
        analyse_window.geometry("900x700")
        analyse_window.configure(bg="#f0f5f9")
        
        # Créer un cadre principal avec défilement
        main_frame = ttk.Frame(analyse_window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Titre
        title_label = tk.Label(
            main_frame, 
            text=f"Analyse du dossier {nom}", 
            font=("Helvetica", 16, "bold"), 
            bg="#f0f5f9"
        )
        title_label.pack(pady=(0, 20))
        
        # Créer un canvas pour permettre le défilement
        canvas = tk.Canvas(main_frame, bg='#f0f5f9', highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Section d'informations générales
        info_frame = ttk.LabelFrame(scrollable_frame, text="Informations générales")
        info_frame.pack(fill="x", padx=10, pady=10)
        
        # Champs importants à afficher
        important_fields = [
            ("SIRET", "siret"),
            ("Nom", "nom"),
            ("Forme juridique", "forme_juri"),
            ("Imposition", "Imposition"),
            ("Code NAF", "code_naf"),
            ("Chiffre d'affaires", "CA"),
            ("Résultat", "resultat"),
            ("Régime TVA", "Regime_de_TVA"),
            ("Emprunt", "emprunt"),
            ("Impôt sur le revenu", "impot_revenu"),
            ("Masse salariale", "m_salariale"),
            ("Trésorerie", "trésorerie")
        ]
        
        # Créer une grille pour les informations
        row = 0
        col = 0
        for label_text, field_key in important_fields:
            # Récupérer la valeur
            value = dossier_dict.get(field_key, "N/A")
            
            # Formater les valeurs numériques
            numeric_fields = ["CA", "resultat", "emprunt", "impot_revenu", "m_salariale", "trésorerie"]
            if field_key in numeric_fields:
                try:
                    value = float(value) if value not in (None, "N/A", "") else 0
                    value_text = f"{value:,.2f} €"
                except (ValueError, TypeError):
                    value_text = str(value)
            else:
                value_text = str(value)
            
            # Cadre pour cette information
            info_item_frame = ttk.Frame(info_frame)
            info_item_frame.grid(row=row, column=col, padx=10, pady=5, sticky="w")
            
            # Label pour le nom du champ
            label = ttk.Label(
                info_item_frame, 
                text=f"{label_text}: ",
                font=("Helvetica", 10, "bold"),
                width=15,
                anchor="e"
            )
            label.grid(row=0, column=0, sticky="e")
            
            # Label pour la valeur
            value_label = ttk.Label(
                info_item_frame, 
                text=value_text,
                font=("Helvetica", 10)
            )
            value_label.grid(row=0, column=1, sticky="w")
            
            # Passer à la colonne suivante ou retourner à la première colonne et ligne suivante
            col += 1
            if col > 2:  # 3 colonnes par ligne
                col = 0
                row += 1
        
        # Section des recommandations
        if recommandations:
            reco_frame = ttk.LabelFrame(scrollable_frame, text="Recommandations")
            reco_frame.pack(fill="x", padx=10, pady=10)
            
            # Cadre pour chaque recommandation
            for i, reco in enumerate(recommandations):
                # Créer un cadre pour cette recommandation
                reco_item_frame = ttk.Frame(reco_frame)
                reco_item_frame.pack(fill="x", padx=10, pady=5)
                
                # Titre avec numéro
                title_label = ttk.Label(
                    reco_item_frame, 
                    text=f"{i+1}. {reco['titre']}",
                    font=("Helvetica", 12, "bold")
                )
                title_label.pack(anchor="w", pady=(5, 2))
                
                # Détail de la recommandation
                detail_label = ttk.Label(
                    reco_item_frame, 
                    text=reco['detail'],
                    font=("Helvetica", 10),
                    wraplength=800,
                    justify="left"
                )
                detail_label.pack(anchor="w", padx=20, pady=(0, 5))
                
                # Bouton d'information pour afficher la formule
                info_button = ttk.Button(
                    reco_item_frame, 
                    text="Formule",
                    width=10,
                    command=lambda r=reco: messagebox.showinfo("Formule", r['formule'])
                )
                info_button.pack(anchor="w", padx=20, pady=(0, 5))
                
                # Séparateur
                ttk.Separator(reco_frame, orient="horizontal").pack(fill="x", padx=5, pady=5)
        else:
            # Message si aucune recommandation n'est trouvée
            no_reco_frame = ttk.LabelFrame(scrollable_frame, text="Recommandations")
            no_reco_frame.pack(fill="x", padx=10, pady=10)
            
            ttk.Label(
                no_reco_frame, 
                text="Aucune recommandation n'a été détectée pour ce dossier.",
                font=("Helvetica", 12),
                padding=20
            ).pack()
        
        # Bouton pour générer un rapport PDF pour ce dossier
        rapport_button = ttk.Button(
            scrollable_frame, 
            text="Générer rapport détaillé",
            command=lambda: self.generer_rapport_dossier(siret, dossier_dict, recommandations)
        )
        rapport_button.pack(pady=20)
        
        # Bouton pour fermer
        close_button = tk.Button(
            scrollable_frame, 
            text="Fermer",
            command=analyse_window.destroy,
            bg="#0f4c75", 
            fg="white",
            font=("Helvetica", 10, "bold"),
            padx=20, 
            pady=5
        )
        close_button.pack(pady=10)

    def generer_rapport_dossier(self, siret, dossier_dict, recommandations):
        """Génère un rapport détaillé pour un dossier spécifique"""
        try:
            # Vérifier si la bibliothèque openpyxl est disponible
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            # Récupérer le nom du dossier
            nom = dossier_dict.get('nom', siret)
            
            # Proposer à l'utilisateur de choisir l'emplacement du fichier
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"Rapport_{siret}.xlsx",
                title="Enregistrer le rapport détaillé"
            )
            
            if not file_path:
                return  # L'utilisateur a annulé
            
            # Créer un nouveau classeur Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Analyse Détaillée"
            
            # Définir les styles
            title_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
            header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
            normal_font = Font(name='Calibri', size=11)
            
            # Couleurs de fond
            header_fill = PatternFill(start_color="3282B8", end_color="3282B8", fill_type="solid")
            alt_row_fill = PatternFill(start_color="EEF4F9", end_color="EEF4F9", fill_type="solid")
            
            # Bordures
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Alignement
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            
            # Titre du rapport
            ws.merge_cells('A1:F1')
            cell = ws.cell(row=1, column=1, value=f"Analyse Détaillée - {nom}")
            cell.font = title_font
            cell.fill = header_fill
            cell.alignment = center_align
            
            # Section des informations générales
            ws.cell(row=3, column=1, value="Informations générales").font = header_font
            
            # En-têtes pour les informations
            ws.cell(row=4, column=1, value="Champ").font = header_font
            ws.cell(row=4, column=1).fill = header_fill
            ws.cell(row=4, column=1).border = thin_border
            
            ws.cell(row=4, column=2, value="Valeur").font = header_font
            ws.cell(row=4, column=2).fill = header_fill
            ws.cell(row=4, column=2).border = thin_border
            
            # Champs importants à afficher
            important_fields = [
                ("SIRET", "siret"),
                ("Nom", "nom"),
                ("Forme juridique", "forme_juri"),
                ("Imposition", "Imposition"),
                ("Code NAF", "code_naf"),
                ("Chiffre d'affaires", "CA"),
                ("Résultat", "resultat"),
                ("Régime TVA", "Regime_de_TVA"),
                ("Emprunt", "emprunt"),
                ("Impôt sur le revenu", "impot_revenu"),
                ("Masse salariale", "m_salariale"),
                ("Trésorerie", "trésorerie")
            ]
            
            # Remplir les informations générales
            for i, (label_text, field_key) in enumerate(important_fields, 5):
                ws.cell(row=i, column=1, value=label_text).font = normal_font
                ws.cell(row=i, column=1).border = thin_border
                
                # Récupérer et formater la valeur
                value = dossier_dict.get(field_key, "N/A")
                numeric_fields = ["CA", "resultat", "emprunt", "impot_revenu", "m_salariale", "trésorerie"]
                
                if field_key in numeric_fields:
                    try:
                        value = float(value) if value not in (None, "N/A", "") else 0
                        cell = ws.cell(row=i, column=2, value=value)
                        cell.number_format = '#,##0.00 €'
                    except (ValueError, TypeError):
                        ws.cell(row=i, column=2, value=str(value))
                else:
                    ws.cell(row=i, column=2, value=str(value))
                
                ws.cell(row=i, column=2).font = normal_font
                ws.cell(row=i, column=2).border = thin_border
                
                # Mise en forme alternée
                if i % 2 == 0:
                    ws.cell(row=i, column=1).fill = alt_row_fill
                    ws.cell(row=i, column=2).fill = alt_row_fill
            
            # Section des recommandations
            row = i + 2
            
            ws.cell(row=row, column=1, value="Recommandations").font = header_font
            row += 1
            
            # En-têtes pour les recommandations
            headers = ["N°", "Recommandation", "Détails", "Formule"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header).font = header_font
                ws.cell(row=row, column=col).fill = header_fill
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).alignment = center_align
            
            row += 1
            
            # Ajuster les largeurs de colonnes
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 30
            
            # Remplir les recommandations
            if recommandations:
                for i, reco in enumerate(recommandations, 1):
                    # Numéro
                    ws.cell(row=row, column=1, value=i).font = normal_font
                    ws.cell(row=row, column=1).border = thin_border
                    
                    # Titre
                    ws.cell(row=row, column=2, value=reco['titre']).font = normal_font
                    ws.cell(row=row, column=2).border = thin_border
                    
                    # Détail
                    cell = ws.cell(row=row, column=3, value=reco['detail'])
                    cell.font = normal_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # Formule
                    cell = ws.cell(row=row, column=4, value=reco['formule'])
                    cell.font = normal_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # Mise en forme alternée
                    if i % 2 == 0:
                        for col in range(1, 5):
                            ws.cell(row=row, column=col).fill = alt_row_fill
                    
                    row += 1
            else:
                # Message si aucune recommandation
                ws.merge_cells(f'A{row}:D{row}')
                cell = ws.cell(row=row, column=1, value="Aucune recommandation n'a été détectée pour ce dossier.")
                cell.font = normal_font
                cell.alignment = center_align
            
            # Sauvegarder le fichier
            wb.save(file_path)
            
            messagebox.showinfo("Rapport généré", f"Le rapport détaillé a été généré avec succès à l'emplacement:\n{file_path}")
            
            # Demander à l'utilisateur s'il souhaite ouvrir le fichier
            if messagebox.askyesno("Ouvrir le rapport?", "Voulez-vous ouvrir le rapport maintenant?"):
                try:
                    import os
                    import subprocess
                    if os.name == 'nt':  # Windows
                        os.startfile(file_path)
                    elif os.name == 'posix':  # macOS et Linux
                        if sys.platform == 'darwin':  # macOS
                            subprocess.call(['open', file_path])
                        else:  # Linux
                            subprocess.call(['xdg-open', file_path])
                except Exception as e:
                    messagebox.showerror("Erreur", f"Impossible d'ouvrir le fichier: {str(e)}")
        except ImportError:
            messagebox.showerror("Erreur", "La bibliothèque openpyxl n'est pas installée. Veuillez l'installer pour utiliser cette fonctionnalité.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur s'est produite lors de la génération du rapport: {str(e)}")
            import traceback
            traceback.print_exc()

# 1. D'abord, ajoutons le bouton dans la méthode setup_analyse_tab_ui:


    def setup_analyse_tab_ui(self):
        """Configure l'interface de l'onglet Analyse"""
        analyse_frame = ttk.Frame(self.analyse_tab)
        analyse_frame.pack(fill=tk.BOTH, expand=True)
        
        self.analyse_canvas = tk.Canvas(analyse_frame, bg='#f5f6fa', highlightthickness=0)
        self.analyse_scrollbar = ttk.Scrollbar(analyse_frame, orient="vertical", command=self.analyse_canvas.yview)
        self.analyse_scrollable_frame = ttk.Frame(self.analyse_canvas)

        self.analyse_canvas.configure(yscrollcommand=self.analyse_scrollbar.set)
        self.analyse_canvas.pack(side="left", fill="both", expand=True)
        self.analyse_scrollbar.pack(side="right", fill="y")

        self.analyse_canvas_frame = self.analyse_canvas.create_window(
            (0, 0), window=self.analyse_scrollable_frame, anchor="nw", width=780
        )

        # Ajout d'une nouvelle section pour les données comptables
        self.create_section("Données Comptables pour l'Analyse", [
            ("🔄 Mise à jour des données comptables pour l'analyse", self.mettre_a_jour_donnees_comptables)
        ], self.analyse_scrollable_frame)



        # Section pour l'analyse des dossiers
        self.create_section("Analyse des données", [
            ("📊 Structure du portefeuille client", self.show_recommendations),
            ("🗺️ Carte des clients", self.show_client_map),
            ("💰 Clients par Trésorerie", self.afficher_clients_par_tresorerie),
            ("💵 Clients par CA", self.afficher_clients_par_ca),
            ("💸 Clients par Impôt sur le Revenu", self.afficher_clients_par_impot),
            ("🏦 Clients par Endettement", self.afficher_clients_par_endettement),
            ("👥 Clients par Charges Salariales", self.afficher_clients_par_charges_salariales)
        ], self.analyse_scrollable_frame)

        # Configuration similaire aux autres onglets pour la gestion du défilement
        self.analyse_scrollable_frame.bind("<Configure>", 
            lambda e: self.analyse_canvas.configure(scrollregion=self.analyse_canvas.bbox("all"))
        )
        self.analyse_canvas.bind("<Configure>", 
            lambda e: self.analyse_canvas.itemconfig(self.analyse_canvas_frame, width=e.width-4)
        )
        
        # Étiquette de statut pour l'onglet Analyse
        self.analyse_status_label = tk.Label(
            self.analyse_scrollable_frame, 
            text="Prêt pour analyser vos données", 
            font=("Helvetica", 10), 
            bg="#f5f6fa"
        )
        self.analyse_status_label.pack(pady=10)
        



    def mettre_a_jour_donnees_comptables(self):
        """Met à jour les données comptables et les indicateurs financiers pour l'analyse"""
        try:
            # Importer le module datetime
            import datetime
            
            # Vérifier si la base de données existe
            if not os.path.exists(self.db_path):
                messagebox.showerror("Erreur", "Base de données introuvable.")
                return
                
            # Connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # Vérifier et ajouter les colonnes manquantes à la table indicateurs
            self.verifier_ajouter_colonnes_indicateurs(cursor, conn)
            
            # Récupérer tous les SIRET uniques
            cursor.execute("SELECT DISTINCT siret FROM rapport_combine WHERE siret != 'moyenne'")
            sirets = cursor.fetchall()
            
            # Nombre total de SIRET pour la barre de progression
            total_sirets = len(sirets)
            
            if total_sirets == 0:
                messagebox.showinfo("Information", "Aucun SIRET trouvé dans la table rapport_combine.\nVeuillez d'abord importer des données.")
                conn.close()
                return
            
            # Initialiser la barre de progression
            self.progress_bar['value'] = 0
            self.progress_label.config(text=f"Mise à jour des données comptables: 0/{total_sirets}")
            self.root.update_idletasks()
            
            # Compteurs pour le suivi
            sirets_mis_a_jour = 0
            sirets_ajoutes = 0
            
            # Traiter chaque SIRET
            for idx, (siret,) in enumerate(sirets):
                # Récupérer les identifiants existants
                cursor.execute("""
                    SELECT nom, nomUniteLegale, prenom1UniteLegale
                    FROM indicateurs
                    WHERE siret = ?
                """, (siret,))
                
                identite = cursor.fetchone()
                nom, nomUniteLegale, prenom1UniteLegale = None, None, None
                
                if identite:
                    nom, nomUniteLegale, prenom1UniteLegale = identite
                
                # Structure pour stocker les valeurs des comptes
                comptes_valeurs = {}
                
                # 1. Calculer les soldes pour chaque type de compte
                mappings_comptes = [
                    ("compte_101_130", "(compte LIKE '10%' OR compte LIKE '11%' OR compte LIKE '12%' OR compte LIKE '13%')"),
                    ("compte_16", "compte LIKE '16%'"),
                    ("compte_20_27", "(compte LIKE '20%' OR compte LIKE '21%' OR compte LIKE '22%' OR compte LIKE '23%' OR compte LIKE '24%' OR compte LIKE '25%' OR compte LIKE '26%' OR compte LIKE '27%')"),
                    ("compte_28_29", "(compte LIKE '28%' OR compte LIKE '29%')"),
                    ("compte_3", "compte LIKE '3%'"),
                    ("compte_40", "compte LIKE '40%'"),
                    ("compte_41", "compte LIKE '41%'"),
                    ("compte_44", "compte LIKE '44%'"),
                    ("compte_46_47", "(compte LIKE '46%' OR compte LIKE '47%')"),
                    ("compte_5", "compte LIKE '5%'"),
                    ("compte_60", "compte LIKE '60%'"),
                    ("compte_61_62", "(compte LIKE '61%' OR compte LIKE '62%')"),
                    ("compte_63", "compte LIKE '63%'"),
                    ("compte_64", "compte LIKE '64%'"),
                    ("compte_65", "compte LIKE '65%'"),
                    ("compte_66", "compte LIKE '66%'"),
                    ("compte_67", "compte LIKE '67%'"),
                    ("compte_68", "compte LIKE '68%'"),
                    ("compte_69", "compte LIKE '69%'"),
                    ("compte_70", "compte LIKE '70%'"),
                    ("compte_71_72", "(compte LIKE '71%' OR compte LIKE '72%')"),
                    ("compte_74", "compte LIKE '74%'"),
                    ("compte_75", "compte LIKE '75%'"),
                    ("compte_76", "compte LIKE '76%'"),
                    ("compte_77", "compte LIKE '77%'"),
                    ("compte_791", "compte LIKE '791%'")  # Ajout spécifique pour le compte 791
                ]
                
                # Calculer chaque solde de compte
                for nom_compte, condition in mappings_comptes:
                    cursor.execute(f"""
                    SELECT SUM(CAST(REPLACE(solde, ',', '.') AS REAL))
                    FROM rapport_combine
                    WHERE siret = ? AND {condition}
                    """, (siret,))
                    comptes_valeurs[nom_compte] = cursor.fetchone()[0] or 0
                
                # 2. Inverser les produits (comptes de classe 7) pour qu'ils apparaissent en positif
                for compte in ["compte_70", "compte_71_72", "compte_74", "compte_75", "compte_76", "compte_77", "compte_791"]:
                    if compte in comptes_valeurs:
                        comptes_valeurs[compte] = -comptes_valeurs[compte]
                
                # 3. Calculer les indicateurs financiers
                
                # Chiffre d'affaires (comptes 70)
                ca = comptes_valeurs.get("compte_70", 0)
                
                # Achats (comptes 60)
                achats = comptes_valeurs.get("compte_60", 0)
                
                # Charges de personnel (comptes 64)
                charges_personnel = comptes_valeurs.get("compte_64", 0)
                
                # Frais financiers (comptes 66)
                frais_financiers = comptes_valeurs.get("compte_66", 0)
                
                # Dotations aux amortissements (comptes 68)
                dotations = comptes_valeurs.get("compte_68", 0)
                
                # Dettes financières (comptes 16)
                dettes_financieres = comptes_valeurs.get("compte_16", 0)
                
                # Capitaux propres (comptes 10 à 14)
                capitaux_propres = comptes_valeurs.get("compte_101_130", 0)
                
                # Créances clients (comptes 41)
                creances_clients = comptes_valeurs.get("compte_41", 0)
                
                # Dettes fournisseurs (comptes 40)
                dettes_fournisseurs = comptes_valeurs.get("compte_40", 0)
                
                # Trésorerie (comptes 5)
                tresorerie = comptes_valeurs.get("compte_5", 0)
                
                # Compte 791 (transferts de charges)
                transfert_charges = comptes_valeurs.get("compte_791", 0)
                
                # Total des charges (somme des comptes 6)
                total_charges = sum([
                    comptes_valeurs.get("compte_60", 0),
                    comptes_valeurs.get("compte_61_62", 0),
                    comptes_valeurs.get("compte_63", 0),
                    comptes_valeurs.get("compte_64", 0),
                    comptes_valeurs.get("compte_65", 0),
                    comptes_valeurs.get("compte_66", 0),
                    comptes_valeurs.get("compte_67", 0),
                    comptes_valeurs.get("compte_68", 0),
                    comptes_valeurs.get("compte_69", 0)
                ])
                
                # Total des produits (somme des comptes 7, sauf 791 qui est traité séparément)
                total_produits = sum([
                    comptes_valeurs.get("compte_70", 0),
                    comptes_valeurs.get("compte_71_72", 0),
                    comptes_valeurs.get("compte_74", 0),
                    comptes_valeurs.get("compte_75", 0),
                    comptes_valeurs.get("compte_76", 0),
                    comptes_valeurs.get("compte_77", 0)
                ])
                
                # EBITDA = CA - Achats - Charges variables - Charges fixes (hors dotations)
                ebitda = total_produits - (total_charges - dotations)
                
                # Capacité d'Autofinancement (CAF)
                caf = ebitda - frais_financiers
                
                # Capacité de remboursement
                capacite_remboursement = abs(dettes_financieres / caf) if caf != 0 else 999
                
                # Ratio de liquidité
                actif_circulant = creances_clients + tresorerie
                passif_circulant = abs(dettes_fournisseurs)
                ratio_liquidite = actif_circulant / passif_circulant if passif_circulant != 0 else 999
                
                # Ratio d'autonomie financière
                total_passif = abs(capitaux_propres) + abs(dettes_financieres) + abs(dettes_fournisseurs)
                ratio_autonomie = (abs(capitaux_propres) / total_passif) if total_passif != 0 else 0
                
                # Frais financiers / CA
                frais_financiers_ca = (frais_financiers / ca) * 100 if ca != 0 else 0
                
                # Délai clients et fournisseurs en jours
                delai_client_jours = (creances_clients / ca) * 360 if ca != 0 else 0
                delai_fournisseur_jours = (abs(dettes_fournisseurs) / achats) * 360 if achats != 0 else 0
                
                # BFR d'exploitation
                bfr_exploitation = creances_clients - abs(dettes_fournisseurs)
                
                # Trésorerie nette
                tresorerie_nette = tresorerie
                
                # Taux de marge EBITDA
                taux_marge_ebitda = (ebitda / ca) * 100 if ca != 0 else 0
                
                # Calcul du Z-Score d'Altman (version simplifiée)
                # Estimation du total actif
                total_actif = abs(comptes_valeurs.get("compte_20_27", 0)) + abs(comptes_valeurs.get("compte_3", 0)) + abs(creances_clients) + abs(tresorerie)
                
                if total_actif > 0:
                    # Composantes du Z-Score
                    x1 = bfr_exploitation / total_actif
                    benefices_non_repartis = capitaux_propres * 0.5  # Estimation
                    x2 = benefices_non_repartis / total_actif
                    x3 = ebitda / total_actif
                    total_dettes = abs(dettes_financieres) + abs(dettes_fournisseurs)
                    x4 = abs(capitaux_propres) / total_dettes if total_dettes > 0 else 3
                    x5 = ca / total_actif
                    
                    # Calcul final du Z-Score (modèle adapté aux PME françaises)
                    z_score = 0.717*x1 + 0.847*x2 + 3.107*x3 + 0.420*x4 + 0.998*x5
                else:
                    z_score = 0
                
                # Détermination du niveau de risque
                if z_score > 2.9:
                    niveau_risque = "Risque faible"
                elif z_score > 1.23:
                    niveau_risque = "Zone grise"
                else:
                    niveau_risque = "Risque élevé"
                
                # Indice de fragilité
                indice_fragilite = (creances_clients - abs(dettes_fournisseurs)) / ca * 360 if ca != 0 else 999
                
                # Résultat (avec prise en compte du compte 791)
                resultat = total_produits - total_charges + transfert_charges
                
                # 4. Vérifier si le SIRET existe déjà dans la table indicateurs
                cursor.execute("SELECT 1 FROM indicateurs WHERE siret = ?", (siret,))
                siret_existe = cursor.fetchone() is not None
                
                if siret_existe:
                    # Préparation de la requête de mise à jour
                    set_clauses = []
                    params = []
                    
                    # Ajouter les comptes
                    for nom_compte, _ in mappings_comptes:
                        set_clauses.append(f"{nom_compte} = ?")
                        params.append(comptes_valeurs.get(nom_compte, 0))
                    
                    # Ajouter les indicateurs
                    indicateurs = {
                        "z_score_altman": z_score,
                        "indice_fragilite": indice_fragilite,
                        "capacite_remboursement": capacite_remboursement,
                        "ratio_liquidite": ratio_liquidite,
                        "ratio_autonomie": ratio_autonomie,
                        "frais_financiers_ca": frais_financiers_ca,
                        "delai_client_jours": delai_client_jours,
                        "delai_fournisseur_jours": delai_fournisseur_jours,
                        "bfr_exploitation": bfr_exploitation,
                        "tresorerie_nette": tresorerie_nette,
                        "niveau_risque": niveau_risque,
                        "ebitda": ebitda,
                        "caf": caf,
                        "taux_marge_ebitda": taux_marge_ebitda,
                        "date_calcul_comptes": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "CA": ca,
                        "resultat": resultat
                    }
                    
                    for nom, valeur in indicateurs.items():
                        set_clauses.append(f"{nom} = ?")
                        params.append(valeur)
                    
                    # Ajouter le SIRET à la fin des paramètres
                    params.append(siret)
                    
                    # Construire et exécuter la requête de mise à jour
                    query = f"UPDATE indicateurs SET {', '.join(set_clauses)} WHERE siret = ?"
                    cursor.execute(query, params)
                    sirets_mis_a_jour += 1
                    
                else:
                    # Création d'un nouvel enregistrement
                    noms_colonnes = [nom_compte for nom_compte, _ in mappings_comptes]
                    noms_colonnes.extend([
                        "z_score_altman", "indice_fragilite", "capacite_remboursement", 
                        "ratio_liquidite", "ratio_autonomie", "frais_financiers_ca",
                        "delai_client_jours", "delai_fournisseur_jours", "bfr_exploitation",
                        "tresorerie_nette", "niveau_risque", "ebitda", "caf", "taux_marge_ebitda",
                        "date_calcul_comptes", "CA", "resultat", "siret", "nom", "nomUniteLegale", "prenom1UniteLegale"
                    ])
                    
                    valeurs = [comptes_valeurs.get(nom_compte, 0) for nom_compte, _ in mappings_comptes]
                    valeurs.extend([
                        z_score, indice_fragilite, capacite_remboursement, 
                        ratio_liquidite, ratio_autonomie, frais_financiers_ca,
                        delai_client_jours, delai_fournisseur_jours, bfr_exploitation,
                        tresorerie_nette, niveau_risque, ebitda, caf, taux_marge_ebitda,
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ca, 
                        resultat, siret, nom, nomUniteLegale, prenom1UniteLegale
                    ])
                    
                    # Construire et exécuter la requête d'insertion
                    placeholders = ", ".join(["?"] * len(valeurs))
                    query = f"INSERT INTO indicateurs ({', '.join(noms_colonnes)}) VALUES ({placeholders})"
                    cursor.execute(query, valeurs)
                    sirets_ajoutes += 1
                
                # Mise à jour de la barre de progression
                self.progress_bar['value'] = (idx + 1) / total_sirets * 100
                self.progress_label.config(text=f"Mise à jour des données comptables: {idx+1}/{total_sirets}")
                
                # Mise à jour périodique de l'interface pour ne pas la bloquer
                if idx % 5 == 0:
                    self.root.update_idletasks()
                
                # Validation périodique pour éviter de perdre le travail en cas d'erreur
                if idx % 20 == 0:
                    conn.commit()
            
            # Validation finale des modifications
            conn.commit()
            conn.close()
            
            # Mise à jour du statut
            self.analyse_status_label.config(text=f"Données comptables mises à jour: {sirets_mis_a_jour} SIRET(s) mis à jour, {sirets_ajoutes} SIRET(s) ajoutés")
            
            messagebox.showinfo("Succès", f"Mise à jour des données comptables terminée avec succès!\n\n{sirets_mis_a_jour} SIRET(s) mis à jour\n{sirets_ajoutes} SIRET(s) ajoutés")
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de la mise à jour des données comptables:\n\n{str(e)}")
            if 'conn' in locals() and conn:
                conn.close()

    def verifier_ajouter_colonnes_indicateurs(self, cursor, conn):
        """Vérifie et ajoute les colonnes manquantes à la table indicateurs"""
        try:
            # Liste des comptes principaux à ajouter
            comptes_principaux = [
                "compte_101_130 REAL",      # Capital et réserves
                "compte_16 REAL",           # Emprunts et dettes financières
                "compte_20_27 REAL",        # Immobilisations
                "compte_28_29 REAL",        # Amortissements et provisions
                "compte_3 REAL",            # Stocks
                "compte_40 REAL",           # Fournisseurs
                "compte_41 REAL",           # Clients
                "compte_44 REAL",           # État, taxes
                "compte_46_47 REAL",        # Débiteurs et créditeurs divers
                "compte_5 REAL",            # Trésorerie
                "compte_60 REAL",           # Achats
                "compte_61_62 REAL",        # Services extérieurs
                "compte_63 REAL",           # Impôts et taxes
                "compte_64 REAL",           # Charges de personnel
                "compte_65 REAL",           # Autres charges de gestion
                "compte_66 REAL",           # Charges financières
                "compte_67 REAL",           # Charges exceptionnelles
                "compte_68 REAL",           # Dotations aux amortissements
                "compte_69 REAL",           # Impôts sur les bénéfices
                "compte_70 REAL",           # Ventes
                "compte_71_72 REAL",        # Production
                "compte_74 REAL",           # Subventions d'exploitation
                "compte_75 REAL",           # Autres produits de gestion
                "compte_76 REAL",           # Produits financiers
                "compte_77 REAL",           # Produits exceptionnels
                "compte_791 REAL",          # Transferts de charges
                "date_calcul_comptes TEXT"  # Date de calcul des comptes
            ]
            
            # Liste des nouveaux indicateurs financiers
            indicateurs_financiers = [
                "z_score_altman REAL",              # Indicateur global de santé financière
                "indice_fragilite REAL",            # Mesure le financement du cycle d'exploitation
                "capacite_remboursement REAL",      # Dette financière nette / CAF
                "ratio_liquidite REAL",             # Actif circulant / Passif circulant
                "ratio_autonomie REAL",             # Fonds propres / Total passif
                "frais_financiers_ca REAL",         # Frais financiers / CA
                "delai_client_jours REAL",          # (Créances clients / CA) * 360
                "delai_fournisseur_jours REAL",     # (Dettes fournisseurs / Achats) * 360
                "bfr_exploitation REAL",            # BFR lié à l'exploitation
                "tresorerie_nette REAL",            # Disponibilités - Concours bancaires
                "niveau_risque TEXT",               # Évaluation textuelle du niveau de risque
                "ebitda REAL",                      # Résultat avant intérêts, impôts, dépréciation et amortissement
                "caf REAL",                         # Capacité d'autofinancement
                "taux_marge_ebitda REAL"            # EBITDA / CA
            ]
            
            # Combinaison des deux listes
            toutes_colonnes = comptes_principaux + indicateurs_financiers
            
            # Récupérer les colonnes existantes
            cursor.execute("PRAGMA table_info(indicateurs)")
            colonnes_existantes = [col[1] for col in cursor.fetchall()]
            
            # Compteurs pour le suivi
            colonnes_ajoutees = 0
            colonnes_deja_existantes = 0
            
            # Ajouter chaque colonne si elle n'existe pas déjà
            for definition_colonne in toutes_colonnes:
                # Extraire le nom de la colonne (tout ce qui précède le premier espace)
                nom_colonne = definition_colonne.split(' ')[0]
                
                if nom_colonne not in colonnes_existantes:
                    try:
                        cursor.execute(f"ALTER TABLE indicateurs ADD COLUMN {definition_colonne}")
                        colonnes_ajoutees += 1
                        print(f"Colonne ajoutée: {definition_colonne}")
                    except sqlite3.OperationalError as e:
                        print(f"Erreur lors de l'ajout de la colonne {nom_colonne}: {str(e)}")
                else:
                    colonnes_deja_existantes += 1
                    print(f"La colonne {nom_colonne} existe déjà")
            
            # Valider les modifications
            conn.commit()
            
            # Afficher un message si des colonnes ont été ajoutées
            if colonnes_ajoutees > 0:
                print(f"\nRésumé de l'opération:")
                print(f"- {colonnes_ajoutees} nouvelles colonnes ajoutées")
                print(f"- {colonnes_deja_existantes} colonnes déjà existantes")
                print(f"- {len(toutes_colonnes)} colonnes au total")
            
            return True
            
        except Exception as e:
            print(f"Erreur lors de la vérification/ajout des colonnes: {str(e)}")
            return False

    def afficher_clients_par_charges_salariales(self):
        """
        Affiche les clients triés par niveau de charges salariales décroissant
        (du plus élevé au plus faible)
        """
        try:
            # Connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            
            # Requête pour récupérer les clients triés par charges salariales
            query = """
            SELECT 
                siret, 
                COALESCE(nom, siret) as nom, 
                m_salariale, 
                code_naf, 
                forme_juri, 
                ville,
                Adresse_eMail,
                nomUniteLegale,
                prenom1UniteLegale,
                CA
            FROM indicateurs 
            WHERE siret != 'moyenne' AND m_salariale IS NOT NULL
            ORDER BY m_salariale DESC
            """
            
            # Exécuter la requête
            df = pd.read_sql_query(query, conn)
            conn.close()
            
            # Vérifier s'il y a des données
            if df.empty:
                messagebox.showinfo("Information", "Aucune donnée de charges salariales disponible.")
                return
            
            # Créer la fenêtre de résultats
            results_window = tk.Toplevel(self.root)
            results_window.title("Clients par Charges Salariales")
            results_window.geometry(f"{self.root.winfo_screenwidth()}x{self.root.winfo_screenheight()}")
            results_window.configure(bg="#f0f5f9")
            
            # Cadre principal
            main_frame = ttk.Frame(results_window)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            # Titre
            title_label = tk.Label(
                main_frame, 
                text="Clients triés par Charges Salariales (du plus élevé au plus faible)", 
                font=("Helvetica", 18, "bold"),
                bg="#f0f5f9"
            )
            title_label.pack(pady=(0, 20))
            
            # Cadre pour le tableau avec défilement
            result_frame = ttk.Frame(main_frame)
            result_frame.pack(fill=tk.BOTH, expand=True)
            
            # Barres de défilement
            scrollbar_y = ttk.Scrollbar(result_frame, orient="vertical")
            scrollbar_x = ttk.Scrollbar(result_frame, orient="horizontal")
            
            # Créer le tableau
            columns = ["SIRET", "Nom", "Charges Salariales (€)", "Chiffre d'Affaires (€)", "Ratio Charges/CA (%)", "Code NAF", "Forme Juridique", "Ville", "Adresse eMail", "Nom Unité Légale", "Prénom Unité Légale"]
            tree = ttk.Treeview(
                result_frame, 
                columns=columns, 
                show='headings', 
                yscrollcommand=scrollbar_y.set, 
                xscrollcommand=scrollbar_x.set
            )
            
            # Configurer les styles pour une police noire
            style = ttk.Style()
            style.configure("Treeview", 
                            foreground="black", 
                            background="white", 
                            fieldbackground="white")
            style.configure("Treeview.Heading", 
                            foreground="black", 
                            font=('Helvetica', 10, 'bold'))
            
            # Configurer les colonnes
            for col in columns:
                tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(tree, c, False))
                tree.column(col, width=100, anchor="center")
            
            # Configurer les barres de défilement
            scrollbar_y.config(command=tree.yview)
            scrollbar_y.pack(side="right", fill="y")
            
            scrollbar_x.config(command=tree.xview)
            scrollbar_x.pack(side="bottom", fill="x")
            
            # Insérer les données
            for _, row in df.iterrows():
                # Formater les charges salariales
                charges = float(row['m_salariale']) if pd.notna(row['m_salariale']) else 0
                charges_formate = f"{charges:,.2f}" if charges != 0 else "0.00"
                
                # Formater le CA
                ca = float(row['CA']) if pd.notna(row['CA']) and row['CA'] != '' else 0
                ca_formate = f"{ca:,.2f}" if ca > 0 else "N/A"
                
                # Calculer et formater le ratio charges/CA
                ratio = (charges / ca * 100) if ca > 0 else float('inf')
                ratio_formate = f"{ratio:.2f}" if ca > 0 else "N/A"
                
                # Insérer la ligne
                tree.insert("", "end", values=[
                    row['siret'], 
                    row['nom'], 
                    charges_formate, 
                    ca_formate,
                    ratio_formate,
                    row['code_naf'], 
                    row['forme_juri'], 
                    row['ville'],
                    row['Adresse_eMail'] if pd.notna(row['Adresse_eMail']) else "N/A",
                    row['nomUniteLegale'] if pd.notna(row['nomUniteLegale']) else "N/A",
                    row['prenom1UniteLegale'] if pd.notna(row['prenom1UniteLegale']) else "N/A"
                ])
            
            # Placer le treeview
            tree.pack(fill="both", expand=True)
            
            # Boutons d'action
            button_frame = tk.Frame(main_frame, bg="#f0f5f9")
            button_frame.pack(pady=10)
            
            # Bouton d'export CSV
            def export_to_csv():
                """Exporte les données du treeview en CSV"""
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".csv", 
                    filetypes=[("CSV files", "*.csv")]
                )
                
                if not file_path:
                    return
                
                try:
                    # Récupérer les données du treeview
                    rows = [tree.item(item)["values"] for item in tree.get_children()]
                    
                    # Écrire dans un fichier CSV
                    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                        csv_writer = csv.writer(csvfile, delimiter=';')
                        # Écrire les en-têtes
                        csv_writer.writerow(columns)
                        # Écrire les données
                        csv_writer.writerows(rows)
                    
                    messagebox.showinfo("Succès", f"Données exportées avec succès dans {file_path}")
                except Exception as e:
                    messagebox.showerror("Erreur", f"Erreur lors de l'export : {str(e)}")
            
            # Bouton d'export CSV
            export_csv_btn = tk.Button(
                button_frame, 
                text="Exporter en CSV", 
                command=export_to_csv,
                bg="#3282b8",
                fg="white"
            )
            export_csv_btn.pack(side=tk.LEFT, padx=10)
            
            # Bouton de fermeture
            close_btn = tk.Button(
                button_frame, 
                text="Fermer", 
                command=results_window.destroy,
                bg="#0f4c75",
                fg="white"
            )
            close_btn.pack(side=tk.LEFT, padx=10)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue : {str(e)}")

    def afficher_clients_par_endettement(self):
        """
        Affiche les clients triés par niveau d'endettement décroissant
        (du plus endetté au moins endetté, en valeur absolue)
        """
        try:
            # Connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            
            # Requête pour récupérer les clients triés par emprunt
            # Note: On utilise ABS() pour trier par valeur absolue car les emprunts sont négatifs
            # Et ORDER BY ABS(emprunt) DESC pour aller du plus endetté au moins endetté
            query = """
            SELECT 
                siret, 
                COALESCE(nom, siret) as nom, 
                emprunt, 
                code_naf, 
                forme_juri, 
                ville,
                Adresse_eMail,
                nomUniteLegale,
                prenom1UniteLegale,
                CA
            FROM indicateurs 
            WHERE siret != 'moyenne' AND emprunt IS NOT NULL
            ORDER BY ABS(emprunt) DESC
            """
            
            # Exécuter la requête
            df = pd.read_sql_query(query, conn)
            conn.close()
            
            # Vérifier s'il y a des données
            if df.empty:
                messagebox.showinfo("Information", "Aucune donnée d'endettement disponible.")
                return
            
            # Créer la fenêtre de résultats
            results_window = tk.Toplevel(self.root)
            results_window.title("Clients par Niveau d'Endettement")
            results_window.geometry(f"{self.root.winfo_screenwidth()}x{self.root.winfo_screenheight()}")
            results_window.configure(bg="#f0f5f9")
            
            # Cadre principal
            main_frame = ttk.Frame(results_window)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            # Titre - Modifié pour refléter le tri inversé
            title_label = tk.Label(
                main_frame, 
                text="Clients triés par Niveau d'Endettement (du plus endetté au moins endetté)", 
                font=("Helvetica", 18, "bold"),
                bg="#f0f5f9"
            )
            title_label.pack(pady=(0, 20))
            
            # Cadre pour le tableau avec défilement
            result_frame = ttk.Frame(main_frame)
            result_frame.pack(fill=tk.BOTH, expand=True)
            
            # Barres de défilement
            scrollbar_y = ttk.Scrollbar(result_frame, orient="vertical")
            scrollbar_x = ttk.Scrollbar(result_frame, orient="horizontal")
            
            # Créer le tableau
            columns = ["SIRET", "Nom", "Endettement (€)", "Chiffre d'Affaires (€)", "Ratio Dette/CA (%)", "Code NAF", "Forme Juridique", "Ville", "Adresse eMail", "Nom Unité Légale", "Prénom Unité Légale"]
            tree = ttk.Treeview(
                result_frame, 
                columns=columns, 
                show='headings', 
                yscrollcommand=scrollbar_y.set, 
                xscrollcommand=scrollbar_x.set
            )
            
            # Configurer les styles pour une police noire
            style = ttk.Style()
            style.configure("Treeview", 
                            foreground="black", 
                            background="white", 
                            fieldbackground="white")
            style.configure("Treeview.Heading", 
                            foreground="black", 
                            font=('Helvetica', 10, 'bold'))
            
            # Configurer les colonnes
            for col in columns:
                tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(tree, c, False))
                tree.column(col, width=100, anchor="center")
            
            # Configurer les barres de défilement
            scrollbar_y.config(command=tree.yview)
            scrollbar_y.pack(side="right", fill="y")
            
            scrollbar_x.config(command=tree.xview)
            scrollbar_x.pack(side="bottom", fill="x")
            
            # Insérer les données
            for _, row in df.iterrows():
                # Formater l'endettement (valeur absolue car les emprunts sont négatifs)
                montant_dette = abs(float(row['emprunt'])) if pd.notna(row['emprunt']) else 0
                endettement_formate = f"{montant_dette:,.2f}" if montant_dette > 0 else "0.00"
                
                # Formater le CA
                ca = float(row['CA']) if pd.notna(row['CA']) and row['CA'] != '' else 0
                ca_formate = f"{ca:,.2f}" if ca > 0 else "N/A"
                
                # Calculer et formater le ratio dette/CA
                ratio = (montant_dette / ca * 100) if ca > 0 else float('inf')
                ratio_formate = f"{ratio:.2f}" if ca > 0 else "N/A"
                
                # Insérer la ligne
                tree.insert("", "end", values=[
                    row['siret'], 
                    row['nom'], 
                    endettement_formate, 
                    ca_formate,
                    ratio_formate,
                    row['code_naf'], 
                    row['forme_juri'], 
                    row['ville'],
                    row['Adresse_eMail'] if pd.notna(row['Adresse_eMail']) else "N/A",
                    row['nomUniteLegale'] if pd.notna(row['nomUniteLegale']) else "N/A",
                    row['prenom1UniteLegale'] if pd.notna(row['prenom1UniteLegale']) else "N/A"
                ])
            
            # Placer le treeview
            tree.pack(fill="both", expand=True)
            
            # Boutons d'action
            button_frame = tk.Frame(main_frame, bg="#f0f5f9")
            button_frame.pack(pady=10)
            
            # Bouton d'export CSV
            def export_to_csv():
                """Exporte les données du treeview en CSV"""
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".csv", 
                    filetypes=[("CSV files", "*.csv")]
                )
                
                if not file_path:
                    return
                
                try:
                    # Récupérer les données du treeview
                    rows = [tree.item(item)["values"] for item in tree.get_children()]
                    
                    # Écrire dans un fichier CSV
                    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                        csv_writer = csv.writer(csvfile, delimiter=';')
                        # Écrire les en-têtes
                        csv_writer.writerow(columns)
                        # Écrire les données
                        csv_writer.writerows(rows)
                    
                    messagebox.showinfo("Succès", f"Données exportées avec succès dans {file_path}")
                except Exception as e:
                    messagebox.showerror("Erreur", f"Erreur lors de l'export : {str(e)}")
            
            # Bouton d'export CSV
            export_csv_btn = tk.Button(
                button_frame, 
                text="Exporter en CSV", 
                command=export_to_csv,
                bg="#3282b8",
                fg="white"
            )
            export_csv_btn.pack(side=tk.LEFT, padx=10)
            
            # Bouton de fermeture
            close_btn = tk.Button(
                button_frame, 
                text="Fermer", 
                command=results_window.destroy,
                bg="#0f4c75",
                fg="white"
            )
            close_btn.pack(side=tk.LEFT, padx=10)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue : {str(e)}")

    def afficher_clients_par_impot(self):
        """
        Affiche les clients triés par niveau d'impôt sur le revenu décroissant
        """
        try:
            # Connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            
            # Requête pour récupérer les clients triés par impôt sur le revenu
            query = """
            SELECT 
                siret, 
                COALESCE(nom, siret) as nom, 
                impot_revenu, 
                code_naf, 
                forme_juri, 
                ville,
                Adresse_eMail,
                nomUniteLegale,
                prenom1UniteLegale
            FROM indicateurs 
            WHERE siret != 'moyenne' AND impot_revenu IS NOT NULL
            ORDER BY impot_revenu DESC
            """
            
            # Exécuter la requête
            df = pd.read_sql_query(query, conn)
            conn.close()
            
            # Vérifier s'il y a des données
            if df.empty:
                messagebox.showinfo("Information", "Aucune donnée d'impôt sur le revenu disponible.")
                return
            
            # Créer la fenêtre de résultats
            results_window = tk.Toplevel(self.root)
            results_window.title("Clients par Niveau d'Impôt sur le Revenu")
            results_window.geometry(f"{self.root.winfo_screenwidth()}x{self.root.winfo_screenheight()}")
            results_window.configure(bg="#f0f5f9")
            
            # Cadre principal
            main_frame = ttk.Frame(results_window)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            # Titre
            title_label = tk.Label(
                main_frame, 
                text="Clients triés par Niveau d'Impôt sur le Revenu", 
                font=("Helvetica", 18, "bold"),
                bg="#f0f5f9"
            )
            title_label.pack(pady=(0, 20))
            
            # Cadre pour le tableau avec défilement
            result_frame = ttk.Frame(main_frame)
            result_frame.pack(fill=tk.BOTH, expand=True)
            
            # Barres de défilement
            scrollbar_y = ttk.Scrollbar(result_frame, orient="vertical")
            scrollbar_x = ttk.Scrollbar(result_frame, orient="horizontal")
            
            # Créer le tableau
            columns = ["SIRET", "Nom", "Impôt sur le Revenu (€)", "Code NAF", "Forme Juridique", "Ville", "Adresse eMail", "Nom Unité Légale", "Prénom Unité Légale"]
            tree = ttk.Treeview(
                result_frame, 
                columns=columns, 
                show='headings', 
                yscrollcommand=scrollbar_y.set, 
                xscrollcommand=scrollbar_x.set
            )
            
            # Configurer les styles pour une police noire
            style = ttk.Style()
            style.configure("Treeview", 
                            foreground="black", 
                            background="white", 
                            fieldbackground="white")
            style.configure("Treeview.Heading", 
                            foreground="black", 
                            font=('Helvetica', 10, 'bold'))
            
            # Configurer les colonnes
            for col in columns:
                tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(tree, c, False))
                tree.column(col, width=100, anchor="center")
            
            # Configurer les barres de défilement
            scrollbar_y.config(command=tree.yview)
            scrollbar_y.pack(side="right", fill="y")
            
            scrollbar_x.config(command=tree.xview)
            scrollbar_x.pack(side="bottom", fill="x")
            
            # Insérer les données
            for _, row in df.iterrows():
                # Formater l'impôt sur le revenu
                impot_formate = f"{float(row['impot_revenu']):,.2f}" if pd.notna(row['impot_revenu']) and row['impot_revenu'] != '' else "N/A"
                
                # Insérer la ligne
                tree.insert("", "end", values=[
                    row['siret'], 
                    row['nom'], 
                    impot_formate, 
                    row['code_naf'], 
                    row['forme_juri'], 
                    row['ville'],
                    row['Adresse_eMail'] if pd.notna(row['Adresse_eMail']) else "N/A",
                    row['nomUniteLegale'] if pd.notna(row['nomUniteLegale']) else "N/A",
                    row['prenom1UniteLegale'] if pd.notna(row['prenom1UniteLegale']) else "N/A"
                ])
            
            # Placer le treeview
            tree.pack(fill="both", expand=True)
            
            # Boutons d'action
            button_frame = tk.Frame(main_frame, bg="#f0f5f9")
            button_frame.pack(pady=10)
            
            # Bouton d'export CSV
            def export_to_csv():
                """Exporte les données du treeview en CSV"""
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".csv", 
                    filetypes=[("CSV files", "*.csv")]
                )
                
                if not file_path:
                    return
                
                try:
                    # Récupérer les données du treeview
                    rows = [tree.item(item)["values"] for item in tree.get_children()]
                    
                    # Écrire dans un fichier CSV
                    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                        csv_writer = csv.writer(csvfile, delimiter=';')
                        # Écrire les en-têtes
                        csv_writer.writerow(columns)
                        # Écrire les données
                        csv_writer.writerows(rows)
                    
                    messagebox.showinfo("Succès", f"Données exportées avec succès dans {file_path}")
                except Exception as e:
                    messagebox.showerror("Erreur", f"Erreur lors de l'export : {str(e)}")
            
            # Bouton d'export CSV
            export_csv_btn = tk.Button(
                button_frame, 
                text="Exporter en CSV", 
                command=export_to_csv,
                bg="#3282b8",
                fg="white"
            )
            export_csv_btn.pack(side=tk.LEFT, padx=10)
            
            # Bouton de fermeture
            close_btn = tk.Button(
                button_frame, 
                text="Fermer", 
                command=results_window.destroy,
                bg="#0f4c75",
                fg="white"
            )
            close_btn.pack(side=tk.LEFT, padx=10)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue : {str(e)}")

    def afficher_clients_par_ca(self):
        """
        Affiche les clients triés par niveau de chiffre d'affaires décroissant
        """
        try:
            # Connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            
            # Requête pour récupérer les clients triés par CA
            query = """
            SELECT 
                siret, 
                COALESCE(nom, siret) as nom, 
                CA, 
                code_naf, 
                forme_juri, 
                ville,
                Adresse_eMail,
                nomUniteLegale,
                prenom1UniteLegale
            FROM indicateurs 
            WHERE siret != 'moyenne' AND CA IS NOT NULL
            ORDER BY CA DESC
            """
            
            # Exécuter la requête
            df = pd.read_sql_query(query, conn)
            conn.close()
            
            # Vérifier s'il y a des données
            if df.empty:
                messagebox.showinfo("Information", "Aucune donnée de chiffre d'affaires disponible.")
                return
            
            # Créer la fenêtre de résultats
            results_window = tk.Toplevel(self.root)
            results_window.title("Clients par Niveau de Chiffre d'Affaires")
            results_window.geometry(f"{self.root.winfo_screenwidth()}x{self.root.winfo_screenheight()}")
            results_window.configure(bg="#f0f5f9")
            
            # Cadre principal
            main_frame = ttk.Frame(results_window)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            # Titre
            title_label = tk.Label(
                main_frame, 
                text="Clients triés par Niveau de Chiffre d'Affaires", 
                font=("Helvetica", 18, "bold"),
                bg="#f0f5f9"
            )
            title_label.pack(pady=(0, 20))
            
            # Cadre pour le tableau avec défilement
            result_frame = ttk.Frame(main_frame)
            result_frame.pack(fill=tk.BOTH, expand=True)
            
            # Barres de défilement
            scrollbar_y = ttk.Scrollbar(result_frame, orient="vertical")
            scrollbar_x = ttk.Scrollbar(result_frame, orient="horizontal")
            
            # Créer le tableau
            columns = ["SIRET", "Nom", "Chiffre d'Affaires (€)", "Code NAF", "Forme Juridique", "Ville", "Adresse eMail", "Nom Unité Légale", "Prénom Unité Légale"]
            tree = ttk.Treeview(
                result_frame, 
                columns=columns, 
                show='headings', 
                yscrollcommand=scrollbar_y.set, 
                xscrollcommand=scrollbar_x.set
            )
            
            # Configurer les styles pour une police noire
            style = ttk.Style()
            style.configure("Treeview", 
                            foreground="black", 
                            background="white", 
                            fieldbackground="white")
            style.configure("Treeview.Heading", 
                            foreground="black", 
                            font=('Helvetica', 10, 'bold'))
            
            # Configurer les colonnes
            for col in columns:
                tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(tree, c, False))
                tree.column(col, width=100, anchor="center")
            
            # Configurer les barres de défilement
            scrollbar_y.config(command=tree.yview)
            scrollbar_y.pack(side="right", fill="y")
            
            scrollbar_x.config(command=tree.xview)
            scrollbar_x.pack(side="bottom", fill="x")
            
            # Insérer les données
            for _, row in df.iterrows():
                # Formater le CA
                ca_formate = f"{row['CA']:,.2f}" if pd.notna(row['CA']) else "N/A"
                
                # Insérer la ligne
                tree.insert("", "end", values=[
                    row['siret'], 
                    row['nom'], 
                    ca_formate, 
                    row['code_naf'], 
                    row['forme_juri'], 
                    row['ville'],
                    row['Adresse_eMail'] if pd.notna(row['Adresse_eMail']) else "N/A",
                    row['nomUniteLegale'] if pd.notna(row['nomUniteLegale']) else "N/A",
                    row['prenom1UniteLegale'] if pd.notna(row['prenom1UniteLegale']) else "N/A"
                ])
            
            # Placer le treeview
            tree.pack(fill="both", expand=True)
            
            # Boutons d'action
            button_frame = tk.Frame(main_frame, bg="#f0f5f9")
            button_frame.pack(pady=10)
            
            # Bouton d'export CSV
            def export_to_csv():
                """Exporte les données du treeview en CSV"""
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".csv", 
                    filetypes=[("CSV files", "*.csv")]
                )
                
                if not file_path:
                    return
                
                try:
                    # Récupérer les données du treeview
                    rows = [tree.item(item)["values"] for item in tree.get_children()]
                    
                    # Écrire dans un fichier CSV
                    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                        csv_writer = csv.writer(csvfile, delimiter=';')
                        # Écrire les en-têtes
                        csv_writer.writerow(columns)
                        # Écrire les données
                        csv_writer.writerows(rows)
                    
                    messagebox.showinfo("Succès", f"Données exportées avec succès dans {file_path}")
                except Exception as e:
                    messagebox.showerror("Erreur", f"Erreur lors de l'export : {str(e)}")
            
            # Bouton d'export CSV
            export_csv_btn = tk.Button(
                button_frame, 
                text="Exporter en CSV", 
                command=export_to_csv,
                bg="#3282b8",
                fg="white"
            )
            export_csv_btn.pack(side=tk.LEFT, padx=10)
            
            # Bouton de fermeture
            close_btn = tk.Button(
                button_frame, 
                text="Fermer", 
                command=results_window.destroy,
                bg="#0f4c75",
                fg="white"
            )
            close_btn.pack(side=tk.LEFT, padx=10)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue : {str(e)}")

    def afficher_clients_par_tresorerie(self):
        """
        Affiche les clients triés par niveau de trésorerie décroissant
        """
        try:
            # Connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            
            # Requête pour récupérer les clients triés par trésorerie
            query = """
            SELECT 
                siret, 
                COALESCE(nom, siret) as nom, 
                trésorerie, 
                code_naf, 
                forme_juri, 
                ville,
                Adresse_eMail,
                nomUniteLegale,
                prenom1UniteLegale
            FROM indicateurs 
            WHERE siret != 'moyenne' AND trésorerie IS NOT NULL
            ORDER BY trésorerie DESC
            """
            
            # Exécuter la requête
            df = pd.read_sql_query(query, conn)
            conn.close()
            
            # Vérifier s'il y a des données
            if df.empty:
                messagebox.showinfo("Information", "Aucune donnée de trésorerie disponible.")
                return
            
            # Créer la fenêtre de résultats
            results_window = tk.Toplevel(self.root)
            results_window.title("Clients par Niveau de Trésorerie")
            results_window.geometry(f"{self.root.winfo_screenwidth()}x{self.root.winfo_screenheight()}")
            results_window.configure(bg="#f0f5f9")
            
            # Cadre principal
            main_frame = ttk.Frame(results_window)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            # Titre
            title_label = tk.Label(
                main_frame, 
                text="Clients triés par Niveau de Trésorerie", 
                font=("Helvetica", 18, "bold"),
                bg="#f0f5f9"
            )
            title_label.pack(pady=(0, 20))
            
            # Cadre pour le tableau avec défilement
            result_frame = ttk.Frame(main_frame)
            result_frame.pack(fill=tk.BOTH, expand=True)
            
            # Barres de défilement
            scrollbar_y = ttk.Scrollbar(result_frame, orient="vertical")
            scrollbar_x = ttk.Scrollbar(result_frame, orient="horizontal")
            
            # Créer le tableau
            columns = ["SIRET", "Nom", "Trésorerie (€)", "Code NAF", "Forme Juridique", "Ville", "Adresse eMail", "Nom Unité Légale", "Prénom Unité Légale"]
            tree = ttk.Treeview(
                result_frame, 
                columns=columns, 
                show='headings', 
                yscrollcommand=scrollbar_y.set, 
                xscrollcommand=scrollbar_x.set
            )
            
            # Configurer les styles pour une police noire
            style = ttk.Style()
            style.configure("Treeview", 
                            foreground="black", 
                            background="white", 
                            fieldbackground="white")
            style.configure("Treeview.Heading", 
                            foreground="black", 
                            font=('Helvetica', 10, 'bold'))
            
            # Configurer les colonnes
            for col in columns:
                tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(tree, c, False))
                tree.column(col, width=100, anchor="center")
            
            # Configurer les barres de défilement
            scrollbar_y.config(command=tree.yview)
            scrollbar_y.pack(side="right", fill="y")
            
            scrollbar_x.config(command=tree.xview)
            scrollbar_x.pack(side="bottom", fill="x")
            
            # Insérer les données
            for _, row in df.iterrows():
                # Formater la trésorerie
                tresorerie_formate = f"{row['trésorerie']:,.2f}" if pd.notna(row['trésorerie']) else "N/A"
                
                # Insérer la ligne
                tree.insert("", "end", values=[
                    row['siret'], 
                    row['nom'], 
                    tresorerie_formate, 
                    row['code_naf'], 
                    row['forme_juri'], 
                    row['ville'],
                    row['Adresse_eMail'] if pd.notna(row['Adresse_eMail']) else "N/A",
                    row['nomUniteLegale'] if pd.notna(row['nomUniteLegale']) else "N/A",
                    row['prenom1UniteLegale'] if pd.notna(row['prenom1UniteLegale']) else "N/A"
                ])
            
            # Placer le treeview
            tree.pack(fill="both", expand=True)
            
            # Boutons d'action
            button_frame = tk.Frame(main_frame, bg="#f0f5f9")
            button_frame.pack(pady=10)
            
            # Bouton d'export CSV
            def export_to_csv():
                """Exporte les données du treeview en CSV"""
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".csv", 
                    filetypes=[("CSV files", "*.csv")]
                )
                
                if not file_path:
                    return
                
                try:
                    # Récupérer les données du treeview
                    rows = [tree.item(item)["values"] for item in tree.get_children()]
                    
                    # Écrire dans un fichier CSV
                    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                        csv_writer = csv.writer(csvfile, delimiter=';')
                        # Écrire les en-têtes
                        csv_writer.writerow(columns)
                        # Écrire les données
                        csv_writer.writerows(rows)
                    
                    messagebox.showinfo("Succès", f"Données exportées avec succès dans {file_path}")
                except Exception as e:
                    messagebox.showerror("Erreur", f"Erreur lors de l'export : {str(e)}")
            
            # Bouton d'export CSV
            export_csv_btn = tk.Button(
                button_frame, 
                text="Exporter en CSV", 
                command=export_to_csv,
                bg="#3282b8",
                fg="white"
            )
            export_csv_btn.pack(side=tk.LEFT, padx=10)
            
            # Bouton de fermeture
            close_btn = tk.Button(
                button_frame, 
                text="Fermer", 
                command=results_window.destroy,
                bg="#0f4c75",
                fg="white"
            )
            close_btn.pack(side=tk.LEFT, padx=10)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue : {str(e)}")



    def get_best_identifier(self, dossier):
        """
        Retourne le meilleur identifiant disponible pour un dossier selon la priorité:
        1. nom (raison sociale)
        2. nomUniteLegale + prenom1UniteLegale
        3. siret
        """
        # Vérifier si nous avons un dictionnaire ou une liste/tuple
        if isinstance(dossier, dict):
            # Cas d'un dictionnaire
            if dossier.get('nom') and str(dossier.get('nom')).strip():
                return str(dossier.get('nom'))
            elif dossier.get('nomUniteLegale'):
                nom_complet = str(dossier.get('nomUniteLegale', ''))
                if dossier.get('prenom1UniteLegale'):
                    nom_complet += ' ' + str(dossier.get('prenom1UniteLegale', ''))
                if nom_complet.strip():
                    return nom_complet
            return str(dossier.get('siret', 'Inconnu'))
        else:
            # Cas d'une liste ou tuple (dépend de l'index dans les résultats SQL)
            # Adapter les indices selon la structure de vos requêtes
            siret_idx = 0  # Indice du SIRET dans le tuple
            nom_idx = 1    # Indice du nom dans le tuple
            
            if len(dossier) > nom_idx and dossier[nom_idx] and str(dossier[nom_idx]).strip():
                return str(dossier[nom_idx])
            # Si nomUniteLegale est disponible dans les résultats, l'utiliser
            if len(dossier) > 2 and dossier[2]:  # Supposons que nomUniteLegale est à l'indice 2
                nom_complet = str(dossier[2])
                if len(dossier) > 3 and dossier[3]:  # Supposons que prenom1UniteLegale est à l'indice 3
                    nom_complet += ' ' + str(dossier[3])
                if nom_complet.strip():
                    return nom_complet
            return str(dossier[siret_idx])

    def show_recommendations(self):
        """Affiche des recommandations pour l'ensemble du portefeuille de clients"""
        try:
            # Vérifier si des données sont disponibles
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Vérifier si la table indicateurs existe et contient des données
            cursor.execute("SELECT COUNT(*) FROM indicateurs WHERE siret != 'moyenne'")
            count = cursor.fetchone()[0]
            
            if count == 0:
                messagebox.showinfo(
                    "Information", 
                    "Aucune donnée disponible. Veuillez d'abord importer des données FEC."
                )
                conn.close()
                return
            
            # Fenêtre pour afficher les recommandations
            rec_window = tk.Toplevel(self.root)
            rec_window.title("Analyse du Portefeuille Client")
            rec_window.geometry("900x700")
            rec_window.configure(bg="#f5f6fa")
            
            # Cadre principal avec défilement
            main_frame = ttk.Frame(rec_window)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
            
            # Titre
            title_label = ttk.Label(
                main_frame, 
                text="Analyse du Portefeuille Client", 
                font=("Arial", 18, "bold")
            )
            title_label.pack(pady=(0, 20), anchor="center")
            
            # Créer un cadre avec défilement pour les résultats
            canvas = tk.Canvas(main_frame, bg='#f5f6fa', highlightthickness=0)
            scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            
            canvas.configure(yscrollcommand=scrollbar.set)
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            
            # Analyser les données pour produire des recommandations globales
            # Synthèse globale du portefeuille
            cursor.execute("""
                SELECT 
                    COUNT(*) as nb_clients, 
                    AVG(CA) as ca_moyen, 
                    SUM(CA) as ca_total,
                    MIN(CA) as ca_min, 
                    MAX(CA) as ca_max,
                    SUM(produits) as total_produits,
                    SUM(charge) as total_charges
                FROM indicateurs 
                WHERE siret != 'moyenne'
            """)
            summary = cursor.fetchone()
            
            nb_clients, ca_moyen, ca_total, ca_min, ca_max, total_produits, total_charges = summary
            
            # Répartition par activité si disponible
            cursor.execute("""
                SELECT activite, COUNT(*) as nb_clients, SUM(CA) as ca_total
                FROM indicateurs
                WHERE siret != 'moyenne' AND activite IS NOT NULL
                GROUP BY activite
                ORDER BY ca_total DESC
            """)
            activites = cursor.fetchall()
            
            # Synthèse par taille d'entreprise
            cursor.execute("""
                SELECT 
                    CASE 
                        WHEN CA < 50000 THEN 'TPE (< 50K€)'
                        WHEN CA < 250000 THEN 'Petite entreprise (50K€ - 250K€)'
                        WHEN CA < 1000000 THEN 'Moyenne entreprise (250K€ - 1M€)'
                        ELSE 'Grande entreprise (> 1M€)'
                    END as categorie,
                    COUNT(*) as nb_clients,
                    SUM(CA) as ca_total
                FROM indicateurs
                WHERE siret != 'moyenne'
                GROUP BY categorie
                ORDER BY ca_total DESC
            """)
            categories = cursor.fetchall()
            
            # Afficher la synthèse globale
            summary_frame = ttk.LabelFrame(scrollable_frame, text="Synthèse du portefeuille client")
            summary_frame.pack(fill=tk.X, padx=10, pady=10, ipady=5)
            
            ttk.Label(
                summary_frame,
                text=f"Nombre de clients: {nb_clients}",
                font=("Helvetica", 11)
            ).pack(anchor="w", padx=10, pady=3)
            
            ttk.Label(
                summary_frame,
                text=f"Chiffre d'affaires total: {ca_total:,.2f} €",
                font=("Helvetica", 11)
            ).pack(anchor="w", padx=10, pady=3)
            
            ttk.Label(
                summary_frame,
                text=f"Chiffre d'affaires moyen: {ca_moyen:,.2f} €",
                font=("Helvetica", 11)
            ).pack(anchor="w", padx=10, pady=3)
            
            ttk.Label(
                summary_frame,
                text=f"Total des produits: {total_produits:,.2f} €",
                font=("Helvetica", 11)
            ).pack(anchor="w", padx=10, pady=3)
            
            ttk.Label(
                summary_frame,
                text=f"Total des charges: {total_charges:,.2f} €",
                font=("Helvetica", 11)
            ).pack(anchor="w", padx=10, pady=3)
            
            # Répartition par taille d'entreprise
            if categories:
                categories_frame = ttk.LabelFrame(scrollable_frame, text="Répartition par taille d'entreprise")
                categories_frame.pack(fill=tk.X, padx=10, pady=10, ipady=5)
                
                for categorie, nb, ca in categories:
                    ttk.Label(
                        categories_frame,
                        text=f"{categorie}: {nb} clients - CA total: {ca:,.2f} €",
                        font=("Helvetica", 11)
                    ).pack(anchor="w", padx=10, pady=3)
            
            # Répartition par activité
            if activites:
                activites_frame = ttk.LabelFrame(scrollable_frame, text="Répartition par activité")
                activites_frame.pack(fill=tk.X, padx=10, pady=10, ipady=5)
                
                for activite, nb, ca in activites[:10]:  # Limiter aux 10 premières activités
                    ttk.Label(
                        activites_frame,
                        text=f"{activite}: {nb} clients - CA total: {ca:,.2f} €",
                        font=("Helvetica", 11)
                    ).pack(anchor="w", padx=10, pady=3)
            
            # Note de conclusion
            ttk.Label(
                scrollable_frame,
                text="Pour des analyses plus détaillées, consultez les rapports individuels des dossiers.",
                wraplength=800,
                font=("Helvetica", 11, "italic")
            ).pack(anchor="w", pady=20, padx=10)
            
            # Bouton d'exportation
            export_button = ttk.Button(
                rec_window, 
                text="Exporter ce rapport", 
                command=lambda: self.export_recommendations_report(
                    summary, categories, activites, None, None, None, None
                )
            )
            export_button.pack(pady=20)
            
            conn.close()
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'analyse des données: {str(e)}")
            if 'conn' in locals() and conn:
                conn.close()

    def export_recommendations_report(self, summary, categories, activites, 
                                    concentration_text, margin_text, portfolio_text, 
                                    more_recommendations):
        """Exporte le rapport de recommandations au format Excel, sans la partie recommandations stratégiques"""
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx", 
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if not file_path:
                return
                
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Analyse Portefeuille"
            
            # Styles
            title_style = openpyxl.styles.NamedStyle(name="title")
            title_style.font = openpyxl.styles.Font(bold=True, size=14)
            
            header_style = openpyxl.styles.NamedStyle(name="header")
            header_style.font = openpyxl.styles.Font(bold=True)
            header_style.fill = openpyxl.styles.PatternFill(
                start_color="3282B8", end_color="3282B8", fill_type="solid"
            )
            header_style.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
            
            subheader_style = openpyxl.styles.NamedStyle(name="subheader")
            subheader_style.font = openpyxl.styles.Font(bold=True)
            subheader_style.fill = openpyxl.styles.PatternFill(
                start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
            )
            
            # Titre
            sheet.cell(row=1, column=1).value = "Analyse du Portefeuille Client"
            sheet.cell(row=1, column=1).style = title_style
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
            
            row = 3
            
            # Section 1: Synthèse globale
            sheet.cell(row=row, column=1).value = "1. Synthèse du portefeuille client"
            sheet.cell(row=row, column=1).style = header_style
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            row += 1
            
            nb_clients, ca_moyen, ca_total, marge_moyenne, resultat_total, ca_min, ca_max = summary
            
            sheet.cell(row=row, column=1).value = "Indicateur"
            sheet.cell(row=row, column=2).value = "Valeur"
            sheet.cell(row=row, column=1).style = subheader_style
            sheet.cell(row=row, column=2).style = subheader_style
            row += 1
            
            metrics = [
                ("Nombre de clients", nb_clients),
                ("Chiffre d'affaires total (€)", ca_total),
                ("Chiffre d'affaires moyen (€)", ca_moyen),
                ("Marge moyenne (%)", marge_moyenne),
                ("Résultat total (€)", resultat_total),
                ("CA minimum (€)", ca_min),
                ("CA maximum (€)", ca_max)
            ]
            
            for metric, value in metrics:
                sheet.cell(row=row, column=1).value = metric
                sheet.cell(row=row, column=2).value = value
                row += 1
            
            row += 2
            
            # Section 2: Répartition par taille d'entreprise
            if categories:
                sheet.cell(row=row, column=1).value = "2. Répartition par taille d'entreprise"
                sheet.cell(row=row, column=1).style = header_style
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                row += 1
                
                sheet.cell(row=row, column=1).value = "Catégorie"
                sheet.cell(row=row, column=2).value = "Nombre de clients"
                sheet.cell(row=row, column=3).value = "CA total (€)"
                sheet.cell(row=row, column=4).value = "Marge moyenne (%)"
                
                for col in range(1, 5):
                    sheet.cell(row=row, column=col).style = subheader_style
                row += 1
                
                for categorie, nb, ca, marge in categories:
                    sheet.cell(row=row, column=1).value = categorie
                    sheet.cell(row=row, column=2).value = nb
                    sheet.cell(row=row, column=3).value = ca
                    sheet.cell(row=row, column=4).value = marge
                    row += 1
                
                row += 2
            
            # Section 3: Répartition par activité
            if activites:
                sheet.cell(row=row, column=1).value = "3. Répartition par activité"
                sheet.cell(row=row, column=1).style = header_style
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                row += 1
                
                sheet.cell(row=row, column=1).value = "Activité"
                sheet.cell(row=row, column=2).value = "Nombre de clients"
                sheet.cell(row=row, column=3).value = "CA total (€)"
                sheet.cell(row=row, column=4).value = "Marge moyenne (%)"
                
                for col in range(1, 5):
                    sheet.cell(row=row, column=col).style = subheader_style
                row += 1
                
                for activite, nb, ca, marge in activites[:10]:
                    sheet.cell(row=row, column=1).value = activite
                    sheet.cell(row=row, column=2).value = nb
                    sheet.cell(row=row, column=3).value = ca
                    sheet.cell(row=row, column=4).value = marge
                    row += 1
                
                row += 2
            
            # Ajuster les largeurs des colonnes
            column_widths = [30, 20, 20, 20, 20]
            for i, width in enumerate(column_widths, 1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
            # Enregistrer le fichier
            workbook.save(file_path)
            messagebox.showinfo("Succès", "Le rapport d'analyse a été exporté avec succès.")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'exportation: {str(e)}")

    def setup_main_tab_ui(self):
        """Configure l'interface de l'onglet principal"""
        main_frame = ttk.Frame(self.main_tab)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        self.canvas = tk.Canvas(main_frame, bg='#f5f6fa', highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        self.canvas_frame = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw", width=780)

        self.create_section("Gestion de la base de données", [
            ("📁 Sélection dossier FEC", self.select_fec_directory_or_files),
            ("💿 Importer dans la base de données", self.update_database),
            ("📊 Création des indicateurs", self.create_indicators),
            ("📤 Export de la base de données", self.export_database)
        ], self.scrollable_frame)
        
        self.create_section("Import des noms et code NAF", [
            ("⚙️ Procédure", self.show_procedure),
            ("📄 Création de fichier avec les SIRET", self.extract_siret),
            ("📥 Import des fichiers dans votre base de données", self.import_nom)
        ], self.scrollable_frame)
        
        self.create_section("Recherche croisée", [
            ("🔍 Recherche Croisée Comptable", self.recherche_croisee_comptable),
            ("📈 Recherche Croisée Indicateur", self.recherche_croisee_indicateur)
        ], self.scrollable_frame)

        self.scrollable_frame.bind("<Configure>", self.on_frame_configure)
        self.canvas.bind("<Configure>", self.on_canvas_configure)
        self.root.bind("<MouseWheel>", self.on_mousewheel)

        self.selected_path_label = tk.Label(self.scrollable_frame, text="Aucun fichier ou dossier sélectionné", font=("Helvetica", 10), bg="#f5f6fa")
        self.selected_path_label.pack(pady=10)

        self.progress_frame = ttk.Frame(self.scrollable_frame)
        self.progress_frame.pack(fill=tk.X, padx=10, pady=5)
        self.progress_bar = ttk.Progressbar(self.progress_frame, length=300, mode='determinate')
        self.progress_bar.pack(side=tk.LEFT)
        self.progress_label = tk.Label(self.progress_frame, text="", bg="#f5f6fa")
        self.progress_label.pack(side=tk.LEFT, padx=5)

    def setup_extra_tab_ui(self):
        """Configure l'interface de l'onglet Extra-Comptable"""
        extra_frame = ttk.Frame(self.extra_tab)
        extra_frame.pack(fill=tk.BOTH, expand=True)
        
        self.extra_canvas = tk.Canvas(extra_frame, bg='#f5f6fa', highlightthickness=0)
        self.extra_scrollbar = ttk.Scrollbar(extra_frame, orient="vertical", command=self.extra_canvas.yview)
        self.extra_scrollable_frame = ttk.Frame(self.extra_canvas)

        self.extra_canvas.configure(yscrollcommand=self.extra_scrollbar.set)
        self.extra_canvas.pack(side="left", fill="both", expand=True)
        self.extra_scrollbar.pack(side="right", fill="y")

        self.extra_canvas_frame = self.extra_canvas.create_window((0, 0), window=self.extra_scrollable_frame, anchor="nw", width=780)

        self.create_section("Données Extra-Comptables", [
            ("📝 Générer fichiers Excel (instructions)", self.generer_instructions_excel),
            ("⚡ Création automatique du fichier Excel", self.creer_fichier_excel_automatique),  # Nouveau bouton
            ("📊 Visualisation des données extra-comptables", self.visualiser_donnees_extra),
            ("📥 Import des données extra-comptables", self.importer_donnees_extra),
            ("🔍 Recherche avancée", self.recherche_avancee)
           
        ], self.extra_scrollable_frame)

        self.extra_scrollable_frame.bind("<Configure>", lambda e: self.extra_canvas.configure(scrollregion=self.extra_canvas.bbox("all")))
        self.extra_canvas.bind("<Configure>", lambda e: self.extra_canvas.itemconfig(self.extra_canvas_frame, width=e.width-4))
        
        self.extra_status_label = tk.Label(self.extra_scrollable_frame, text="Aucune donnée extra-comptable importée", font=("Helvetica", 10), bg="#f5f6fa")
        self.extra_status_label.pack(pady=10)

    def creer_fichier_excel_automatique(self):
        """
        Crée automatiquement un fichier Excel avec les colonnes requises
        et pré-remplit uniquement les SIRET depuis la base de données
        """
        try:
            # Vérifier si openpyxl est disponible
            import openpyxl
            
            # Ouvrir la connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Récupérer tous les SIRET de la table indicateurs
            cursor.execute("SELECT siret FROM indicateurs WHERE siret != 'moyenne' AND siret GLOB '[0-9]*'")
            sirets_data = cursor.fetchall()
            
            # Fermer la connexion
            conn.close()
            
            # Vérifier si des SIRET ont été trouvés
            if not sirets_data:
                messagebox.showinfo("Information", "Aucun SIRET n'a été trouvé dans la base de données.\nImportez d'abord des données FEC.")
                return
            
            # Créer un nouveau classeur Excel
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            
            # Définir les en-têtes des colonnes
            headers = [
                "Code du dossier", "Code NAF", "Imposition", "OGA", 
                "Regime_de_TVA", "Adresse_eMail", "coll_compta", "date_clot", 
                "siret", "Site", "code_postal", "forme_juri", 
                "responsable", "Adresse", "ville", "impot_revenu"
            ]
            
            # Ajouter les en-têtes à la première ligne
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)
                
                # Ajouter un remplissage de couleur pour les en-têtes
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="1E88E5", end_color="1E88E5", fill_type="solid"
                )
                cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
            
            # Pré-remplir uniquement les SIRET à partir de la ligne 2
            for row_idx, siret_row in enumerate(sirets_data, 2):
                siret = siret_row[0]
                
                # Ajouter le SIRET dans la colonne I (index 9)
                sheet.cell(row=row_idx, column=9).value = siret
                
                # Ne pas remplir la colonne Code du dossier
            
            # Ajuster la largeur des colonnes
            for col_idx in range(1, len(headers) + 1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
            
            # Proposer à l'utilisateur de sauvegarder le fichier
            default_path = self.get_resource_path('extra_comptable.xlsx')
            file_path = filedialog.asksaveasfilename(
                initialfile="extra_comptable.xlsx",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialdir=self.base_path
            )
            
            if not file_path:
                return  # L'utilisateur a annulé
            
            # Sauvegarder le fichier
            workbook.save(file_path)
            
            messagebox.showinfo(
                "Succès", 
                f"Le fichier a été créé avec succès avec {len(sirets_data)} SIRET(s).\n"
                f"Emplacement: {file_path}\n\n"
                "N'oubliez pas de compléter les autres colonnes avant d'importer."
            )
            
            # Demander si l'utilisateur veut ouvrir le fichier
            if messagebox.askyesno("Ouvrir le fichier", "Voulez-vous ouvrir le fichier Excel maintenant?"):
                try:
                    if os.name == 'nt':  # Windows
                        os.startfile(file_path)
                    elif os.name == 'posix':  # macOS, Linux
                        if sys.platform == 'darwin':  # macOS
                            subprocess.call(['open', file_path])
                        else:  # Linux
                            subprocess.call(['xdg-open', file_path])
                except Exception as e:
                    messagebox.showerror("Erreur", f"Impossible d'ouvrir le fichier: {str(e)}")
        
        except ImportError:
            messagebox.showerror("Erreur", "La bibliothèque openpyxl n'est pas installée. Veuillez l'installer pour utiliser cette fonctionnalité.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de la création du fichier Excel: {str(e)}")
            import traceback
            traceback.print_exc()

    def generer_instructions_excel(self):
        """Affiche les instructions pour créer le fichier Excel des données extra-comptables avec scrollbar"""
        instructions_window = tk.Toplevel(self.root)
        instructions_window.title("Instructions pour créer extra_comptable.xlsx")
        instructions_window.geometry("700x600")  # Taille fixe pour une meilleure cohérence
        
        # Cadre principal
        main_frame = tk.Frame(instructions_window)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Créer un canvas avec scrollbar pour le contenu défilant
        canvas = tk.Canvas(main_frame, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')
        
        # Configurer le défilement
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Empaqueter le canvas et la scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Titre principal
        titre = tk.Label(
            scrollable_frame, 
            text="Instructions pour créer extra_comptable.xlsx", 
            font=("Helvetica", 16, "bold"),
            bg='white'
        )
        titre.pack(pady=(10, 20), padx=10)
        
        # Instructions importantes en rouge
        important_text = """⚠️ CONDITIONS IMPORTANTES DE FONCTIONNEMENT ⚠️"""
        important_label = tk.Label(
            scrollable_frame, 
            text=important_text,
            font=("Helvetica", 12, "bold"),
            fg="red",
            bg='white',
            justify=tk.LEFT
        )
        important_label.pack(fill='x', pady=(0, 5), padx=10, anchor='w')
        
        conditions_text = """- Le fichier DOIT être nommé exactement: extra_comptable.xlsx
    - Le fichier DOIT être placé dans le même dossier que l'outil
    - Le SIRET doit contenir EXACTEMENT 9 chiffres (sans espaces ni tirets)
    - IMPORTANT: Seuls les dossiers dont le FEC a déjà été importé seront mis à jour"""
        conditions_label = tk.Label(
            scrollable_frame, 
            text=conditions_text,
            font=("Helvetica", 11, "bold"),
            fg="red",
            bg='white',
            justify=tk.LEFT,
            anchor='w'
        )
        conditions_label.pack(fill='x', pady=(0, 20), padx=10, anchor='w')
        
        # Séparateur
        separator1 = ttk.Separator(scrollable_frame, orient='horizontal')
        separator1.pack(fill='x', padx=10, pady=(0, 15))
        
        # Instructions étape par étape
        instructions_title = tk.Label(
            scrollable_frame,
            text="📋 Création du fichier Excel pour données extra-comptables",
            font=("Helvetica", 14, "bold"),
            bg='white',
            fg='#2c3e50'
        )
        instructions_title.pack(pady=(0, 15), padx=10, anchor='w')
        
        # Étapes numérotées
        etapes = [
            ("1. Ouvrir un tableur", "Ouvrez Microsoft Excel ou un tableur compatible (LibreOffice Calc, Google Sheets)"),
            ("2. Créer un nouveau classeur", "Créez un nouveau fichier de type classeur Excel"),
            ("3. Entrer les en-têtes", "Entrez les en-têtes EXACTEMENT dans cet ordre :")
        ]
        
        for titre_etape, description in etapes:
            # Titre de l'étape
            etape_title = tk.Label(
                scrollable_frame,
                text=titre_etape,
                font=("Helvetica", 12, "bold"),
                bg='white',
                fg='#34495e'
            )
            etape_title.pack(anchor='w', padx=10, pady=(5, 2))
            
            # Description de l'étape
            etape_desc = tk.Label(
                scrollable_frame,
                text=description,
                font=("Helvetica", 11),
                bg='white',
                justify=tk.LEFT,
                anchor='w'
            )
            etape_desc.pack(anchor='w', padx=20, pady=(0, 8))
        
        # Liste des colonnes dans un cadre spécial
        colonnes_frame = tk.Frame(scrollable_frame, bg='#f8f9fa', relief='solid', bd=1)
        colonnes_frame.pack(fill='x', padx=10, pady=10)
        
        colonnes_title = tk.Label(
            colonnes_frame,
            text="En-têtes des colonnes (ordre obligatoire) :",
            font=("Helvetica", 11, "bold"),
            bg='#f8f9fa'
        )
        colonnes_title.pack(pady=(5, 10))
        
        # Liste des colonnes en deux colonnes pour économiser l'espace
        colonnes_text = """A1: Code du dossier          B1: Code NAF
    C1: Imposition                D1: OGA
    E1: Regime_de_TVA            F1: Adresse_eMail
    G1: coll_compta              H1: date_clot
    I1: siret                    J1: Site
    K1: code_postal              L1: forme_juri
    M1: responsable              N1: Adresse
    O1: ville                    P1: impot_revenu"""
        
        colonnes_label = tk.Label(
            colonnes_frame, 
            text=colonnes_text, 
            font=("Courier", 10),
            justify=tk.LEFT,
            anchor='w',
            bg='#f8f9fa'
        )
        colonnes_label.pack(padx=10, pady=(0, 10))
        
        # Étape 4
        etape4_title = tk.Label(
            scrollable_frame,
            text="4. Saisir les données",
            font=("Helvetica", 12, "bold"),
            bg='white',
            fg='#34495e'
        )
        etape4_title.pack(anchor='w', padx=10, pady=(15, 5))
        
        etape4_desc = tk.Label(
            scrollable_frame,
            text="Commencez à saisir vos données à partir de la ligne 2",
            font=("Helvetica", 11),
            bg='white'
        )
        etape4_desc.pack(anchor='w', padx=20, pady=(0, 10))
        
        # Conseils importants
        conseils_frame = tk.Frame(scrollable_frame, bg='#e8f5e8', relief='solid', bd=1)
        conseils_frame.pack(fill='x', padx=10, pady=15)
        
        conseils_title = tk.Label(
            conseils_frame,
            text="⚠️ Conseils importants :",
            font=("Helvetica", 11, "bold"),
            bg='#e8f5e8',
            fg='#27ae60'
        )
        conseils_title.pack(pady=(10, 5))
        
        conseils_text = """• Respectez strictement l'ordre des colonnes
    • Vérifiez l'orthographe des en-têtes
    • Les cellules vides sont autorisées
    • Sauvegardez au format .xlsx
    • Seules les données correspondant à des SIRET déjà importés via FEC seront mises à jour"""
        
        conseils_label = tk.Label(
            conseils_frame,
            text=conseils_text,
            font=("Helvetica", 10),
            justify=tk.LEFT,
            anchor='w',
            bg='#e8f5e8'
        )
        conseils_label.pack(padx=15, pady=(0, 10), anchor='w')
        
        # Exemple de format SIRET
        exemple_frame = tk.Frame(scrollable_frame, bg='#fff3cd', relief='solid', bd=1)
        exemple_frame.pack(fill='x', padx=10, pady=15)
        
        exemple_title = tk.Label(
            exemple_frame,
            text="Exemple de format SIRET correct :",
            font=("Helvetica", 11, "bold"),
            bg='#fff3cd',
            fg='#856404'
        )
        exemple_title.pack(pady=(10, 5))
        
        exemple_text = """Dans la colonne I (siret), mettre : 123456789 (9 chiffres exactement)
    NE PAS METTRE : 12345678900000 ou 123-456-789 ou 123 456 789"""
        exemple_label = tk.Label(
            exemple_frame,
            text=exemple_text,
            font=("Courier", 10, "bold"),
            fg="red",
            justify=tk.LEFT,
            anchor='w',
            bg='#fff3cd'
        )
        exemple_label.pack(padx=15, pady=(0, 10), anchor='w')
        
        # Cadre pour le bouton de fermeture (fixe en bas)
        button_frame = tk.Frame(instructions_window, bg='#f0f0f0')
        button_frame.pack(side='bottom', fill='x', pady=10)
        
        # Bouton de fermeture
        fermer_btn = tk.Button(
            button_frame, 
            text="Fermer", 
            command=instructions_window.destroy,
            bg="#0f4c75",
            fg="white",
            font=("Helvetica", 11, "bold"),
            padx=20, pady=10
        )
        fermer_btn.pack()
        
        # Activer le défilement avec la molette de la souris
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        # Lier l'événement de la molette à la fenêtre et au canvas
        def bind_mousewheel(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        def unbind_mousewheel(event):
            canvas.unbind_all("<MouseWheel>")
        
        canvas.bind('<Enter>', bind_mousewheel)
        canvas.bind('<Leave>', unbind_mousewheel)
        
        # S'assurer que la fenêtre est focalisée et que le contenu est visible
        instructions_window.focus_set()
        instructions_window.after(100, lambda: canvas.yview_moveto(0))


    def recherche_avancee(self):
        """Recherche croisée sur l'ensemble des colonnes de la table indicateurs"""
        # Créer une fenêtre pour la recherche
        recherche_window = tk.Toplevel(self.root)
        recherche_window.title("Recherche Avancée")
        recherche_window.geometry("1200x700")
        recherche_window.configure(bg="#f0f5f9")
        
        # Créer un frame avec défilement
        main_frame = ttk.Frame(recherche_window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        canvas = tk.Canvas(main_frame, bg='#f5f6fa', highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        canvas_frame = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas_frame, width=e.width-4))
        recherche_window.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        
        # Titre et instructions
        ttk.Label(scrollable_frame, text="Recherche sur toutes les colonnes", 
                  font=("Helvetica", 16, "bold")).pack(anchor="center", pady=(0, 5))
        ttk.Label(scrollable_frame, text="Remplissez un ou plusieurs champs pour filtrer les résultats", 
                  font=("Helvetica", 10)).pack(anchor="center", pady=(0, 15))
        
        # Connexion à la base de données
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Récupérer les colonnes de la table indicateurs
        cursor.execute("PRAGMA table_info(indicateurs)")
        columns_info = cursor.fetchall()
        
        # Champs à traiter spécifiquement comme numériques (min/max)
        numeric_fields = ['siret', 'CA', 'assurance', 'deplacement', 'loyer', 'CFE', 'TNS', 'publicité', 
                         'honoraires', 'banque', 'emprunt', 'm_salariale', 'produits', 'charge', 
                         'trésorerie', 'resultat', 'impot_revenu']
        
        # Créer des widgets pour chaque colonne
        self.search_widgets = {}
        
        # Organisation en 6 niveaux avec 6 champs par niveau
        fields_per_row = 6
        total_fields = len(columns_info)
        
        for level in range(6):  # 6 niveaux
            # Créer un cadre pour ce niveau
            level_frame = ttk.Frame(scrollable_frame)
            level_frame.pack(fill=tk.X, pady=10)
            
            # Configurer les poids des colonnes pour une distribution uniforme
            for i in range(fields_per_row):
                level_frame.columnconfigure(i, weight=1)
            
            # Ajouter les champs pour ce niveau
            for i in range(fields_per_row):
                idx = level * fields_per_row + i
                if idx < total_fields:
                    col_info = columns_info[idx]
                    col_name = col_info[1]
                    col_type = col_info[2]
                    
                    # Créer un frame pour ce champ
                    field_frame = ttk.Frame(level_frame)
                    field_frame.grid(row=0, column=i, padx=5, pady=5, sticky="nsew")
                    
                    # Label du champ
                    ttk.Label(field_frame, text=col_name, font=("Helvetica", 10, "bold")).pack(anchor="center")
                    
                    # Différents types de widgets selon le type de colonne
                    if col_name in numeric_fields or col_type.upper() in ('INTEGER', 'REAL', 'FLOAT', 'DECIMAL', 'NUMERIC'):
                        # Cadre pour min/max
                        min_max_frame = ttk.Frame(field_frame)
                        min_max_frame.pack(fill="x", pady=2)
                        
                        # Entrée minimum
                        min_frame = ttk.Frame(min_max_frame)
                        min_frame.pack(fill="x")
                        ttk.Label(min_frame, text="Min:", width=4).pack(side="left")
                        min_entry = ttk.Entry(min_frame, width=10)
                        min_entry.pack(side="left", fill="x", expand=True)
                        
                        # Entrée maximum
                        max_frame = ttk.Frame(min_max_frame)
                        max_frame.pack(fill="x", pady=2)
                        ttk.Label(max_frame, text="Max:", width=4).pack(side="left")
                        max_entry = ttk.Entry(max_frame, width=10)
                        max_entry.pack(side="left", fill="x", expand=True)
                        
                        self.search_widgets[col_name] = {'min': min_entry, 'max': max_entry}
                        
                    else:
                        # Pour les champs non numériques, créer un Combobox au lieu d'un Entry simple
                        # Récupérer les valeurs uniques pour ce champ
                        try:
                            cursor.execute(f"SELECT DISTINCT {col_name} FROM indicateurs WHERE {col_name} IS NOT NULL")
                            unique_values = [row[0] for row in cursor.fetchall() if row[0]]
                            
                            # Créer un combobox avec auto-complétion
                            combo = ttk.Combobox(field_frame, values=unique_values, width=15)
                            combo.pack(pady=2, fill="x")
                            
                            # Configuration pour auto-complétion
                            combo.bind('<KeyRelease>', lambda event, cb=combo, vals=unique_values: self.auto_complete(event, cb, vals))
                            
                            self.search_widgets[col_name] = combo
                        except Exception as e:
                            # En cas d'erreur, revenir à un champ texte simple
                            print(f"Erreur lors de la création du combobox pour {col_name}: {e}")
                            entry = ttk.Entry(field_frame, width=15)
                            entry.pack(pady=2, fill="x")
                            self.search_widgets[col_name] = entry
        
        # Boutons en bas
        buttons_frame = ttk.Frame(scrollable_frame)
        buttons_frame.pack(pady=20)
        
        ttk.Button(buttons_frame, text="Annuler", command=recherche_window.destroy).pack(side="right", padx=10)
        ttk.Button(buttons_frame, text="Rechercher", style="Accent.TButton", 
                command=lambda: self.execute_advanced_search(recherche_window)).pack(side="right", padx=10)
        
        conn.close()

    # Ajouter cette nouvelle méthode pour gérer l'auto-complétion
    def auto_complete(self, event, combobox, values):
        """Fonction d'auto-complétion pour les combobox"""
        current_text = combobox.get()
        if not current_text:
            combobox['values'] = values
            return
        
        # Recherche des valeurs correspondant à la saisie actuelle (insensible à la casse)
        matching_values = [val for val in values if str(val).lower().startswith(current_text.lower())]
        
        # Mettre à jour la liste déroulante
        combobox['values'] = matching_values
        
        # Si une seule valeur correspond exactement, la sélectionner
        if len(matching_values) == 1 and matching_values[0].lower() == current_text.lower():
            combobox.set(matching_values[0])
            combobox.selection_range(len(current_text), 'end')
        
        # Si au moins une valeur correspond, ouvrir la liste déroulante
        if matching_values and event.keysym != 'BackSpace':
            combobox.event_generate('<Down>')

    def execute_advanced_search(self, recherche_window):
        """Exécuter la recherche avancée avec les critères spécifiés"""
        query = "SELECT * FROM indicateurs WHERE 1=1"
        params = []
        
        # Construire la requête en fonction des critères de recherche
        for col_name, widget in self.search_widgets.items():
            if isinstance(widget, dict):  # Pour les colonnes numériques (min/max)
                min_value = widget['min'].get().strip()
                max_value = widget['max'].get().strip()
                
                if min_value:
                    try:
                        query += f" AND {col_name} >= ?"
                        params.append(float(min_value))
                    except ValueError:
                        pass  # Ignorer les valeurs non numériques
                
                if max_value:
                    try:
                        query += f" AND {col_name} <= ?"
                        params.append(float(max_value))
                    except ValueError:
                        pass  # Ignorer les valeurs non numériques
            elif isinstance(widget, ttk.Combobox):  # Pour les combobox
                value = widget.get().strip()
                if value:
                    query += f" AND {col_name} = ?"
                    params.append(value)
            else:  # Pour les champs de texte
                value = widget.get().strip()
                if value:
                    query += f" AND {col_name} LIKE ?"
                    params.append(f"%{value}%")
        
        # Exécuter la requête
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute(query, params)
            results = cursor.fetchall()
            
            # Récupérer les noms des colonnes
            cursor.execute("PRAGMA table_info(indicateurs)")
            columns = [info[1] for info in cursor.fetchall()]
            
            # Fermer la fenêtre de recherche et ouvrir la fenêtre de résultats
            recherche_window.destroy()
            self.display_advanced_search_results(results, columns)
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'exécution de la recherche: {str(e)}")
        finally:
            conn.close()

    def display_advanced_search_results(self, results, columns):
        """Affiche les résultats de la recherche avancée"""
        if not results:
            messagebox.showinfo("Information", "Aucun résultat ne correspond aux critères de recherche.")
            return
        
        # Créer une fenêtre pour les résultats
        results_window = tk.Toplevel(self.root)
        results_window.title("Résultats de la Recherche Avancée")
        results_window.geometry(f"{self.root.winfo_screenwidth()}x{self.root.winfo_screenheight()}")
        
        # Créer un cadre principal
        main_frame = ttk.Frame(results_window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Ajouter une étiquette d'information
        ttk.Label(main_frame, text=f"{len(results)} résultat(s) trouvé(s)", font=("Helvetica", 12, "bold")).pack(anchor="w", pady=(0, 10))
        
        # Créer un cadre pour le tableau avec défilement
        result_frame = ttk.Frame(main_frame)
        result_frame.pack(fill=tk.BOTH, expand=True)
        
        # Ajouter des barres de défilement
        scrollbar_y = ttk.Scrollbar(result_frame, orient="vertical")
        scrollbar_x = ttk.Scrollbar(result_frame, orient="horizontal")
        
        # Créer le tableau des résultats
        tree = ttk.Treeview(result_frame, columns=columns, show='headings', 
                          yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        # Configurer le style pour texte noir au lieu de gris
        self.style.configure("Treeview", foreground="black", background="white", rowheight=25)
        self.style.configure("Treeview.Heading", foreground="black", font=('Helvetica', 10, 'bold'))
        
        # Configurer les en-têtes de colonnes
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100, anchor="center")
        
        # Configurer les barres de défilement
        scrollbar_y.config(command=tree.yview)
        scrollbar_y.pack(side="right", fill="y")
        
        scrollbar_x.config(command=tree.xview)
        scrollbar_x.pack(side="bottom", fill="x")
        
        tree.pack(fill="both", expand=True)
        
        # Insérer les données
        for row in results:
            tree.insert("", "end", values=row)
        
        # Cadre pour les boutons
        button_frame = ttk.Frame(results_window)
        button_frame.pack(pady=10)
        
        # Ajouter un bouton d'exportation
        export_button = ttk.Button(button_frame, text="Exporter en CSV", style="Accent.TButton",
                                 command=lambda: self.export_advanced_search_to_csv(tree, columns))
        export_button.pack(side="left", padx=10)

    def export_advanced_search_to_csv(self, tree, columns):
        """Exporte les résultats de la recherche avancée en CSV"""
        file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        
        if not file_path:
            return
        
        rows = [tree.item(item)["values"] for item in tree.get_children()]
        
        try:
            with open(file_path, mode="w", newline='', encoding="utf-8") as file:
                writer = csv.writer(file, delimiter=';')
                writer.writerow(columns)
                writer.writerows(rows)
            
            messagebox.showinfo("Succès", "Les données ont été exportées avec succès en CSV.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'exportation en CSV: {str(e)}")

    def on_mousewheel(self, event):
        """Gestion améliorée du défilement qui inclut l'onglet Analyse"""
        current_tab = self.notebook.index(self.notebook.select())
        if current_tab == 0:  # Onglet principal
            self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        elif current_tab == 1:  # Onglet extra-comptable
            self.extra_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        elif current_tab == 2:  # Onglet analyse
            self.analyse_canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def create_section(self, title, buttons, parent_frame):
        frame = ttk.Frame(parent_frame)
        frame.pack(fill=tk.X, padx=10, pady=5)
        
        label = tk.Label(frame, text=title, font=('Arial', 12, 'bold'), bg='#f5f6fa')
        label.pack(anchor=tk.W)
        
        separator = ttk.Separator(frame, orient=tk.HORIZONTAL)
        separator.pack(fill=tk.X, pady=5)
        
        for text, command in buttons:
            btn = CustomButton(frame, text=text, command=command)
            btn.pack(fill=tk.X, pady=2)

    def visualiser_donnees_extra(self):
        """Ouvre un explorateur de fichiers puis affiche les données extra-comptables"""
        try:
            # Chemin du fichier Excel fixe (dans le même répertoire que l'application)
            file_path = self.get_resource_path('extra_comptable.xlsx')
            
            # Vérifier si le fichier existe
            if not os.path.exists(file_path):
                # Si le fichier n'existe pas, demander à l'utilisateur d'en sélectionner un
                file_path = filedialog.askopenfilename(
                    filetypes=[
                        ("Fichiers Excel", "*.xlsx *.xls"),
                        ("Tous les fichiers", "*.*")
                    ],
                    title="Ouvrir le fichier extra-comptable"
                )
                if not file_path:
                    return
            
            try:
                # Utiliser un bloc with pour s'assurer que les ressources sont fermées
                with pd.ExcelFile(file_path) as xls:
                    sheet_names = xls.sheet_names
                    
                    if len(sheet_names) == 1:
                        # S'il n'y a qu'une seule feuille, l'utiliser directement
                        df = pd.read_excel(file_path, sheet_name=sheet_names[0])
                        self.display_extra_file_results(df, os.path.basename(file_path))
                    else:
                        # S'il y a plusieurs feuilles, demander laquelle utiliser
                        sheet_dialog = tk.Toplevel(self.root)
                        sheet_dialog.title("Sélection de la feuille")
                        sheet_dialog.geometry("300x200")
                        
                        ttk.Label(sheet_dialog, text="Veuillez choisir la feuille à visualiser:").pack(pady=10)
                        
                        sheet_var = tk.StringVar(value=sheet_names[0])
                        sheet_combo = ttk.Combobox(sheet_dialog, textvariable=sheet_var, values=sheet_names)
                        sheet_combo.pack(pady=10, padx=20, fill="x")
                        
                        def confirm_sheet():
                            sheet_name = sheet_var.get()
                            sheet_dialog.destroy()
                            try:
                                df = pd.read_excel(file_path, sheet_name=sheet_name)
                                self.display_extra_file_results(df, os.path.basename(file_path))
                            except Exception as e:
                                messagebox.showerror("Erreur", f"Impossible de lire la feuille Excel: {str(e)}")
                        
                        ttk.Button(sheet_dialog, text="Confirmer", command=confirm_sheet).pack(pady=10)
                        sheet_dialog.transient(self.root)
                        sheet_dialog.grab_set()
                        self.root.wait_window(sheet_dialog)
                        
            except Exception as e:
                messagebox.showerror("Erreur", f"Impossible d'ouvrir le fichier Excel: {str(e)}")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue: {str(e)}")

    def display_extra_file_results(self, df, filename):
        """Affiche les résultats d'un fichier Excel dans une nouvelle fenêtre avec fonctionnalités d'édition"""
        # Garder une référence au DataFrame original et au chemin du fichier
        self.current_excel_file = filename
        self.current_excel_df = df.copy()
        self.excel_modified = False  # Pour suivre si des modifications ont été effectuées
        
        results_window = tk.Toplevel(self.root)
        results_window.title(f"Données Extra-Comptables - {filename}")
        results_window.geometry(f"{self.root.winfo_screenwidth()}x{self.root.winfo_screenheight()}")

        # Intercepter la fermeture de la fenêtre
        results_window.protocol("WM_DELETE_WINDOW", lambda: self.on_editor_close(results_window))

        # Menu principal
        menu_bar = tk.Menu(results_window)
        results_window.config(menu=menu_bar)
        
        # Menu Fichier
        file_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Fichier", menu=file_menu)
        file_menu.add_command(label="Enregistrer", command=lambda: self.save_excel_changes())
        file_menu.add_command(label="Fermer", command=lambda: self.on_editor_close(results_window))
        
        # Menu Édition
        edit_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="Édition", menu=edit_menu)
        edit_menu.add_command(label="Ajouter une ligne", command=lambda: self.add_excel_row(tree))
        edit_menu.add_command(label="Supprimer la sélection", command=lambda: self.delete_excel_selection(tree))
        
        # Panneau principal
        main_panel = ttk.PanedWindow(results_window, orient=tk.VERTICAL)
        main_panel.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Barre d'outils
        toolbar_frame = ttk.Frame(main_panel)
        main_panel.add(toolbar_frame, weight=0)
        
        ttk.Button(toolbar_frame, text="Ajouter une ligne", 
                  command=lambda: self.add_excel_row(tree)).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar_frame, text="Supprimer la sélection", 
                  command=lambda: self.delete_excel_selection(tree)).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar_frame, text="Enregistrer", 
                  command=lambda: self.save_excel_changes()).pack(side=tk.RIGHT, padx=5)
        ttk.Button(toolbar_frame, text="Fermer", 
                  command=lambda: self.on_editor_close(results_window)).pack(side=tk.RIGHT, padx=5)
        
        # Tableau principal avec les données
        result_frame = ttk.Frame(main_panel)
        main_panel.add(result_frame, weight=1)
        
        # Ajout de barres de défilement
        scrollbar_y = ttk.Scrollbar(result_frame, orient="vertical")
        scrollbar_x = ttk.Scrollbar(result_frame, orient="horizontal")

        # Création du treeview avec les colonnes
        columns = list(df.columns)
        tree = ttk.Treeview(result_frame, columns=columns, show='headings', 
                           yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set,
                           selectmode='extended')  # Mode sélection multiple
        
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100, anchor="center")

        # Configuration du style
        self.style.configure("Treeview", foreground="black", background="white")
        self.style.configure("Treeview.Heading", foreground="black", font=('Helvetica', 10, 'bold'))

        # Configuration des barres de défilement
        scrollbar_y.config(command=tree.yview)
        scrollbar_y.pack(side="right", fill="y")

        scrollbar_x.config(command=tree.xview)
        scrollbar_x.pack(side="bottom", fill="x")

        tree.pack(fill="both", expand=True)

        # Insertion des données
        for i, row in df.iterrows():
            row_values = list(row)
            tree.insert("", "end", values=row_values, tags=(str(i),))  # Tag avec l'index
        
        # Barre d'état
        status_frame = ttk.Frame(results_window)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        self.status_label = ttk.Label(status_frame, text=f"Prêt - {len(df)} lignes")
        self.status_label.pack(side=tk.LEFT, padx=5)
        
        # Configurer l'édition des cellules
        tree.bind("<Double-1>", lambda event: self.on_cell_double_click(event, tree, columns))
        
        # Configurer la fonction de tri en cliquant sur les en-têtes
        for col in columns:
            tree.heading(col, command=lambda c=col: self.sort_treeview(tree, c, False))
        
        # Garder une référence au treeview pour les modifications
        self.current_excel_tree = tree
        results_window.focus_set()


    def on_cell_double_click(self, event, tree, columns):
        """Gérer le double-clic sur une cellule pour l'édition"""
        # Identifier la région où le clic a eu lieu
        region = tree.identify("region", event.x, event.y)
        
        # Si c'est une cellule (pas un en-tête)
        if region == "cell":
            # Obtenir l'identifiant de l'item (ligne) et de la colonne
            item_id = tree.identify_row(event.y)
            column_id = tree.identify_column(event.x)
            
            if not item_id or not column_id:
                return
            
            # Convertir l'identifiant de colonne (comme #1, #2) en numéro d'indice
            column_index = int(column_id.replace('#', '')) - 1
            column_name = columns[column_index]
            
            # Obtenir les coordonnées de la cellule
            x, y, width, height = tree.bbox(item_id, column_id)
            
            # Obtenir la valeur actuelle
            current_value = tree.item(item_id, 'values')[column_index]
            
            # Créer un widget d'entrée pour éditer la valeur
            entry_edit = ttk.Entry(tree)
            entry_edit.insert(0, current_value)
            entry_edit.select_range(0, tk.END)  # Sélectionner tout le texte
            entry_edit.focus()
            
            # Positionner l'entrée sur la cellule
            entry_edit.place(x=x, y=y, width=width, height=height)
            
            def on_edit_done(event=None):
                """Finaliser l'édition et mettre à jour la valeur"""
                new_value = entry_edit.get()
                
                # Mettre à jour la valeur dans le treeview
                values = list(tree.item(item_id, 'values'))
                values[column_index] = new_value
                tree.item(item_id, values=values)
                
                # Mettre à jour le DataFrame
                row_tag = tree.item(item_id, 'tags')[0]
                try:
                    row_index = int(row_tag)
                    self.current_excel_df.at[row_index, column_name] = new_value
                    self.excel_modified = True  # Marquer comme modifié
                    self.status_label.config(text=f"Modifié - {len(self.current_excel_df)} lignes")
                except:
                    pass
                
                # Détruire le widget d'entrée
                entry_edit.destroy()
            
            # Associer les événements
            entry_edit.bind("<Return>", on_edit_done)
            entry_edit.bind("<FocusOut>", on_edit_done)

    def add_excel_row(self, tree):
        """Ajouter une nouvelle ligne au tableau et au DataFrame"""
        # Créer une série avec des valeurs vides pour toutes les colonnes
        new_row = pd.Series([""] * len(self.current_excel_df.columns), index=self.current_excel_df.columns)
        
        # Ajouter la ligne au DataFrame
        self.current_excel_df = pd.concat([self.current_excel_df, pd.DataFrame([new_row])], ignore_index=True)
        
        # Ajouter la ligne au Treeview
        new_idx = len(self.current_excel_df) - 1
        tree.insert("", "end", values=list(new_row), tags=(str(new_idx),))
        
        self.excel_modified = True
        self.status_label.config(text=f"Modifié - {len(self.current_excel_df)} lignes")

    def delete_excel_selection(self, tree):
        """Supprimer les lignes sélectionnées"""
        # Obtenir toutes les lignes sélectionnées
        selected_items = tree.selection()
        
        if not selected_items:
            messagebox.showinfo("Information", "Aucune ligne sélectionnée.")
            return
            
        # Confirmer la suppression
        confirm = messagebox.askyesno("Confirmation", f"Voulez-vous vraiment supprimer {len(selected_items)} ligne(s) ?")
        
        if not confirm:
            return
            
        # Récupérer les indices des lignes à supprimer
        rows_to_delete = []
        for item_id in selected_items:
            try:
                row_tag = tree.item(item_id, 'tags')[0]
                row_index = int(row_tag)
                rows_to_delete.append(row_index)
            except:
                pass
        
        # Supprimer du Treeview
        for item_id in selected_items:
            tree.delete(item_id)
        
        # Supprimer du DataFrame
        if rows_to_delete:
            self.current_excel_df = self.current_excel_df.drop(rows_to_delete).reset_index(drop=True)
            
            # Mettre à jour les tags sur les lignes restantes
            self.rebuild_treeview(tree)
            
            self.excel_modified = True
            self.status_label.config(text=f"Modifié - {len(self.current_excel_df)} lignes")

    def rebuild_treeview(self, tree):
        """Reconstruire le treeview pour mettre à jour les indices"""
        # Vider le treeview
        tree.delete(*tree.get_children())
        
        # Réinsérer les données à partir du DataFrame
        for i, row in self.current_excel_df.iterrows():
            row_values = list(row)
            tree.insert("", "end", values=row_values, tags=(str(i),))

    def save_excel_changes(self):
        """Enregistrer les modifications dans le fichier extra_comptable.xlsx"""
        try:
            # Importer les modules nécessaires
            import os
            import gc
            import time
            import uuid
            import shutil
            
            # Nettoyer explicitement les références
            self.current_excel_tree = None
            gc.collect()
            
            # Déterminer le chemin du dossier de l'application
            app_dir = self.base_path
            target_file = os.path.join(app_dir, "extra_comptable.xlsx")
            
            # Créer un fichier temporaire avec un nom unique
            temp_name = f"temp_{uuid.uuid4().hex}.xlsx"
            temp_path = os.path.join(app_dir, temp_name)
            
            # Enregistrer dans le fichier temporaire
            self.current_excel_df.to_excel(temp_path, index=False)
            
            # Attendre un peu pour s'assurer que l'écriture est terminée
            time.sleep(0.5)
            
            # Tenter de supprimer l'ancien fichier s'il existe
            try:
                if os.path.exists(target_file):
                    # Essayer différentes méthodes pour fermer le fichier si ouvert
                    gc.collect()  # Forcer le ramasse-miettes
                    
                    # Utiliser le module win32file si sur Windows
                    if os.name == 'nt':
                        try:
                            import win32file
                            win32file.DeleteFile(target_file)
                        except:
                            os.remove(target_file)
                    else:
                        os.remove(target_file)
                    
                    # Pause après suppression
                    time.sleep(0.5)
            except Exception as del_error:
                # Si on ne peut pas supprimer, déplacer le fichier temporaire vers l'emplacement final
                # sans essayer de supprimer l'original
                messagebox.showwarning("Avertissement", 
                    f"Impossible de remplacer le fichier existant. Vos modifications ont été enregistrées dans {temp_path}")
                
                # Mettre à jour les références
                self.current_excel_file = temp_path
                self.excel_modified = False
                self.status_label.config(text=f"Enregistré - {len(self.current_excel_df)} lignes")
                return
            
            # Renommer le fichier temporaire vers le nom cible
            try:
                os.rename(temp_path, target_file)
            except:
                # Si le renommage échoue, essayer de copier
                shutil.copy2(temp_path, target_file)
                os.remove(temp_path)
            
            # Mettre à jour les références
            self.current_excel_file = target_file
            self.excel_modified = False
            self.status_label.config(text=f"Enregistré - {len(self.current_excel_df)} lignes")
            
            messagebox.showinfo("Succès", "Les modifications ont été enregistrées avec succès.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'enregistrement: {str(e)}")

    def on_editor_close(self, window):
        """Gérer la fermeture de l'éditeur"""
        if self.excel_modified:
            response = messagebox.askyesnocancel(
                "Modifications non enregistrées",
                "Des modifications n'ont pas été enregistrées. Voulez-vous les enregistrer avant de fermer?"
            )
            
            if response is None:  # Annuler
                return
            elif response:  # Oui, enregistrer
                self.save_excel_changes()
                if self.excel_modified:  # Si l'enregistrement a échoué ou a été annulé
                    return
        
        # Nettoyer les références
        self.current_excel_file = None
        self.current_excel_df = None
        self.current_excel_tree = None
        self.excel_modified = False
        
        # Fermer la fenêtre
        window.destroy()

    def sort_treeview(self, tree, col, reverse):
        """Trier le treeview par colonne"""
        # Récupérer toutes les lignes
        data = [(tree.set(item, col), item) for item in tree.get_children('')]
        
        # Trier les données
        data.sort(reverse=reverse)
        
        # Réorganiser les lignes
        for index, (val, item) in enumerate(data):
            tree.move(item, '', index)
        
        # Inverser l'ordre pour le prochain clic
        tree.heading(col, command=lambda: self.sort_treeview(tree, col, not reverse))
    
    def export_extra_to_csv(self, tree, columns):
        """Exporte les données extra-comptables en CSV"""
        file_path = filedialog.asksaveasfilename(defaultextension=".csv", 
                                                filetypes=[("CSV files", "*.csv")])
        
        if not file_path:
            return
            
        rows = [tree.item(item)["values"] for item in tree.get_children()]
        
        try:
            with open(file_path, mode="w", newline='', encoding="utf-8") as file:
                writer = csv.writer(file, delimiter=';')
                writer.writerow(columns)
                writer.writerows(rows)
                
            messagebox.showinfo("Succès", "Les données ont été exportées avec succès en CSV.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'exportation en CSV: {str(e)}")
    
    def export_extra_to_xlsx(self, tree, columns, results):
        """Exporte les données extra-comptables en XLSX avec mise en forme"""
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                filetypes=[("Excel files", "*.xlsx")])
        
        if not file_path:
            return
            
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Données Extra-Comptables"
        
        # En-têtes
        for col_idx, col_name in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="3282B8", end_color="3282B8", fill_type="solid")
            cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
        
        # Données
        for row_idx, row_data in enumerate(results, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                # Alternance de couleurs pour les lignes
                if row_idx % 2 == 0:
                    cell.fill = openpyxl.styles.PatternFill(start_color="F0F5F9", end_color="F0F5F9", fill_type="solid")
        
        try:
            workbook.save(file_path)
            messagebox.showinfo("Succès", "Les données ont été exportées avec succès au format Excel.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'exportation en XLSX: {str(e)}")

    def importer_donnees_extra(self):
        """Importe les données extra-comptables depuis un fichier Excel fixe vers la table indicateurs"""
        # Variables pour gérer les ressources
        excel_file = None
        conn = None
        
        try:
            # Chemin du fichier Excel fixe (dans le même répertoire que l'application)
            file_path = self.get_resource_path('extra_comptable.xlsx')
            
            # Vérifier si le fichier existe
            if not os.path.exists(file_path):
                messagebox.showerror("Erreur", f"Le fichier 'extra_comptable.xlsx' n'existe pas dans le répertoire de l'application.\nChemin recherché: {file_path}")
                return
            
            # Afficher un message de débogage
            print(f"Fichier trouvé: {file_path}")
            
            # Lire le fichier Excel avec gestion explicite des ressources
            try:
                excel_file = pd.ExcelFile(file_path)
                # Forcer la lecture de tous les types de données sans conversion automatique
                df_extra = excel_file.parse('Sheet1', dtype=str)
                # Fermer immédiatement le fichier après la lecture
                excel_file.close()
                excel_file = None
                
                print(f"Fichier Excel lu avec succès. Colonnes: {df_extra.columns.tolist()}")
                print(f"Nombre de lignes: {len(df_extra)}")
            except Exception as e:
                messagebox.showerror("Erreur", f"Impossible de lire le fichier Excel: {str(e)}")
                return
            
            # Connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Avant d'importer, vérifier et ajouter les colonnes manquantes
            self.add_missing_columns(conn, df_extra.columns.tolist())
            
            # Rechercher une colonne qui pourrait contenir des SIRET
            siret_column = None
            possible_siret_names = ['siret', 'SIRET', 'Siret', 'numéro siret', 'Numéro SIRET', 'num_siret']
            
            for col_name in possible_siret_names:
                if col_name in df_extra.columns:
                    siret_column = col_name
                    print(f"Colonne SIRET trouvée: {siret_column}")
                    break
            
            if not siret_column:
                # Si aucune colonne standard n'est trouvée, chercher par contenu
                for col in df_extra.columns:
                    # Vérifier si la colonne contient des valeurs qui ressemblent à des SIRET (14 chiffres)
                    if df_extra[col].astype(str).str.replace(' ', '').str.replace('-', '').str.match(r'^\d{14}$').any():
                        siret_column = col
                        print(f"Colonne SIRET détectée par contenu: {siret_column}")
                        break
            
            # Vérifier si une colonne SIRET a été trouvée
            if not siret_column:
                all_cols = ", ".join(df_extra.columns.tolist())
                messagebox.showerror("Erreur", f"Aucune colonne contenant des SIRET n'a été trouvée.\nColonnes disponibles: {all_cols}")
                return
            
            # Récupérer les colonnes de la table indicateurs
            cursor.execute("PRAGMA table_info(indicateurs)")
            indicateurs_columns = [col[1] for col in cursor.fetchall()]
            print(f"Colonnes dans la table indicateurs: {indicateurs_columns}")
            
            # Trouver les colonnes communes entre le fichier Excel et la table indicateurs
            common_columns = [col for col in df_extra.columns if col in indicateurs_columns and col != siret_column]
            print(f"Colonnes communes: {common_columns}")
            
            # Vérifier s'il y a des colonnes communes à mettre à jour
            if not common_columns:
                messagebox.showwarning("Avertissement", "Aucune colonne commune (hors SIRET) n'a été trouvée entre le fichier Excel et la table indicateurs.")
                conn.close()
                return
            
            # Récupérer tous les SIRET de la table indicateurs
            cursor.execute("SELECT siret FROM indicateurs WHERE siret != 'moyenne'")
            db_sirets_raw = [row[0] for row in cursor.fetchall()]
            
            # Normaliser les SIRET de la base de données (supprimer espaces, tirets, etc.)
            db_sirets = {str(siret).replace(' ', '').replace('-', ''): siret for siret in db_sirets_raw}
            print(f"Nombre de SIRET dans la base de données: {len(db_sirets)}")
            
            # Pour chaque SIRET dans le fichier Excel, mettre à jour les données si le SIRET existe dans la table
            updates_count = 0
            for _, row in df_extra.iterrows():
                # Normaliser le SIRET du fichier Excel
                excel_siret_raw = str(row[siret_column])
                excel_siret = excel_siret_raw.replace(' ', '').replace('-', '')
                
                # Vérifier si le SIRET normalisé existe dans la base de données
                if excel_siret in db_sirets:
                    # Récupérer le SIRET original de la base
                    original_siret = db_sirets[excel_siret]
                    
                    # Filtrer les valeurs non vides (inclure zéro, nombres négatifs, etc.)
                    valid_columns = []
                    values = []
                    
                    for col in common_columns:
                        if col in row.index and pd.notna(row[col]) and row[col] != '':
                            # Tentative de conversion spécifique pour les colonnes numériques comme impot_revenu
                            if col == 'impot_revenu' or any(col.lower().endswith(suffix) for suffix in ['_revenu', '_montant', '_ca', '_impot']):
                                try:
                                    # Nettoyer la valeur et remplacer la virgule par un point
                                    cleaned_value = str(row[col]).replace(',', '.').strip()
                                    # Tenter de convertir en float pour les colonnes numériques
                                    value = float(cleaned_value) if cleaned_value else None
                                    valid_columns.append(col)
                                    values.append(value)
                                    print(f"Valeur numérique importée pour {col}: {value} (original: {row[col]})")
                                except (ValueError, TypeError) as e:
                                    print(f"Erreur de conversion pour {col}, valeur '{row[col]}': {e}")
                                    # Si la conversion échoue, essayer d'insérer comme chaîne
                                    valid_columns.append(col)
                                    values.append(row[col])
                            else:
                                valid_columns.append(col)
                                values.append(row[col])
                    
                    if valid_columns:
                        # Construire la requête SQL pour mettre à jour les colonnes valides
                        set_clause = ', '.join([f"{col} = ?" for col in valid_columns])
                        query = f"UPDATE indicateurs SET {set_clause} WHERE siret = ?"
                        
                        # Ajouter le SIRET à la fin des valeurs
                        values.append(original_siret)
                        
                        # Exécuter la requête
                        cursor.execute(query, values)
                        updates_count += 1
                        print(f"Mise à jour du SIRET {original_siret} avec {len(valid_columns)} colonnes")
            
            # Sauvegarder les modifications
            conn.commit()
            
            # Mettre à jour le libellé de statut
            self.extra_status_label.config(text=f"Importation réussie: {updates_count} entrées mises à jour")
            
            messagebox.showinfo("Succès", f"Les données ont été importées avec succès. {updates_count} entrées ont été mises à jour dans la table indicateurs.")
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Erreur détaillée: {error_details}")
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'importation des données: {str(e)}")
            if conn:
                conn.rollback()
        finally:
            # S'assurer que toutes les ressources sont libérées
            if excel_file is not None:
                try:
                    excel_file.close()
                except:
                    pass
            
            if conn is not None:
                try:
                    conn.close()
                except:
                    pass


    def add_missing_columns(self, conn, excel_columns):
        """Ajoute les colonnes manquantes à la table indicateurs"""
        try:
            cursor = conn.cursor()
            
            # Liste de colonnes standard à vérifier (en plus des colonnes Excel)
            standard_columns = [
                "Imposition", 
                "OGA", 
                "Regime_de_TVA", 
                "Adresse_eMail", 
                "coll_compta", 
                "date_clot", 
                "Site", 
                "code_postal", 
                "forme_juri", 
                "responsable", 
                "Adresse", 
                "ville", 
                "impot_revenu"
            ]
            
            # Combinaison des colonnes Excel et standard
            all_columns_to_check = list(set(excel_columns + standard_columns))
            
            # Récupérer les colonnes existantes
            cursor.execute("PRAGMA table_info(indicateurs)")
            existing_columns = [col[1] for col in cursor.fetchall()]
            
            # Ajouter les colonnes manquantes
            columns_added = []
            for column in all_columns_to_check:
                if column not in existing_columns:
                    try:
                        # Nettoyer le nom de colonne pour éviter les problèmes SQL
                        clean_column = self.clean_column_name(column)
                        
                        query = f"ALTER TABLE indicateurs ADD COLUMN {clean_column} TEXT"
                        cursor.execute(query)
                        columns_added.append(clean_column)
                        print(f"Colonne '{clean_column}' ajoutée avec succès.")
                    except sqlite3.Error as e:
                        print(f"Erreur lors de l'ajout de la colonne '{column}': {e}")
            
            # Valider les modifications
            conn.commit()
            
            if columns_added:
                print(f"Colonnes ajoutées: {', '.join(columns_added)}")
                messagebox.showinfo("Information", f"Les colonnes suivantes ont été ajoutées à la table indicateurs: \n{', '.join(columns_added)}")
            else:
                print("Aucune colonne n'a été ajoutée, toutes existent déjà.")
            
            return True
        
        except sqlite3.Error as e:
            print(f"Erreur SQLite: {e}")
            conn.rollback()
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'ajout des colonnes: {str(e)}")
            return False


    def process_extra_comptable_data(self, df, file_path):
        """Traite les données extra-comptables et les importe dans la base de données"""
        # Vérifier si des données existent déjà
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Vérifier si la table existe
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='extra_comptable'")
            table_exists = cursor.fetchone()
            
            if table_exists:
                # Demander si l'utilisateur veut remplacer ou ajouter les données
                response = messagebox.askyesnocancel(
                    "Données existantes", 
                    "Des données extra-comptables existent déjà. Souhaitez-vous :\n\n"
                    "Oui: Remplacer les données existantes\n"
                    "Non: Ajouter à la suite des données existantes\n"
                    "Annuler: Abandonner l'importation"
                )
                
                if response is None:  # Annuler
                    conn.close()
                    return
                elif response:  # Oui = remplacer
                    cursor.execute("DROP TABLE IF EXISTS extra_comptable")
                    conn.commit()
            
            # Clean DataFrame column names (remove special characters, spaces)
            df.columns = [self.clean_column_name(col) for col in df.columns]
            
            # Import into the database
            df.to_sql('extra_comptable', conn, if_exists='append', index=False)
            
            # Update the label
            self.extra_status_label.config(text=f"Données importées depuis: {os.path.basename(file_path)}")
            
            messagebox.showinfo("Succès", "Les données extra-comptables ont été importées avec succès.")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'importation des données: {str(e)}")
        finally:
            if conn:
                conn.close()
    
    def clean_column_name(self, column_name):
        """Nettoie les noms de colonnes pour les rendre compatibles avec SQLite"""
        # Remplacer les espaces par des underscores
        clean_name = str(column_name).strip().replace(' ', '_')
        # Supprimer les caractères spéciaux
        clean_name = re.sub(r'[^\w]', '', clean_name)
        # Assurer qu'il commence par une lettre ou un underscore
        if re.match(r'^[0-9]', clean_name):
            clean_name = 'col_' + clean_name
        return clean_name

    # Méthodes existantes du FECExplorer original
    def select_fec_directory_or_files(self):
        response = messagebox.askquestion("Sélection", "Voulez-vous sélectionner un dossier (Oui) ou des fichiers (Non) ?")
        
        if response == 'yes':
            folder_selected = filedialog.askdirectory()
            if folder_selected:
                self.selected_path_label.config(text=f"Dossier sélectionné : {folder_selected}")
                threading.Thread(target=self.process_fec_files_with_progress, args=(folder_selected, None), daemon=True).start()
        else:
            files_selected = filedialog.askopenfilenames(filetypes=[("Fichiers texte", "*.txt")])
            if files_selected:
                self.selected_path_label.config(text=f"{len(files_selected)} fichier(s) sélectionné(s)")
                threading.Thread(target=self.process_fec_files_with_progress, args=(None, files_selected), daemon=True).start()

    def process_fec_files_with_progress(self, fec_directory_path=None, fec_files=None):
        start_time = time.time()
        
        def update_progress(value, message):
            self.update_progress(value, message)
        
        output_file_path = self.process_fec_files(fec_directory_path, fec_files, progress_callback=update_progress)
        
        elapsed_time = time.time() - start_time
        self.update_progress(100, f"Terminé en {elapsed_time:.2f} secondes")

        if output_file_path and os.path.exists(output_file_path):
            self.root.after(0, self.enable_import_button)

    def update_progress(self, value, message):
        self.root.after(0, self._update_progress_ui, value, message)

    def _update_progress_ui(self, value, message):
        self.progress_bar['value'] = value
        self.progress_label.config(text=message)
        self.root.update_idletasks()

    def enable_import_button(self):
        # Cette méthode pourrait être utilisée pour activer le bouton d'importation
        pass

    def process_fec_files(self, fec_directory_path=None, fec_files=None, progress_callback=None):
        if not fec_directory_path and not fec_files:
            messagebox.showwarning("Erreur", "Aucun fichier ou dossier FEC n'a été sélectionné.")
            return

        self.clear_output_directory(self.reports_dir)

        merged_output_file_name = "RapportCombiné.txt"
        merged_output_file_path = os.path.join(self.reports_dir, merged_output_file_name)

        if fec_directory_path:
            pattern = r"^\d{9}FEC\d{8}\.txt$"
            files = [f for f in os.listdir(fec_directory_path) if re.match(pattern, f) and f.endswith('.txt')]
            files = [os.path.join(fec_directory_path, f) for f in files]
        else:
            files = fec_files

        total_files = len(files)
        for file_index, file_path in enumerate(files, start=1):
            file_name = os.path.basename(file_path)
            nine_digits = file_name[:9]
            output_file_path = os.path.join(self.reports_dir, f"{nine_digits}_BalanceGenerale.txt")

            with open(file_path, 'r', encoding='ISO-8859-1') as f:
                reader = csv.reader(f, delimiter='\t')
                next(reader)  # Ignorer la première ligne
                soldes = {}
                for row_num, row in enumerate(reader, start=2):
                    try:
                        if len(row) < 13:
                            raise IndexError(f"La ligne {row_num} ne contient pas suffisamment de colonnes")
                        
                        compte_num = row[4]
                        debit = float(row[11].replace(',', '.')) if row[11] else 0
                        credit = float(row[12].replace(',', '.')) if row[12] else 0
                        
                        if compte_num not in soldes:
                            soldes[compte_num] = {"Debit": 0, "Credit": 0}
                        soldes[compte_num]["Debit"] += debit
                        soldes[compte_num]["Credit"] += credit
                    except IndexError as e:
                        logging.error(f"Erreur dans le fichier {file_name}, ligne {row_num}: {str(e)}")
                        continue
                    except ValueError as e:
                        logging.error(f"Erreur de conversion dans le fichier {file_name}, ligne {row_num}: {str(e)}")
                        continue

            with open(output_file_path, 'w', encoding='utf-8') as f:
                f.write("Compte\tTotal Debit\tTotal Credit\tSolde\n")
                for compte_num, solde in soldes.items():
                    solde_total = solde["Debit"] - solde["Credit"]
                    debit_str = f"{solde['Debit']:.2f}".replace('.', ',')
                    credit_str = f"{solde['Credit']:.2f}".replace('.', ',')
                    solde_total_str = f"{solde_total:.2f}".replace('.', ',')
                    f.write(f"{compte_num}\t{debit_str}\t{credit_str}\t{solde_total_str}\n")

            if progress_callback:
                progress_callback(file_index / total_files * 100, f"Traitement du fichier {file_index}/{total_files}")

        balance_files = [f for f in os.listdir(self.reports_dir) if f.endswith('_BalanceGenerale.txt')]
        if not balance_files:
            messagebox.showwarning("Avertissement", "Aucun fichier de balance générale trouvé.")
            return None

        with open(merged_output_file_path, 'w', encoding='utf-8') as merged_file:
            for balance_file in balance_files:
                balance_file_path = os.path.join(self.reports_dir, balance_file)
                file_name_prefix = balance_file.split("_")[0]
                with open(balance_file_path, 'r', encoding='utf-8') as file:
                    lines = file.readlines()
                    if len(lines) > 1:
                        for line in lines[1:]:
                            merged_file.write(f"{file_name_prefix}\t{line}")

        return merged_output_file_path

    def clear_output_directory(self, output_directory):
        if os.path.exists(output_directory):
            for file_name in os.listdir(output_directory):
                file_path = os.path.join(output_directory, file_name)
                try:
                    if os.path.isfile(file_path) or os.path.islink(file_path):
                        os.unlink(file_path)
                    elif os.path.isdir(file_path):
                        os.rmdir(file_path)
                except Exception as e:
                    logging.error(f"Erreur lors de la suppression du fichier {file_path}: {e}")
        else:
            os.makedirs(output_directory)
            
    def update_database(self):
        output_file_path = os.path.join(self.reports_dir, "RapportCombiné.txt")
        
        if os.path.exists(output_file_path):
            self.import_to_database(output_file_path)
            self.selected_path_label.config(text="Importer dans la base de données a bien été exécuté")
        else:
            messagebox.showwarning("Erreur", "Le fichier RapportCombiné.txt n'existe pas encore. Veuillez d'abord traiter les fichiers FEC.")

    def import_to_database(self, file_path):
        if not os.path.exists(file_path):
            messagebox.showwarning("Erreur", f"Le fichier {file_path} n'existe pas. Impossible de l'importer.")
            return

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            cursor.execute("DROP TABLE IF EXISTS rapport_combine")
            conn.commit()
            logging.info("Table 'rapport_combine' supprimée avec succès.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la suppression de la table rapport_combine : {str(e)}")
            return

        try:
            df = pd.read_csv(file_path, delimiter='\t', header=None)
            if df.empty:
                messagebox.showwarning("Erreur", "Le fichier de rapport est vide ou mal formaté.")
                return

            df.columns = ['siret', 'compte', 'debit', 'credit', 'solde']
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la lecture du fichier: {str(e)}")
            return

        try:
            df.to_sql('rapport_combine', conn, if_exists='replace', index=False)
            conn.commit()
            messagebox.showinfo("Succès", "Les données ont été importées avec succès dans la base de données.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'importation des données dans la base de données: {str(e)}")
        finally:
            conn.close()

    def create_indicators(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Création de la table indicateurs
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS indicateurs (
            siret TEXT PRIMARY KEY,
            CA REAL,
            assurance REAL,
            deplacement REAL,
            loyer REAL,
            CFE REAL,
            TNS REAL,
            publicité REAL,
            honoraires REAL,
            banque REAL,
            emprunt REAL,
            m_salariale REAL,
            produits REAL,
            charge REAL,
            trésorerie REAL,
            compte_791 REAL,
            resultat REAL,
            prestation TEXT,
            multitva TEXT
        )
        ''')
        conn.commit()

        # Charger les données de la table rapport_combine
        df = pd.read_sql_query('SELECT * FROM rapport_combine', conn)

        # Nettoyage et préparation des données
        df['solde'] = df['solde'].apply(lambda x: Decimal(str(x).replace(',', '.').strip()))
        df['compte'] = df['compte'].astype(str)

        # Initialisation et calcul des indicateurs
        indicators = {
            'CA': lambda row: -row['solde'] if row['compte'].startswith('70') else Decimal('0'),
            'assurance': lambda row: row['solde'] if row['compte'].startswith('616') else Decimal('0'),
            'deplacement': lambda row: row['solde'] if row['compte'].startswith('625') else Decimal('0'),
            'loyer': lambda row: row['solde'] if row['compte'].startswith('613') else Decimal('0'),
            'CFE': lambda row: row['solde'] if row['compte'].startswith('63511') else Decimal('0'),
            'TNS': lambda row: row['solde'] if row['compte'].startswith('644') else Decimal('0'),
            'publicité': lambda row: row['solde'] if row['compte'].startswith('623') else Decimal('0'),
            'honoraires': lambda row: row['solde'] if row['compte'].startswith('6226') else Decimal('0'),
            'banque': lambda row: row['solde'] if row['compte'].startswith('627') else Decimal('0'),
            'emprunt': lambda row: row['solde'] if row['compte'].startswith('16') else Decimal('0'),
            'm_salariale': lambda row: row['solde'] if row['compte'].startswith('64') else Decimal('0'),
            'produits': lambda row: -row['solde'] if row['compte'].startswith('7') and not row['compte'].startswith('791') else Decimal('0'),
            'compte_791': lambda row: -row['solde'] if row['compte'].startswith('791') else Decimal('0'),
            'charge': lambda row: row['solde'] if row['compte'].startswith('6') else Decimal('0'),
            'trésorerie': lambda row: row['solde'] if row['compte'].startswith('5') else Decimal('0'),
        }

        for indicator, func in indicators.items():
            df[indicator] = df.apply(func, axis=1)

        # Grouper par siret et calculer les sommes
        df_grouped = df.groupby('siret').agg({k: 'sum' for k in indicators.keys()}).reset_index()

        # Calculer le résultat en ajoutant le compte 791 au résultat
        df_grouped['resultat'] = df_grouped['produits'] - df_grouped['charge'] + df_grouped['compte_791']

        # Vérification des conditions pour 'prestation'
        grouped_prestation = df.groupby('siret').agg(
            has_707=pd.NamedAgg(column='compte', aggfunc=lambda x: any(str(compte).startswith('707') for compte in x)),
            has_706=pd.NamedAgg(column='compte', aggfunc=lambda x: any(str(compte).startswith('706') for compte in x))
        ).reset_index()
        grouped_prestation['prestation'] = grouped_prestation.apply(
            lambda row: 'presta' if not row['has_707'] and row['has_706'] else None, axis=1
        )

        # Vérification des conditions pour 'multitva'
        df_multitva = df[df['compte'].str.startswith('4457')]
        df_grouped_multitva = df_multitva.groupby('siret').size().reset_index(name='count_4457')
        df_grouped_multitva['multitva'] = df_grouped_multitva['count_4457'].apply(lambda x: 'multitva' if x > 1 else None)

        # Joindre les informations
        df_final = pd.merge(df_grouped, grouped_prestation[['siret', 'prestation']], on='siret', how='left')
        df_final = pd.merge(df_final, df_grouped_multitva[['siret', 'multitva']], on='siret', how='left')

        # Conversion en float pour SQLite
        float_columns = [col for col in df_final.columns if col not in ['siret', 'prestation', 'multitva']]
        df_final[float_columns] = df_final[float_columns].astype(float)

        # Insertion dans la table indicateurs
        df_final.to_sql('indicateurs', conn, if_exists='replace', index=False)

        conn.commit()
        conn.close()

        messagebox.showinfo("Succès", "Les indicateurs ont été créés avec succès.")

        # Appel de la méthode pour calculer les moyennes
        self.calculate_averages()

    def calculate_averages(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Récupérer les noms des colonnes
        cursor.execute("PRAGMA table_info(indicateurs)")
        columns = [col[1] for col in cursor.fetchall()]
        columns_to_ignore = ['siret', 'prestation', 'multitva']
        columns_to_calculate = [col for col in columns if col not in columns_to_ignore]

        # Calculer les moyennes
        avg_query = ", ".join([f"AVG({col})" for col in columns_to_calculate])
        cursor.execute(f"SELECT {avg_query} FROM indicateurs")
        averages = cursor.fetchone()

        # Préparer les données pour l'insertion
        values = ['moyenne'] + list(averages) + [None, None]  # None pour prestation et multitva

        # Insérer ou mettre à jour la ligne de moyenne
        placeholders = ', '.join(['?'] * len(columns))
        cursor.execute(f"INSERT OR REPLACE INTO indicateurs ({', '.join(columns)}) VALUES ({placeholders})", values)

        conn.commit()
        conn.close()

        messagebox.showinfo("Succès", "Les moyennes ont été calculées et ajoutées à la table indicateurs.")

    def export_database(self):
        def remove_accents(input_str):
            nfkd_form = unicodedata.normalize('NFKD', input_str)
            return u"".join([c for c in nfkd_form if not unicodedata.combining(c)])

        try:
            conn = sqlite3.connect(self.db_path)
            df = pd.read_sql_query("SELECT * FROM indicateurs", conn)
            conn.close()

            if df.empty:
                messagebox.showwarning("Erreur", "La table 'indicateurs' est vide ou n'existe pas.")
                return

            df.columns = [remove_accents(col) for col in df.columns]

            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("Fichier CSV", "*.csv")],
                title="Enregistrer le fichier CSV"
            )

            if file_path:
                df.to_csv(file_path, sep=';', index=False, encoding='utf-8')
                messagebox.showinfo("Succès", f"Les données ont été exportées avec succès vers {file_path}")
            else:
                messagebox.showwarning("Annulé", "L'exportation a été annulée.")

        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'exportation des données : {str(e)}")

    def show_procedure(self):
        def open_link(event):
            webbrowser.open("https://annuaire-entreprises.data.gouv.fr/export-sirene")

        procedure_window = tk.Toplevel(self.root)
        procedure_window.title("Procédures pour affecter des noms et codes NAF à vos dossiers")

        process_text = (
            "Procédure de gestion de la boîte de dialogue\n\n"
            "**Opération pour moins de 500 dossiers**\n"
            "Pour plus de dossiers, créer des fichiers par blocs de 500 puis créer un seul CSV avec tous les dossiers\n\n"
            "1. Ouvrir le lien : Accédez à Sirene.fr via ce lien : "
        )

        label = tk.Label(procedure_window, text=process_text, font=("Helvetica", 12), justify="left")
        label.pack(padx=10, pady=5, anchor="w")

        link_label = tk.Label(procedure_window, text="https://annuaire-entreprises.data.gouv.fr/export-sirene", 
                              fg="blue", cursor="hand2", font=("Helvetica", 12))
        link_label.pack(padx=10, anchor="w")
        link_label.bind("<Button-1>", open_link)

        more_text = (
            "\n2. Sélectionner un fichier :\n"
            "   Cliquez sur « Choisir un fichier » dans la section charger une liste de SIREN/SIRET.\n"
            "   Recherchez et sélectionnez le fichier créé via l'outil de consolidation FEC (Bouton : Création du fichier avec les SIRET).\n\n"
            "3. Constituer la liste : Cliquez sur « Calculer les résultats » sur Sirene.fr.\n\n"
            "4. Télécharger la liste : Cliquez sur « Télécharger la liste ».\n\n"
            "5. Importer dans la base de données :\n"
            "   Retournez dans l'outil de consolidation FEC.\n"
            "   Cliquez sur « Import des fichiers dans votre base de données » et sélectionnez le fichier téléchargé."
        )

        more_text_label = tk.Label(procedure_window, text=more_text, font=("Helvetica", 12), justify="left")
        more_text_label.pack(padx=10, pady=5, anchor="w")

        close_button = tk.Button(procedure_window, text="Fermer", command=procedure_window.destroy, font=("Helvetica", 14))
        close_button.pack(pady=10)

    def extract_siret(self):
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute("SELECT siret FROM indicateurs WHERE siret != 'moyenne' AND siret GLOB '[0-9]*'")
            sirets = cursor.fetchall()

            emplacement_fichier = filedialog.asksaveasfilename(
                defaultextension=".txt", 
                filetypes=[("Text files", "*.txt")],
                title="Choisissez l'emplacement de sauvegarde"
            )

            if emplacement_fichier:
                with open(emplacement_fichier, 'w') as fichier:
                    for siret in sirets:
                        fichier.write(f"{siret[0]}\n")

                messagebox.showinfo("Succès", f"Extraction réussie. Fichier enregistré à : {emplacement_fichier}")
            else:
                messagebox.showinfo("Information", "Aucun fichier sélectionné. Opération annulée.")

        except sqlite3.Error as erreur:
            messagebox.showerror("Erreur", f"Erreur lors de la connexion à la base de données : {erreur}")

        finally:
            if conn:
                conn.close()

    def import_nom(self):
        try:
            # Ouvrir l'explorateur de fichiers pour choisir le fichier CSV
            chemin_csv = filedialog.askopenfilename(
                filetypes=[("CSV files", "*.csv")],
                title="Choisissez un fichier CSV à importer"
            )

            # Vérification si un fichier a été sélectionné
            if not chemin_csv:
                messagebox.showinfo("Information", "Aucun fichier sélectionné. Opération annulée.")
                return

            # Connexion à la base de données SQLite
            connexion = sqlite3.connect(self.db_path)
            curseur = connexion.cursor()

            # 1. CRÉER LA TABLE nomdossier EN PREMIER (avec toutes les colonnes nécessaires)
            curseur.execute('''
                CREATE TABLE IF NOT EXISTS nomdossier (
                    siren TEXT,
                    denominationUniteLegale TEXT,
                    activitePrincipaleEtablissement TEXT,
                    nomUniteLegale TEXT,
                    prenom1UniteLegale TEXT
                )
            ''')
            connexion.commit()
            print("Table nomdossier créée ou vérifiée avec succès.")

            # 2. Suppression des anciennes données de la table
            curseur.execute("DELETE FROM nomdossier")
            connexion.commit()
            print("Anciennes données supprimées.")

            # 3. Lecture du fichier CSV en récupérant les colonnes spécifiées
            try:
                df = pd.read_csv(chemin_csv, usecols=[
                    'siren', 
                    'denominationUniteLegale', 
                    'activitePrincipaleEtablissement',
                    'nomUniteLegale',
                    'prenom1UniteLegale'
                ], sep=',')
                print(f"Fichier CSV lu avec succès. {len(df)} lignes trouvées.")
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de la lecture du fichier CSV: {str(e)}")
                connexion.close()
                return

            # 4. Insertion des données du CSV dans la table `nomdossier`
            compteur_inserts = 0
            for _, row in df.iterrows():
                try:
                    curseur.execute('''
                        INSERT INTO nomdossier (siren, denominationUniteLegale, activitePrincipaleEtablissement, nomUniteLegale, prenom1UniteLegale) 
                        VALUES (?, ?, ?, ?, ?)
                    ''', (
                        row['siren'], 
                        row['denominationUniteLegale'], 
                        row['activitePrincipaleEtablissement'], 
                        row['nomUniteLegale'], 
                        row['prenom1UniteLegale']
                    ))
                    compteur_inserts += 1
                except Exception as e:
                    print(f"Erreur lors de l'insertion de la ligne: {str(e)}")
                    continue

            # Sauvegarde des modifications dans la base de données
            connexion.commit()
            print(f"Données importées avec succès: {compteur_inserts} lignes insérées.")

            # 5. Vérifier que la table indicateurs existe
            curseur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='indicateurs'")
            if not curseur.fetchone():
                messagebox.showerror("Erreur", "La table 'indicateurs' n'existe pas. Veuillez d'abord créer les indicateurs.")
                connexion.close()
                return

            # 6. Ajouter les colonnes `code_naf`, `nom`, `nomUniteLegale`, et `prenom1UniteLegale` à la table `indicateurs` si elles n'existent pas
            curseur.execute("PRAGMA table_info(indicateurs)")
            colonnes_existantes = [colonne[1] for colonne in curseur.fetchall()]

            colonnes_a_ajouter = ['code_naf', 'nom', 'nomUniteLegale', 'prenom1UniteLegale']
            for colonne in colonnes_a_ajouter:
                if colonne not in colonnes_existantes:
                    try:
                        curseur.execute(f"ALTER TABLE indicateurs ADD COLUMN {colonne} TEXT")
                        print(f"Colonne '{colonne}' ajoutée à la table indicateurs.")
                    except Exception as e:
                        print(f"Erreur lors de l'ajout de la colonne {colonne}: {str(e)}")

            # Sauvegarde après ajout des colonnes
            connexion.commit()

            # 7. Mise à jour des colonnes dans la table `indicateurs`
            try:
                curseur.execute('''
                    UPDATE indicateurs
                    SET code_naf = (
                        SELECT activitePrincipaleEtablissement 
                        FROM nomdossier 
                        WHERE nomdossier.siren = substr(indicateurs.siret, 1, 9)
                    ),
                    nom = COALESCE(nom, (
                        SELECT denominationUniteLegale
                        FROM nomdossier 
                        WHERE nomdossier.siren = substr(indicateurs.siret, 1, 9)
                    )),
                    nomUniteLegale = (
                        SELECT nomUniteLegale 
                        FROM nomdossier 
                        WHERE nomdossier.siren = substr(indicateurs.siret, 1, 9)
                    ),
                    prenom1UniteLegale = (
                        SELECT prenom1UniteLegale 
                        FROM nomdossier 
                        WHERE nomdossier.siren = substr(indicateurs.siret, 1, 9)
                    )
                    WHERE EXISTS (
                        SELECT 1 
                        FROM nomdossier 
                        WHERE nomdossier.siren = substr(indicateurs.siret, 1, 9)
                    )
                ''')

                # Sauvegarde des modifications dans la base de données
                connexion.commit()
                print("Mise à jour de la table indicateurs terminée.")

                # Vérifier combien de lignes ont été mises à jour
                curseur.execute("SELECT COUNT(*) FROM indicateurs WHERE code_naf IS NOT NULL")
                nb_maj = curseur.fetchone()[0]
                print(f"Nombre de dossiers mis à jour: {nb_maj}")

            except Exception as e:
                print(f"Erreur lors de la mise à jour: {str(e)}")
                messagebox.showerror("Erreur", f"Erreur lors de la mise à jour de la table indicateurs: {str(e)}")
                connexion.close()
                return

            # 8. Appel de la méthode pour mettre à jour les libellés NAF
            try:
                self.mise_a_jour_activite()
                print("Mise à jour des activités terminée.")
            except Exception as e:
                print(f"Erreur lors de la mise à jour des activités: {str(e)}")

            messagebox.showinfo("Succès", f"Importation réussie!\n\n"
                                         f"- {compteur_inserts} lignes importées dans nomdossier\n"
                                         f"- {nb_maj} dossiers mis à jour dans indicateurs\n"
                                         f"- Activités mises à jour")

        except (sqlite3.Error, pd.errors.EmptyDataError) as erreur:
            logging.error(f"Erreur lors du traitement : {erreur}")
            messagebox.showerror("Erreur", f"Une erreur est survenue : {erreur}")

        except Exception as erreur:
            logging.error(f"Erreur inattendue : {erreur}")
            messagebox.showerror("Erreur", f"Une erreur inattendue est survenue : {erreur}")
            import traceback
            traceback.print_exc()

        finally:
            if 'connexion' in locals():
                connexion.close()
                print("Connexion à la base de données fermée.")

    def mise_a_jour_activite(self):
        try:
            connexion = sqlite3.connect(self.db_path)
            curseur = connexion.cursor()

            # 1. Vérifier si la colonne `activite` existe déjà, sinon l'ajouter
            curseur.execute("PRAGMA table_info(indicateurs)")
            

            colonnes_existantes = [colonne[1] for colonne in curseur.fetchall()]

            if 'activite' not in colonnes_existantes:
                curseur.execute("ALTER TABLE indicateurs ADD COLUMN activite TEXT")
                connexion.commit()
                logging.info("Colonne 'activite' ajoutée à la table `indicateurs`.")

            # 2. Mettre à jour la colonne `activite` dans la table `indicateurs`
            curseur.execute('''
                UPDATE indicateurs
                SET activite = (
                    SELECT Libellé
                    FROM libelle_naf
                    WHERE CAST(SUBSTR(indicateurs.code_naf, 1, 2) AS INTEGER) = CAST(SUBSTR(libelle_naf.Code, 1, 2) AS INTEGER)
                )
                WHERE EXISTS (
                    SELECT 1
                    FROM libelle_naf
                    WHERE CAST(SUBSTR(indicateurs.code_naf, 1, 2) AS INTEGER) = CAST(SUBSTR(libelle_naf.Code, 1, 2) AS INTEGER)
                )
            ''')

            connexion.commit()
            logging.info("La colonne 'activite' de la table `indicateurs` a été mise à jour avec succès.")

        except sqlite3.Error as erreur:
            logging.error(f"Erreur lors de la mise à jour de l'activité : {erreur}")
            messagebox.showerror("Erreur", f"Erreur lors de la mise à jour de l'activité : {erreur}")

        finally:
            if connexion:
                connexion.close()

    def recherche_croisee_comptable(self):
        recherche_window = tk.Toplevel(self.root)
        recherche_window.title("Recherche Croisée Comptable")
        recherche_window.geometry("860x510")
        recherche_window.configure(bg="#f0f5f9")
        recherche_window.resizable(True, False)

        # En-tête
        header_frame = tk.Frame(recherche_window, bg="#2c3e50", height=58)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        tk.Label(header_frame, text="🔍  Recherche Croisée Comptable",
                 font=("Helvetica", 13, "bold"), fg="white", bg="#2c3e50").pack(side=tk.LEFT, padx=20, pady=15)

        # Corps — 3 cartes de critères côte à côte
        body = ttk.Frame(recherche_window, padding=(20, 15, 20, 5))
        body.pack(fill=tk.BOTH, expand=True)
        for i in range(3):
            body.grid_columnconfigure(i, weight=1)

        TYPES = ["Premier chiffre", "Deux premiers chiffres",
                 "Trois premiers chiffres", "Quatre premiers chiffres"]
        self.compte_widgets = []

        for i in range(3):
            card = ttk.LabelFrame(body, text=f"  Critère {i+1}  ", padding=(12, 10))
            card.grid(row=0, column=i, padx=8, pady=5, sticky="nsew")

            ttk.Label(card, text="Précision du compte", font=("Helvetica", 9, "bold"),
                      foreground="#555555").pack(anchor="w")
            compte_cb = ttk.Combobox(card, values=TYPES, width=26, state="readonly")
            compte_cb.pack(fill=tk.X, pady=(3, 10))

            ttk.Label(card, text="Numéro de compte", font=("Helvetica", 9, "bold"),
                      foreground="#555555").pack(anchor="w")
            compte_entry = ttk.Entry(card, width=28)
            compte_entry.pack(fill=tk.X, pady=(3, 10))

            ttk.Separator(card, orient="horizontal").pack(fill=tk.X, pady=6)

            range_row = ttk.Frame(card)
            range_row.pack(fill=tk.X)
            ttk.Label(range_row, text="Min", width=4, foreground="#555555").pack(side=tk.LEFT)
            montant_min = ttk.Entry(range_row, width=9)
            montant_min.pack(side=tk.LEFT, padx=(3, 10))
            ttk.Label(range_row, text="Max", width=4, foreground="#555555").pack(side=tk.LEFT)
            montant_max = ttk.Entry(range_row, width=9)
            montant_max.pack(side=tk.LEFT, padx=(3, 0))

            self.compte_widgets.append((compte_cb, compte_entry))
            setattr(self, f'montant_min_entry{i+1}', montant_min)
            setattr(self, f'montant_max_entry{i+1}', montant_max)

        # Note explicative sur les conventions de solde
        note = ttk.LabelFrame(body, text="  ℹ️  Conventions de solde  ", padding=(10, 6))
        note.grid(row=1, column=0, columnspan=3, padx=0, pady=(12, 0), sticky="ew")
        note.grid_columnconfigure(0, weight=1)
        note.grid_columnconfigure(1, weight=1)

        left_note = tk.Frame(note, bg="#fef9e7", bd=1, relief="solid")
        left_note.grid(row=0, column=0, padx=(0, 6), pady=2, sticky="nsew")
        tk.Label(left_note, text="Solde négatif  (créditeur)",
                 font=("Helvetica", 9, "bold"), fg="#7d6608", bg="#fef9e7").pack(anchor="w", padx=10, pady=(6, 3))
        for txt in ("Classe 7  —  Produits",
                    "Classe 1  —  Passif / Capitaux propres",
                    "Compte 401  —  Fournisseurs"):
            tk.Label(left_note, text=f"  •  {txt}",
                     font=("Helvetica", 9), fg="#5d4037", bg="#fef9e7").pack(anchor="w", padx=10)
        tk.Label(left_note, text="", bg="#fef9e7").pack(pady=2)

        right_note = tk.Frame(note, bg="#eaf4fb", bd=1, relief="solid")
        right_note.grid(row=0, column=1, padx=(6, 0), pady=2, sticky="nsew")
        tk.Label(right_note, text="Solde positif  (débiteur)",
                 font=("Helvetica", 9, "bold"), fg="#1a5276", bg="#eaf4fb").pack(anchor="w", padx=10, pady=(6, 3))
        for txt in ("Classe 6  —  Charges",
                    "Classes 2, 3, 5  —  Actif (immo., stocks, tréso.)",
                    "Compte 411  —  Clients"):
            tk.Label(right_note, text=f"  •  {txt}",
                     font=("Helvetica", 9), fg="#154360", bg="#eaf4fb").pack(anchor="w", padx=10)
        tk.Label(right_note, text="", bg="#eaf4fb").pack(pady=2)

        # Barre d'actions fixe en bas
        action_bar = tk.Frame(recherche_window, bg="#ecf0f1")
        action_bar.pack(fill=tk.X, side=tk.BOTTOM)
        ttk.Separator(action_bar, orient="horizontal").pack(fill=tk.X)
        btn_row = tk.Frame(action_bar, bg="#ecf0f1")
        btn_row.pack(pady=12, padx=20)

        def reset_filters():
            for cb, entry in self.compte_widgets:
                cb.set("")
                entry.delete(0, tk.END)
            for j in range(1, 4):
                getattr(self, f'montant_min_entry{j}').delete(0, tk.END)
                getattr(self, f'montant_max_entry{j}').delete(0, tk.END)

        ttk.Button(btn_row, text="Réinitialiser", command=reset_filters, width=14).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_row, text="🔍  Rechercher", command=self.execute_cross_search,
                   style="Accent.TButton", width=16).pack(side=tk.LEFT)

    def execute_cross_search(self):
        results = set()
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        active_criteria = []  # liste de (substr_len, compte_value, label_critere)

        SUBSTR_MAP = {
            "premier chiffre": 1,
            "deux premiers chiffres": 2,
            "trois premiers chiffres": 3,
            "quatre premiers chiffres": 4,
        }

        for i in range(1, 4):
            query = f"""
                SELECT [siret], SUM(CAST(solde AS REAL)) AS total_solde{i}
                FROM rapport_combine
                WHERE 1=1
            """
            query_params = []

            compte_cb, compte_entry = self.compte_widgets[i-1]
            compte_type = compte_cb.get()
            compte_value = compte_entry.get()

            substr_len = None
            if compte_type and compte_value:
                for key, length in SUBSTR_MAP.items():
                    if key in compte_type.lower():
                        substr_len = length
                        break
                if substr_len:
                    query += f" AND SUBSTR(compte, 1, {substr_len}) = ?"
                    query_params.append(compte_value)
                    active_criteria.append((substr_len, compte_value, f"Critère {i} — {compte_type} = {compte_value}"))

            montant_min = getattr(self, f'montant_min_entry{i}').get()
            montant_max = getattr(self, f'montant_max_entry{i}').get()

            if montant_min and montant_max:
                query += f" GROUP BY [siret] HAVING total_solde{i} BETWEEN ? AND ?"
                query_params.append(float(montant_min))
                query_params.append(float(montant_max))
            elif montant_min:
                query += f" GROUP BY [siret] HAVING total_solde{i} >= ?"
                query_params.append(float(montant_min))
            elif montant_max:
                query += f" GROUP BY [siret] HAVING total_solde{i} <= ?"
                query_params.append(float(montant_max))
            else:
                query += f" GROUP BY [siret]"

            try:
                cursor.execute(query, tuple(query_params))
                current_results = set(row[0] for row in cursor.fetchall())
                if i == 1:
                    results = current_results
                else:
                    results &= current_results
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de l'exécution de la recherche croisée pour Numéro de compte {i} : {str(e)}")

        self.display_cross_search_results(results, conn, cursor, active_criteria)

    def display_cross_search_results(self, results, conn, cursor, active_criteria=None):
        if not results:
            messagebox.showinfo("Information", "Aucun dossier ne correspond aux critères de recherche.")
            return

        # --- Données indicateurs ---
        cursor.execute("PRAGMA table_info(indicateurs)")
        ind_columns = [info[1] for info in cursor.fetchall()]
        excluded_columns = ['siret', 'prestation', 'multitva']

        cursor.execute(f"SELECT * FROM indicateurs WHERE siret IN ({','.join(['?']*len(results))})", list(results))
        ind_rows = cursor.fetchall()

        numeric_totals = {col: 0 for col in ind_columns if col not in excluded_columns}
        count_rows = len(ind_rows)
        for row in ind_rows:
            for i, col in enumerate(ind_columns):
                if col not in excluded_columns:
                    try:
                        numeric_totals[col] += float(row[i])
                    except (ValueError, TypeError):
                        pass

        # --- Données comptes depuis rapport_combine ---
        compte_rows = []
        compte_columns = ["siret", "compte", "solde_total"]
        if active_criteria:
            conditions = " OR ".join([f"SUBSTR(compte, 1, {n}) = ?" for n, _, _ in active_criteria])
            criteria_params = [v for _, v, _ in active_criteria]
            placeholders = ','.join(['?'] * len(results))
            rc_query = f"""
                SELECT siret, compte, SUM(CAST(solde AS REAL)) AS solde_total
                FROM rapport_combine
                WHERE ({conditions})
                AND siret IN ({placeholders})
                GROUP BY siret, compte
                ORDER BY siret, compte
            """
            try:
                cursor.execute(rc_query, criteria_params + list(results))
                compte_rows = cursor.fetchall()
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur récupération comptes : {str(e)}")

        results_window = tk.Toplevel(self.root)
        results_window.title("Resultats - Recherche Croisee Comptable")
        w = min(self.root.winfo_screenwidth(), 1400)
        h = min(self.root.winfo_screenheight(), 860)
        results_window.geometry(f"{w}x{h}")
        results_window.configure(bg="#f0f5f9")

        # En-tête
        header_frame = tk.Frame(results_window, bg="#27ae60", height=55)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        tk.Label(header_frame,
                 text=f"✅  {count_rows} dossier(s)  —  {len(compte_rows)} ligne(s) de compte(s)",
                 font=("Helvetica", 13, "bold"), fg="white", bg="#27ae60").pack(side=tk.LEFT, padx=20, pady=14)

        # Styles tableau
        self.style.configure("Treeview", foreground="black", background="white")
        self.style.configure("Treeview.Heading", foreground="black", font=("Helvetica", 10, "bold"))

        # Notebook 2 onglets
        notebook = ttk.Notebook(results_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # ---- Onglet 1 : Comptes (rapport_combine) ----
        tab_comptes = ttk.Frame(notebook)
        notebook.add(tab_comptes, text=f"  Comptes sélectionnés ({len(compte_rows)} lignes)  ")

        if active_criteria:
            crit_bar = tk.Frame(tab_comptes, bg="#eaf2fb")
            crit_bar.pack(fill=tk.X, padx=5, pady=(5, 0))
            labels = "  |  ".join(lbl for _, _, lbl in active_criteria)
            tk.Label(crit_bar, text=f"Filtres actifs :  {labels}",
                     font=("Helvetica", 9, "italic"), fg="#2c3e50", bg="#eaf2fb").pack(side=tk.LEFT, padx=8, pady=4)

        tree_c = None
        if compte_rows:
            cf = ttk.Frame(tab_comptes, padding=(5, 5, 5, 0))
            cf.pack(fill=tk.BOTH, expand=True)
            sy1 = ttk.Scrollbar(cf, orient="vertical")
            sx1 = ttk.Scrollbar(cf, orient="horizontal")
            tree_c = ttk.Treeview(cf, columns=compte_columns, show="headings",
                                  yscrollcommand=sy1.set, xscrollcommand=sx1.set)
            col_labels = {"siret": "SIRET", "compte": "N° Compte", "solde_total": "Solde total (EUR)"}
            for col in compte_columns:
                tree_c.heading(col, text=col_labels.get(col, col), anchor="center")
                tree_c.column(col, width=250 if col == "siret" else 180, minwidth=100, anchor="center")
            sy1.config(command=tree_c.yview)
            sy1.pack(side="right", fill="y")
            sx1.config(command=tree_c.xview)
            sx1.pack(side="bottom", fill="x")
            tree_c.pack(fill="both", expand=True)
            tree_c.tag_configure("even", background="#ffffff")
            tree_c.tag_configure("odd", background="#eaf2fb")
            for i, row in enumerate(compte_rows):
                tree_c.insert("", "end", values=row, tags=("odd" if i % 2 else "even",))
        else:
            ttk.Label(tab_comptes,
                      text="Aucune donnée de compte — vérifiez les critères saisis.",
                      font=("Helvetica", 11), foreground="#888").pack(expand=True)

        # ---- Onglet 2 : Indicateurs ----
        tab_ind = ttk.Frame(notebook)
        notebook.add(tab_ind, text=f"  Indicateurs ({count_rows} dossiers)  ")

        tree_i = None
        if ind_rows:
            rf = ttk.Frame(tab_ind, padding=(5, 5, 5, 0))
            rf.pack(fill=tk.BOTH, expand=True)
            sy2 = ttk.Scrollbar(rf, orient="vertical")
            sx2 = ttk.Scrollbar(rf, orient="horizontal")
            tree_i = ttk.Treeview(rf, columns=ind_columns, show="headings",
                                  yscrollcommand=sy2.set, xscrollcommand=sx2.set)
            for col in ind_columns:
                tree_i.heading(col, text=col, anchor="center")
                tree_i.column(col, width=120, minwidth=80, anchor="center")
            sy2.config(command=tree_i.yview)
            sy2.pack(side="right", fill="y")
            sx2.config(command=tree_i.xview)
            sx2.pack(side="bottom", fill="x")
            tree_i.pack(fill="both", expand=True)
            tree_i.tag_configure("even", background="#ffffff")
            tree_i.tag_configure("odd", background="#eaf2fb")
            tree_i.tag_configure("avg", background="#fef9e7", font=("Helvetica", 9, "bold"))
            for i, row in enumerate(ind_rows):
                tree_i.insert("", "end", values=row, tags=("odd" if i % 2 else "even",))
            if count_rows > 0:
                averages = []
                for col in ind_columns:
                    if col == 'siret':
                        averages.append("Moyenne")
                    elif col in excluded_columns:
                        averages.append("")
                    else:
                        averages.append(f"{numeric_totals[col] / count_rows:.2f}")
                tree_i.insert("", "end", values=averages, tags=("avg",))
        else:
            ttk.Label(tab_ind, text="Aucun indicateur disponible.",
                      font=("Helvetica", 11), foreground="#888").pack(expand=True)

        # Barre d'actions bas
        bottom_bar = tk.Frame(results_window, bg="#ecf0f1")
        bottom_bar.pack(fill=tk.X, side=tk.BOTTOM)
        ttk.Separator(bottom_bar, orient="horizontal").pack(fill=tk.X)
        btn_row = tk.Frame(bottom_bar, bg="#ecf0f1")
        btn_row.pack(pady=10, padx=20)
        if tree_c:
            ttk.Button(btn_row, text="📥  Comptes (CSV)",
                       command=lambda: self.export_to_csv(tree_c),
                       width=18).pack(side=tk.LEFT, padx=(0, 8))
        if tree_i:
            ttk.Button(btn_row, text="📥  Indicateurs (CSV)",
                       command=lambda: self.export_to_csv(tree_i),
                       style="Accent.TButton", width=20).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_row, text="Fermer", command=results_window.destroy, width=12).pack(side=tk.LEFT)

    def export_to_csv(self, tree):
        file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])

        if not file_path:
            return

        columns = [tree.heading(col)["text"] for col in tree["columns"]]
        rows = [tree.item(item)["values"] for item in tree.get_children()]

        try:
            with open(file_path, mode="w", newline='', encoding="utf-8") as file:
                writer = csv.writer(file, delimiter=';')
                writer.writerow(columns)
                writer.writerows(rows)

            messagebox.showinfo("Succès", "Les données ont été exportées avec succès en CSV.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'exportation en CSV : {str(e)}")

    def recherche_croisee_indicateur(self):
        recherche_window = tk.Toplevel(self.root)
        recherche_window.title("Recherche Croisée Indicateur")
        recherche_window.geometry("920x680")
        recherche_window.configure(bg="#f0f5f9")
        recherche_window.resizable(True, True)

        # En-tête
        header_frame = tk.Frame(recherche_window, bg="#2c3e50", height=58)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        tk.Label(header_frame, text="📈  Recherche Croisée Indicateur",
                 font=("Helvetica", 13, "bold"), fg="white", bg="#2c3e50").pack(side=tk.LEFT, padx=20, pady=15)

        # Zone scrollable pour les filtres
        container = tk.Frame(recherche_window, bg="#f0f5f9")
        container.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(container, bg="#f0f5f9", highlightthickness=0)
        v_scroll = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scroll_frame = ttk.Frame(canvas)

        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=v_scroll.set)

        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        numeric_columns = self.get_numeric_columns()[:12]
        non_numeric_columns = self.get_non_numeric_columns()
        self.numeric_entries = {}
        self.non_numeric_comboboxes = {}

        # --- Section filtres numériques ---
        if numeric_columns:
            num_section = ttk.LabelFrame(scroll_frame, text="  Filtres numériques (plages de valeurs)  ", padding=(15, 10))
            num_section.pack(fill=tk.X, padx=20, pady=(15, 8))
            COLS = 3
            for i in range(COLS):
                num_section.grid_columnconfigure(i, weight=1)
            for idx, column in enumerate(numeric_columns):
                r, c = divmod(idx, COLS)
                cell = ttk.Frame(num_section)
                cell.grid(row=r, column=c, padx=10, pady=8, sticky="ew")
                ttk.Label(cell, text=column, font=("Helvetica", 10, "bold")).pack(anchor="w")
                range_row = ttk.Frame(cell)
                range_row.pack(fill=tk.X, pady=(4, 0))
                ttk.Label(range_row, text="Min", width=4, foreground="#555555").pack(side=tk.LEFT)
                min_entry = ttk.Entry(range_row, width=9)
                min_entry.pack(side=tk.LEFT, padx=(3, 10))
                self.numeric_entries[f"{column}_min"] = min_entry
                ttk.Label(range_row, text="Max", width=4, foreground="#555555").pack(side=tk.LEFT)
                max_entry = ttk.Entry(range_row, width=9)
                max_entry.pack(side=tk.LEFT, padx=(3, 0))
                self.numeric_entries[f"{column}_max"] = max_entry

        # --- Section filtres textuels ---
        if non_numeric_columns:
            txt_section = ttk.LabelFrame(scroll_frame, text="  Filtres textuels  ", padding=(15, 10))
            txt_section.pack(fill=tk.X, padx=20, pady=(8, 15))
            COLS = 3
            for i in range(COLS):
                txt_section.grid_columnconfigure(i, weight=1)
            for idx, column in enumerate(non_numeric_columns):
                r, c = divmod(idx, COLS)
                cell = ttk.Frame(txt_section)
                cell.grid(row=r, column=c, padx=10, pady=8, sticky="ew")
                ttk.Label(cell, text=column, font=("Helvetica", 10, "bold")).pack(anchor="w")
                values = self.get_unique_values(column)
                combo = ttk.Combobox(cell, values=[""] + [str(v) for v in values], width=24)
                combo.pack(fill=tk.X, pady=(4, 0))
                self.non_numeric_comboboxes[column] = combo

        # --- Barre d'actions fixe en bas ---
        action_bar = tk.Frame(recherche_window, bg="#ecf0f1")
        action_bar.pack(fill=tk.X, side=tk.BOTTOM)
        ttk.Separator(action_bar, orient="horizontal").pack(fill=tk.X)
        btn_row = tk.Frame(action_bar, bg="#ecf0f1")
        btn_row.pack(pady=12, padx=20)

        def reset_filters():
            for entry in self.numeric_entries.values():
                entry.delete(0, tk.END)
            for combo in self.non_numeric_comboboxes.values():
                combo.set("")

        ttk.Button(btn_row, text="Réinitialiser", command=reset_filters, width=14).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_row, text="🔍  Rechercher", command=self.execute_indicator_search,
                   style="Accent.TButton", width=16).pack(side=tk.LEFT)

        recherche_window.protocol("WM_DELETE_WINDOW",
                                  lambda: [canvas.unbind_all("<MouseWheel>"), recherche_window.destroy()])

    def get_numeric_columns(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(indicateurs)")
        all_columns = [col[1] for col in cursor.fetchall()]
        exclude_columns = ['siret', 'produits', 'charge', 'code_naf', 'nom', 'nomUniteLegale', 'prenom1UniteLegale']
        numeric_columns = [col for col in all_columns if col not in exclude_columns and self.is_numeric_column(col)]
        conn.close()
        return numeric_columns

    def is_numeric_column(self, column):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute(f"SELECT {column} FROM indicateurs LIMIT 10")
        rows = cursor.fetchall()
        conn.close()
        try:
            for row in rows:
                if row[0] is not None:
                    float(row[0])
            return True
        except (ValueError, TypeError):
            return False

    def get_non_numeric_columns(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(indicateurs)")
        all_columns = [col[1] for col in cursor.fetchall()]
        exclude_columns = ['siret', 'produits', 'charge', 'code_naf', 'nom', 'nomUniteLegale', 'prenom1UniteLegale']
        numeric_columns = self.get_numeric_columns()
        non_numeric_columns = [col for col in all_columns if col not in exclude_columns and col not in numeric_columns]
        conn.close()
        return non_numeric_columns

    def get_unique_values(self, column):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute(f"SELECT DISTINCT {column} FROM indicateurs ORDER BY {column}")
        values = [row[0] for row in cursor.fetchall()]
        conn.close()
        return values

    def execute_indicator_search(self):
        query = "SELECT * FROM indicateurs WHERE 1=1"
        params = []

        for column in self.numeric_entries:
            if "min" in column and self.numeric_entries[column].get():
                query += f" AND {column[:-4]} >= ?"
                params.append(float(self.numeric_entries[column].get()))
            elif "max" in column and self.numeric_entries[column].get():
                query += f" AND {column[:-4]} <= ?"
                params.append(float(self.numeric_entries[column].get()))

        for column, combobox in self.non_numeric_comboboxes.items():
            if combobox.get():
                query += f" AND {column} = ?"
                params.append(combobox.get())

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute(query, params)
        results = cursor.fetchall()
        conn.close()

        self.display_indicator_results(results)

    def display_indicator_results(self, results):
        if not results:
            messagebox.showinfo("Aucun résultat", "Aucun résultat trouvé pour ces critères.")
            return

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(indicateurs)")
        columns = [info[1] for info in cursor.fetchall()]
        conn.close()

        siret_index = columns.index("siret")
        filtered_results = [row for row in results if "moyenne" not in str(row[siret_index]).strip().lower()]

        results_window = tk.Toplevel(self.root)
        results_window.title("Résultats — Recherche Croisée Indicateur")
        w = min(self.root.winfo_screenwidth(), 1400)
        h = min(self.root.winfo_screenheight(), 820)
        results_window.geometry(f"{w}x{h}")
        results_window.configure(bg="#f0f5f9")

        # En-tête avec compteur
        header_frame = tk.Frame(results_window, bg="#27ae60", height=55)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        tk.Label(header_frame,
                 text=f"✅  {len(filtered_results)} résultat(s) trouvé(s)",
                 font=("Helvetica", 13, "bold"), fg="white", bg="#27ae60").pack(side=tk.LEFT, padx=20, pady=14)

        # Tableau
        result_frame = ttk.Frame(results_window, padding=(10, 10, 10, 0))
        result_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar_y = ttk.Scrollbar(result_frame, orient="vertical")
        scrollbar_x = ttk.Scrollbar(result_frame, orient="horizontal")

        tree = ttk.Treeview(result_frame, columns=columns, show="headings",
                            yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        for col in columns:
            tree.heading(col, text=col, anchor="center")
            tree.column(col, width=120, minwidth=80, anchor="center")

        scrollbar_y.config(command=tree.yview)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.config(command=tree.xview)
        scrollbar_x.pack(side="bottom", fill="x")
        tree.pack(fill="both", expand=True)

        tree.tag_configure("even", background="#ffffff")
        tree.tag_configure("odd", background="#eaf2fb")
        for i, row in enumerate(filtered_results):
            tree.insert("", "end", values=row, tags=("odd" if i % 2 else "even",))

        # Barre d'actions bas
        bottom_bar = tk.Frame(results_window, bg="#ecf0f1")
        bottom_bar.pack(fill=tk.X, side=tk.BOTTOM)
        ttk.Separator(bottom_bar, orient="horizontal").pack(fill=tk.X)
        btn_row = tk.Frame(bottom_bar, bg="#ecf0f1")
        btn_row.pack(pady=10, padx=20)
        ttk.Button(btn_row, text="📥  Exporter en XLSX",
                   command=lambda: self.export_to_xlsx(tree, columns, filtered_results),
                   style="Accent.TButton", width=20).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_row, text="Fermer", command=results_window.destroy, width=12).pack(side=tk.LEFT)

    def export_to_xlsx(self, tree, columns, filtered_results):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

        if not file_path:
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Résultats de la Recherche"

        for col_idx, col_name in enumerate(columns, 1):
            sheet.cell(row=1, column=col_idx).value = col_name

        for row_idx, row_data in enumerate(filtered_results, 2):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx).value = value

        numeric_columns = [col for col in columns if col != 'siret']
        numeric_column_indexes = [columns.index(col) + 1 for col in numeric_columns]

        averages = {}
        row_count = len(filtered_results)
        for col_idx in numeric_column_indexes:
            values = [sheet.cell(row=row_idx, column=col_idx).value for row_idx in range(2, row_count + 2)]
            numeric_values = [value for value in values if isinstance(value, (int, float))]
            if numeric_values:
                average = sum(numeric_values) / len(numeric_values)
                averages[col_idx] = average
                sheet.cell(row=row_count + 2, column=col_idx).value = average

        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        for row_idx in range(2, row_count + 2):
            for col_idx in numeric_column_indexes:
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if isinstance(cell_value, (int, float)) and col_idx in averages:
                    average = averages[col_idx]
                    if cell_value > 1.3 * average:
                        sheet.cell(row=row_idx, column=col_idx).fill = red_fill
                    elif cell_value < 0.7 * average:
                        sheet.cell(row=row_idx, column=col_idx).fill = green_fill

        try:
            workbook.save(file_path)
            messagebox.showinfo("Succès", "Les données ont été exportées avec succès en XLSX.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'exportation en XLSX : {str(e)}")

    def setup_signature(self):
        signature_frame = tk.Frame(self.root, bg='#f5f6fa')
        signature_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=5)

        signature_label = tk.Label(signature_frame, text="Expert comptable Benjamin Hofman", bg='#f5f6fa')
        signature_label.pack(side=tk.LEFT, padx=5)

        linkedin_button = tk.Button(signature_frame, text="[in]", command=self.open_linkedin, bg='#0077B5', fg='white')
        linkedin_button.pack(side=tk.LEFT)

    def on_frame_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas_frame, width=event.width-4)

    def open_linkedin(self):
        webbrowser.open("https://www.linkedin.com/in/benjamin-hofman-22b1b5296/")

    def run(self):
        self.root.mainloop()

    def show_client_map(self):
        """Affiche une carte avec la localisation des clients"""
        try:
            # Vérifier si les bibliothèques nécessaires sont installées
            try:
                import folium
                import requests
                from geopy.geocoders import Nominatim
            except ImportError:
                response = messagebox.askyesno(
                    "Bibliothèques manquantes", 
                    "Certaines bibliothèques sont nécessaires pour générer la carte. Voulez-vous les installer maintenant ?\n\n"
                    "Bibliothèques nécessaires: folium, requests, geopy"
                )
                if response:
                    import subprocess
                    subprocess.call([sys.executable, "-m", "pip", "install", "folium", "requests", "geopy"])
                    messagebox.showinfo("Installation terminée", "Les bibliothèques ont été installées. La carte va être générée.")
                    # Importer après installation
                    import folium
                    import requests
                    from geopy.geocoders import Nominatim
                else:
                    return
            
            # Connexion à la base de données
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Vérifier si la table contient des informations sur la ville
            cursor.execute("PRAGMA table_info(indicateurs)")
            columns = [info[1] for info in cursor.fetchall()]
            
            ville_column = None
            for possible_col in ['ville', 'Ville', 'city', 'City', 'adresseEtablissement']:
                if possible_col in columns:
                    ville_column = possible_col
                    break
            
            if not ville_column:
                # Si la colonne de ville n'existe pas, proposer d'en ajouter une
                response = messagebox.askyesno(
                    "Information manquante", 
                    "Aucune information de ville n'a été trouvée dans la base de données. "
                    "Voulez-vous ajouter une colonne 'ville' et importer des données depuis un fichier CSV?"
                )
                
                if response:
                    # Ajouter la colonne ville
                    cursor.execute("ALTER TABLE indicateurs ADD COLUMN ville TEXT")
                    conn.commit()
                    
                    # Demander à l'utilisateur de choisir un fichier CSV avec les villes
                    file_path = filedialog.askopenfilename(
                        filetypes=[("CSV files", "*.csv")],
                        title="Choisissez un fichier CSV contenant les villes"
                    )
                    
                    if file_path:
                        # Importer les données du CSV
                        import pandas as pd
                        try:
                            df = pd.read_csv(file_path)
                            if 'siret' in df.columns and 'ville' in df.columns:
                                for _, row in df.iterrows():
                                    cursor.execute(
                                        "UPDATE indicateurs SET ville = ? WHERE siret = ?",
                                        (row['ville'], row['siret'])
                                    )
                                conn.commit()
                                ville_column = 'ville'
                            else:
                                messagebox.showerror(
                                    "Format incorrect", 
                                    "Le fichier CSV doit contenir les colonnes 'siret' et 'ville'."
                                )
                                conn.close()
                                return
                        except Exception as e:
                            messagebox.showerror("Erreur", f"Erreur lors de l'importation du CSV: {str(e)}")
                            conn.close()
                            return
                    else:
                        conn.close()
                        return
                else:
                    conn.close()
                    return
            
            # Récupérer les informations des clients avec leur ville
            cursor.execute(f"""
                SELECT 
                    COALESCE(nom, siret) as nom,
                    siret,
                    {ville_column},
                    code_naf,
                    forme_juri,
                    CA
                FROM indicateurs 
                WHERE siret != 'moyenne' AND {ville_column} IS NOT NULL AND {ville_column} != ''
            """)
            clients = cursor.fetchall()
            conn.close()
            
            if not clients:
                messagebox.showinfo(
                    "Information", 
                    "Aucune donnée de ville n'est disponible pour les clients."
                )
                return
            
            # Initialiser le géocodeur
            geolocator = Nominatim(user_agent="fec_explorer")
            
            # Créer une carte centrée sur la France
            m = folium.Map(location=[46.603354, 1.888334], zoom_start=6)
            
            # Compter le nombre de clients par ville
            ville_client_count = {}
            for client in clients:
                _, _, ville, _, _, _ = client
                ville_client_count[ville] = ville_client_count.get(ville, 0) + 1
            
            # Ajouter les clients à la carte
            processed_clients = 0
            client_locations = {}  # Pour éviter de géocoder plusieurs fois la même ville
            ville_clients = {}  # Stocker les clients par ville
            
            # Créer une fenêtre de progression
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Génération de la carte")
            progress_window.geometry("400x150")
            progress_window.configure(bg="#f5f6fa")
            
            progress_label = ttk.Label(
                progress_window, 
                text=f"Géocodage des clients... (0/{len(clients)})",
                font=("Helvetica", 10)
            )
            progress_label.pack(pady=10)
            
            progress_bar = ttk.Progressbar(
                progress_window, 
                orient="horizontal", 
                length=350, 
                mode="determinate"
            )
            progress_bar.pack(pady=10)
            
            progress_window.update()
            
            # Méthode pour afficher les clients d'une ville
            def show_city_clients(ville):
                """Affiche les clients d'une ville triés par CA"""
                if ville not in ville_clients:
                    messagebox.showinfo("Information", f"Aucun client trouvé pour {ville}")
                    return
                
                # Trier les clients par chiffre d'affaires décroissant
                city_clients = sorted(
                    ville_clients[ville], 
                    key=lambda x: float(x['ca']) if x['ca'] is not None and x['ca'] != '' else 0, 
                    reverse=True
                )
                
                # Créer la fenêtre des clients
                clients_window = tk.Toplevel(self.root)
                clients_window.title(f"Clients de {ville}")
                clients_window.geometry("800x600")
                
                # Créer un cadre avec scrollbar
                frame = tk.Frame(clients_window)
                frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
                
                # Créer un canvas avec scrollbar
                canvas = tk.Canvas(frame)
                scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
                scrollable_frame = tk.Frame(canvas)
                
                scrollable_frame.bind(
                    "<Configure>",
                    lambda e: canvas.configure(
                        scrollregion=canvas.bbox("all")
                    )
                )
                
                canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
                canvas.configure(yscrollcommand=scrollbar.set)
                
                canvas.pack(side="left", fill="both", expand=True)
                scrollbar.pack(side="right", fill="y")
                
                # Titre
                titre_label = tk.Label(
                    scrollable_frame, 
                    text=f"Clients de {ville}", 
                    font=("Helvetica", 16, "bold")
                )
                titre_label.pack(pady=10)
                
                # En-têtes
                headers = ["Nom", "SIRET", "Code NAF", "Forme Juridique", "Chiffre d'Affaires"]
                header_frame = tk.Frame(scrollable_frame)
                header_frame.pack(fill='x', padx=10)
                
                for i, header in enumerate(headers):
                    label = tk.Label(
                        header_frame, 
                        text=header, 
                        font=("Helvetica", 12, "bold")
                    )
                    label.grid(row=0, column=i, padx=5, pady=5, sticky='w')
                
                # Afficher les clients
                for i, client in enumerate(city_clients, 1):
                    # Cadre pour chaque ligne de client
                    ligne_client = tk.Frame(scrollable_frame)
                    ligne_client.pack(fill='x', padx=10)
                    
                    # Formater le CA
                    try:
                        ca_formate = f"{float(client['ca']):,.2f} €" if client['ca'] is not None else 'N/A'
                    except (ValueError, TypeError):
                        ca_formate = 'N/A'
                    
                    # Données à afficher
                    donnees = [
                        client['nom'], 
                        client['siret'], 
                        client['code_naf'], 
                        client['forme_juri'], 
                        ca_formate
                    ]
                    
                    # Créer les labels pour chaque donnée
                    for j, donnee in enumerate(donnees):
                        label = tk.Label(
                            ligne_client, 
                            text=str(donnee), 
                            font=("Helvetica", 10)
                        )
                        label.grid(row=0, column=j, padx=5, pady=5, sticky='w')
            
            # Fonction pour ouvrir la sélection de ville
            def open_city_selection():
                # Créer une fenêtre de sélection de ville
                city_window = tk.Toplevel(self.root)
                city_window.title("Sélection de ville")
                city_window.geometry("300x500")
                
                # Créer un label
                label = tk.Label(city_window, text="Sélectionnez une ville", font=("Helvetica", 14))
                label.pack(pady=10)
                
                # Créer un cadre avec scrollbar
                frame = tk.Frame(city_window)
                frame.pack(fill=tk.BOTH, expand=True)
                
                canvas = tk.Canvas(frame)
                scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
                scrollable_frame = tk.Frame(canvas)
                
                scrollable_frame.bind(
                    "<Configure>",
                    lambda e: canvas.configure(
                        scrollregion=canvas.bbox("all")
                    )
                )
                
                canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
                canvas.configure(yscrollcommand=scrollbar.set)
                
                canvas.pack(side="left", fill="both", expand=True)
                scrollbar.pack(side="right", fill="y")
                
                # Créer un bouton pour chaque ville
                for ville, count in sorted(ville_client_count.items(), key=lambda x: x[1], reverse=True):
                    btn = tk.Button(
                        scrollable_frame, 
                        text=f"{ville} ({count} clients)", 
                        command=lambda v=ville: show_city_clients(v)
                    )
                    btn.pack(fill='x', padx=10, pady=5)
            
            # Ajouter des marqueurs pour chaque ville
            for client in clients:
                nom, siret, ville, code_naf, forme_juri, ca = client
                
                # Mise à jour de la barre de progression
                processed_clients += 1
                progress = (processed_clients / len(clients)) * 100
                progress_bar["value"] = progress
                progress_label.config(text=f"Géocodage des clients... ({processed_clients}/{len(clients)})")
                progress_window.update()
                
                # Si la ville a déjà été géocodée, utiliser les coordonnées existantes
                if ville in client_locations:
                    lat, lon = client_locations[ville]
                else:
                    try:
                        # Géocoder la ville
                        location = geolocator.geocode(f"{ville}, France", timeout=10)
                        if location:
                            lat, lon = location.latitude, location.longitude
                            client_locations[ville] = (lat, lon)
                        else:
                            continue
                    except Exception as e:
                        continue
                
                # Déterminer la taille du marqueur en fonction du nombre de clients dans la ville
                radius = min(20, max(5, ville_client_count[ville] * 2))
                
                # Stocker les clients par ville
                if ville not in ville_clients:
                    ville_clients[ville] = []
                ville_clients[ville].append({
                    'nom': nom,
                    'siret': siret,
                    'ville': ville,
                    'code_naf': code_naf,
                    'forme_juri': forme_juri,
                    'ca': ca,
                    'lat': lat,
                    'lon': lon
                })
                
                # Créer le texte de l'info-bulle
                popup_text = f"""
                Ville: {ville}<br>
                Nombre de clients: {ville_client_count[ville]}
                """

                # Ajouter le marqueur à la carte
                marker = folium.CircleMarker(
                    location=[lat, lon],
                    radius=radius,
                    popup=folium.Popup(popup_text, max_width=300),
                    tooltip=f"{ville} ({ville_client_count[ville]} clients)",
                    color='blue',
                    fill=True,
                    fill_color='blue',
                    fill_opacity=0.7
                ).add_to(m)
            
            # Fermer la fenêtre de progression
            progress_window.destroy()
            
            # Sauvegarder la carte dans un fichier HTML temporaire
            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix='.html') as f:
                map_path = f.name
            m.save(map_path)
            
            # Ouvrir la carte dans le navigateur par défaut
            webbrowser.open('file://' + map_path)
            
            # Ajouter un bouton pour sélectionner une ville
            select_city_button = tk.Button(
                text="Sélectionner une ville", 
                command=open_city_selection
            )
            select_city_button.pack()
            
            # Proposer d'enregistrer la carte
            response = messagebox.askyesno(
                "Carte générée", 
                "La carte a été générée et ouverte dans votre navigateur. Voulez-vous l'enregistrer?"
            )
            
            if response:
                save_path = filedialog.asksaveasfilename(
                    defaultextension=".html",
                    filetypes=[("HTML files", "*.html")],
                    title="Enregistrer la carte"
                )
                
                if save_path:
                    import shutil
                    shutil.copy2(map_path, save_path)
                    messagebox.showinfo(
                        "Succès", 
                        f"La carte a été enregistrée à l'emplacement: {save_path}"
                    )
            
        except Exception as e:
                messagebox.showerror("Erreur", f"Une erreur est survenue lors de la génération de la carte: {str(e)}")
                import traceback
                traceback.print_exc()



class GradientFrame(tk.Canvas):
    def __init__(self, parent, color1="#1E88E5", color2="#1DE9B6", **kwargs):
        tk.Canvas.__init__(self, parent, **kwargs)
        self._color1 = color1
        self._color2 = color2
        self.bind("<Configure>", self._draw_gradient)

    def _draw_gradient(self, event=None):
        self.delete("gradient")
        width = self.winfo_width()
        height = self.winfo_height()
        
        for i in range(width):
            r1, g1, b1 = self.winfo_rgb(self._color1)
            r2, g2, b2 = self.winfo_rgb(self._color2)
            
            r = (r1 + int((r2-r1) * i/width)) >> 8
            g = (g1 + int((g2-g1) * i/width)) >> 8
            b = (b1 + int((b2-b1) * i/width)) >> 8
            
            color = f'#{r:02x}{g:02x}{b:02x}'
            self.create_line(i, 0, i, height, tags=("gradient",), fill=color)
        
        self.tag_lower("gradient")

class CustomButton(tk.Button):
    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)
        self.configure(
            background="#3498db",
            foreground="white",
            activebackground="#2980b9",
            activeforeground="white",
            relief="flat",
            borderwidth=0,
            padx=10,
            pady=8,
            anchor="w",
            font=('Arial', 10),
            cursor="hand2"
        )
        self.bind("<Enter>", self.on_enter)
        self.bind("<Leave>", self.on_leave)

    def on_enter(self, e):
        self.configure(background="#2980b9")

    def on_leave(self, e):
        self.configure(background="#3498db")

if __name__ == "__main__":
    app = FECExplorer()
    app.run()